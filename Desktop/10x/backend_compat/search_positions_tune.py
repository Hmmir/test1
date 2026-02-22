import argparse
import itertools
import json
import math
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple


def _to_int(value: Any) -> int:
    try:
        if isinstance(value, str) and value.strip() in {"-", "—"}:
            return 0
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_date_ymd(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value or "").strip()
    if len(text) >= 10:
        return text[:10]
    return ""


def _median_int(values: List[int]) -> int:
    if not values:
        return 0
    items = sorted(int(v) for v in values)
    return int(items[len(items) // 2])


def _load_targets_from_xlsx(
    xlsx_path: Path,
    sheet_name: str,
    reduce_mode: str,
) -> Tuple[Dict[Tuple[int, str], Tuple[int, int]], Dict[Tuple[int, str], Dict[str, Any]]]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"sheet '{sheet_name}' not found in {xlsx_path}")
    ws = wb[sheet_name]

    rows: List[Tuple[str, int, str, int, int]] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        # Row 1: technical header, Row 2: localized header in the template.
        if idx < 2:
            continue
        if not row or len(row) < 3:
            continue
        day = _to_date_ymd(row[0])
        nm_id = _to_int(row[1])
        key = str(row[2] or "").strip()
        if not day or not nm_id or not key:
            continue
        pos = _to_int(row[3] if len(row) > 3 else 0)
        promo = _to_int(row[4] if len(row) > 4 else 0)
        rows.append((day, nm_id, key, pos, promo))

    by_pair: Dict[Tuple[int, str], List[Tuple[str, int, int]]] = {}
    for day, nm_id, key, pos, promo in rows:
        by_pair.setdefault((nm_id, key), []).append((day, pos, promo))

    mode = str(reduce_mode or "").strip().lower() or "last"
    targets: Dict[Tuple[int, str], Tuple[int, int]] = {}
    meta: Dict[Tuple[int, str], Dict[str, Any]] = {}

    for pair, obs in by_pair.items():
        obs_sorted = sorted(obs, key=lambda x: x[0])
        if mode in {"last", "latest"}:
            day, pos, promo = obs_sorted[-1]
        elif mode in {"first", "earliest"}:
            day, pos, promo = obs_sorted[0]
        elif mode in {"median"}:
            day = obs_sorted[-1][0]
            pos = _median_int([x[1] for x in obs_sorted])
            promo = _median_int([x[2] for x in obs_sorted])
        elif mode in {"median_nonzero", "median_nz"}:
            day = obs_sorted[-1][0]
            pos_vals = [x[1] for x in obs_sorted if int(x[1] or 0) > 0]
            promo_vals = [x[2] for x in obs_sorted if int(x[2] or 0) > 0]
            pos = _median_int(pos_vals) if pos_vals else 0
            promo = _median_int(promo_vals) if promo_vals else 0
        else:
            raise RuntimeError(f"unknown reduce mode: {reduce_mode}")

        targets[pair] = (int(pos or 0), int(promo or 0))
        meta[pair] = {
            "observations": len(obs_sorted),
            "date_min": obs_sorted[0][0],
            "date_max": obs_sorted[-1][0],
            "target_date": day,
        }

    return targets, meta


def _build_items(targets: Dict[Tuple[int, str], Tuple[int, int]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for (nm_id, key) in sorted(targets.keys(), key=lambda x: (x[0], x[1])):
        out.append({"itemNumber": int(nm_id), "query": str(key)})
    return out


def _evaluate(
    predicted: Dict[Tuple[int, str], Tuple[int, int]],
    targets: Dict[Tuple[int, str], Tuple[int, int]],
) -> Dict[str, Any]:
    total = len(targets)
    if total <= 0:
        return {
            "pairs": 0,
            "pos_exact_ratio": 0.0,
            "promo_exact_ratio": 0.0,
            "both_exact_ratio": 0.0,
            "pos_mae": 0.0,
        }

    pos_exact = 0
    promo_exact = 0
    both_exact = 0
    pos_abs_sum = 0.0
    missing_pred = 0

    for pair, (tpos, tpromo) in targets.items():
        pred = predicted.get(pair)
        if pred is None:
            missing_pred += 1
            pred_pos, pred_promo = 0, 0
        else:
            pred_pos, pred_promo = pred
        if int(pred_pos) == int(tpos):
            pos_exact += 1
        if int(pred_promo) == int(tpromo):
            promo_exact += 1
        if int(pred_pos) == int(tpos) and int(pred_promo) == int(tpromo):
            both_exact += 1
        pos_abs_sum += abs(float(int(pred_pos) - int(tpos)))

    return {
        "pairs": total,
        "missing_pred": missing_pred,
        "pos_exact_ratio": round(pos_exact / float(total), 6),
        "promo_exact_ratio": round(promo_exact / float(total), 6),
        "both_exact_ratio": round(both_exact / float(total), 6),
        "pos_mae": round(pos_abs_sum / float(total), 6),
    }


def _powerset_sizes(items: List[str], min_k: int, max_k: int) -> Iterable[Tuple[str, ...]]:
    n = len(items)
    lo = max(0, int(min_k))
    hi = min(n, int(max_k))
    for k in range(lo, hi + 1):
        for combo in itertools.combinations(items, k):
            yield combo


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="clone_export_latest.xlsx", help="Path to exported XLSX")
    ap.add_argument("--sheet", default="search_positions", help="Sheet name with positions")
    ap.add_argument(
        "--reduce",
        default="last",
        help="How to reduce multiple dates per (nm_id,search_key): last|first|median|median_nonzero",
    )
    ap.add_argument(
        "--mode",
        default="multi",
        help="Request mode: multi (per-dest requests + agg) | single (dest is comma-joined subset)",
    )
    ap.add_argument(
        "--dests",
        default="",
        help="Candidate dest codes (comma-separated). Example: -1257786,-1250619",
    )
    ap.add_argument("--min-dests", type=int, default=2)
    ap.add_argument("--max-dests", type=int, default=5)
    ap.add_argument(
        "--agg-modes",
        default="first,first_nonzero,best,min,median",
        help="Aggregation modes for multi mode",
    )
    ap.add_argument("--max-pages", type=int, default=10)
    ap.add_argument(
        "--max-configs",
        type=int,
        default=5000,
        help="Safety cap for number of configs evaluated (mainly for multi mode permutations)",
    )
    ap.add_argument("--output", default="", help="Output JSON path")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise RuntimeError(f"xlsx not found: {xlsx_path}")

    targets, targets_meta = _load_targets_from_xlsx(
        xlsx_path=xlsx_path,
        sheet_name=str(args.sheet),
        reduce_mode=str(args.reduce),
    )
    items = _build_items(targets)

    dest_candidates = [d.strip() for d in str(args.dests or "").split(",") if d.strip()]
    if not dest_candidates:
        # Default pool based on common WB "dest" codes seen in the wild (RU).
        dest_candidates = [
            "-1257786",  # Moscow (commonly used)
            "-1029256",
            "-102269",
            "-1282181",
            "-456807",
        ]

    mode = str(args.mode or "").strip().lower() or "multi"
    max_pages = max(1, int(args.max_pages))

    report: Dict[str, Any] = {
        "xlsx": str(xlsx_path),
        "sheet": str(args.sheet),
        "reduce": str(args.reduce),
        "mode": mode,
        "pairs": len(targets),
        "dest_candidates": dest_candidates,
        "max_pages": max_pages,
        "targets_meta": {
            f"{nm_id}|{key}": meta for (nm_id, key), meta in targets_meta.items()
        },
    }

    # Lazy import: wb_client uses repo-style imports.
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from wb_client import fetch_search_positions_multi  # type: ignore
    from wb_client import _aggregate_positions  # type: ignore

    if mode == "multi":
        agg_modes = [m.strip() for m in str(args.agg_modes).split(",") if m.strip()]
        if not agg_modes:
            agg_modes = ["first", "first_nonzero", "best", "min", "median"]

        # Fetch raw per-dest positions once for the full candidate list.
        _rows2d, raw_rows = fetch_search_positions_multi(
            items, max_pages=max_pages, dests=dest_candidates, agg_mode="first"
        )
        per_item: Dict[Tuple[int, str], Dict[str, Tuple[int, int]]] = {}
        for row in raw_rows:
            try:
                nm_id = _to_int(row.get("nm_id"))
                key = str(row.get("search_key") or "").strip()
                dest = str(row.get("dest") or "").strip()
                pos = _to_int(row.get("position"))
                promo = _to_int(row.get("promo_position"))
            except Exception:
                continue
            if not nm_id or not key or not dest:
                continue
            per_item.setdefault((nm_id, key), {})[dest] = (pos, promo)

        # Evaluate configs: permutations (order matters for first/first_nonzero).
        configs: List[Dict[str, Any]] = []
        evaluated = 0

        for subset in _powerset_sizes(dest_candidates, int(args.min_dests), int(args.max_dests)):
            for perm in itertools.permutations(subset, len(subset)):
                for agg in agg_modes:
                    evaluated += 1
                    if evaluated > int(args.max_configs):
                        break

                    predicted: Dict[Tuple[int, str], Tuple[int, int]] = {}
                    for pair in targets.keys():
                        per_dest = per_item.get(pair, {})
                        # Build only the dests used by this config.
                        per_dest_conf = {d: per_dest.get(d, (0, 0)) for d in perm}
                        pos, promo = _aggregate_positions(per_dest_conf, list(perm), agg)
                        predicted[pair] = (int(pos or 0), int(promo or 0))

                    metrics = _evaluate(predicted, targets)
                    configs.append(
                        {
                            "dests": list(perm),
                            "agg_mode": str(agg),
                            **metrics,
                        }
                    )
                if evaluated > int(args.max_configs):
                    break
            if evaluated > int(args.max_configs):
                break

        configs.sort(
            key=lambda x: (
                float(x.get("both_exact_ratio") or 0.0),
                float(x.get("pos_exact_ratio") or 0.0),
                -float(x.get("pos_mae") or 0.0),
                -len(x.get("dests") or []),
            ),
            reverse=True,
        )
        report["configs_evaluated"] = evaluated
        report["top"] = configs[:50]
        report["best"] = configs[0] if configs else None

        if configs:
            best = configs[0]
            report["recommended_env"] = {
                "BTLZ_WB_SEARCH_DESTS": ",".join(best.get("dests") or []),
                "BTLZ_WB_SEARCH_POS_AGG": str(best.get("agg_mode") or ""),
            }

    elif mode == "single":
        # Evaluate dest subsets by making a single request per subset, using dest=csv.
        configs: List[Dict[str, Any]] = []
        evaluated = 0

        for subset in _powerset_sizes(dest_candidates, int(args.min_dests), int(args.max_dests)):
            evaluated += 1
            if evaluated > int(args.max_configs):
                break
            dest_csv = ",".join(subset)
            rows2d, _raw = fetch_search_positions_multi(
                items, max_pages=max_pages, dests=[dest_csv], agg_mode="first"
            )
            predicted: Dict[Tuple[int, str], Tuple[int, int]] = {}
            for row in rows2d:
                if not row or len(row) < 5:
                    continue
                nm_id = _to_int(row[1])
                key = str(row[2] or "").strip()
                pos = _to_int(row[3])
                promo = _to_int(row[4])
                if nm_id and key:
                    predicted[(nm_id, key)] = (pos, promo)

            metrics = _evaluate(predicted, targets)
            configs.append(
                {
                    "dest": dest_csv,
                    "dests": list(subset),
                    **metrics,
                }
            )

        configs.sort(
            key=lambda x: (
                float(x.get("both_exact_ratio") or 0.0),
                float(x.get("pos_exact_ratio") or 0.0),
                -float(x.get("pos_mae") or 0.0),
                -len(x.get("dests") or []),
            ),
            reverse=True,
        )
        report["configs_evaluated"] = evaluated
        report["top"] = configs[:50]
        report["best"] = configs[0] if configs else None
        if configs:
            best = configs[0]
            report["recommended_env"] = {
                "BTLZ_WB_SEARCH_DEST": str(best.get("dest") or ""),
            }
    else:
        raise RuntimeError(f"unknown mode: {args.mode}")

    if not args.output:
        ts = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
        out_path = Path(__file__).resolve().parent / "data" / f"search_positions_tune_report_{ts}.json"
    else:
        out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(str(out_path))


if __name__ == "__main__":
    main()
