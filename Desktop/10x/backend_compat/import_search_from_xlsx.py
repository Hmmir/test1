import argparse
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List

import openpyxl  # type: ignore

from storage import Storage


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_date_ymd(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            pass
    text = str(value).strip()
    if len(text) >= 10:
        return text[:10]
    return text


def _to_pos(value: Any) -> int:
    text = str(value or "").strip()
    if not text or text == "-":
        return 0
    return max(_to_int(text), 0)


def import_search_keys(
    wb: Any, storage: Storage, spreadsheet_id: str, source: str = "xlsx_import"
) -> int:
    if "search_keys" not in wb.sheetnames:
        return 0
    sh = wb["search_keys"]
    items: List[Dict[str, Any]] = []
    # Exported files usually contain labels in row 2; data starts from row 3.
    for row in sh.iter_rows(min_row=3, values_only=True):
        nm_id = _to_int(row[0] if len(row) > 0 else 0)
        search_key = str(row[1] if len(row) > 1 else "").strip()
        if not nm_id or not search_key:
            continue
        item: Dict[str, Any] = {
            "nm_id": nm_id,
            "search_key": search_key,
        }
        key_id = str(row[2] if len(row) > 2 else "").strip()
        if key_id:
            item["search_key_id"] = key_id
        items.append(item)
    if not items:
        return 0
    return storage.upsert_search_keys(spreadsheet_id, items, source=source)


def import_search_positions(
    wb: Any, storage: Storage, spreadsheet_id: str, source: str = "xlsx_import"
) -> int:
    if "search_positions" not in wb.sheetnames:
        return 0
    sh = wb["search_positions"]
    by_day: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for row in sh.iter_rows(min_row=3, values_only=True):
        day = _to_date_ymd(row[0] if len(row) > 0 else "")
        nm_id = _to_int(row[1] if len(row) > 1 else 0)
        search_key = str(row[2] if len(row) > 2 else "").strip()
        if not day or not nm_id or not search_key:
            continue
        by_day[day].append(
            {
                "date": day,
                "captured_at": f"{day} 00:00:00",
                "nm_id": nm_id,
                "search_key": search_key,
                "position": _to_pos(row[3] if len(row) > 3 else 0),
                "promo_position": _to_pos(row[4] if len(row) > 4 else 0),
                "source": source,
            }
        )

    total = 0
    for day, rows in by_day.items():
        total += storage.upsert_daily_positions(spreadsheet_id, day, rows)
    return total


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import search_keys + search_positions from XLSX export into backend DB"
    )
    parser.add_argument("--xlsx", required=True, help="Path to XLSX export")
    parser.add_argument("--spreadsheet-id", required=True, help="Spreadsheet id")
    parser.add_argument(
        "--db-path",
        default=str(Path(__file__).resolve().parent / "data" / "btlz.db"),
        help="Path to sqlite DB",
    )
    parser.add_argument(
        "--source",
        default="xlsx_import",
        help="Source label for imported rows",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    storage = Storage(args.db_path)
    keys_count = import_search_keys(wb, storage, args.spreadsheet_id, source=args.source)
    pos_count = import_search_positions(
        wb, storage, args.spreadsheet_id, source=args.source
    )
    print(
        f"Imported search_keys={keys_count} search_positions={pos_count} "
        f"spreadsheet_id={args.spreadsheet_id}"
    )


if __name__ == "__main__":
    main()
