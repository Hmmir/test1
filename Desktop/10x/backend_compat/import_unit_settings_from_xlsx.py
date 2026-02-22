import argparse
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _ymd(value: Any) -> str:
    if value is None:
        return ""
    try:
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
    except Exception:
        pass
    return str(value).strip()[:10]


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _load_unit_rows(
    xlsx_path: str,
    dt_default: str,
    min_row: int = 2,
    only_nm_ids: Optional[set] = None,
) -> List[Dict[str, Any]]:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required to import xlsx fixtures") from exc

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        raise RuntimeError(f"Failed to load xlsx: {xlsx_path}") from exc

    if "UNIT" not in wb.sheetnames:
        raise RuntimeError('XLSX does not contain sheet "UNIT"')
    sh = wb["UNIT"]

    header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
    headers = [str(x).strip() if x is not None else "" for x in (header or [])]
    col = {name: idx for idx, name in enumerate(headers) if name}

    nm_idx = col.get("nm_id")
    if nm_idx is None:
        raise RuntimeError('UNIT sheet: missing "nm_id" header in row 1')
    dt_idx = col.get("date")

    out: List[Dict[str, Any]] = []
    for row in sh.iter_rows(min_row=max(2, int(min_row)), values_only=True):
        if not row or nm_idx >= len(row):
            continue
        nm_id = _to_int(row[nm_idx])
        if not nm_id:
            continue
        if only_nm_ids and nm_id not in only_nm_ids:
            continue

        dt = ""
        if dt_idx is not None and dt_idx < len(row):
            dt = _ymd(row[dt_idx])
        if not dt:
            dt = dt_default

        item: Dict[str, Any] = {"nm_id": nm_id, "date": dt}
        for key, idx in col.items():
            if key in {"nm_id", "date"}:
                continue
            if idx < 0 or idx >= len(row):
                continue
            val = row[idx]
            if _is_blank(val):
                continue
            item[key] = val
        out.append(item)

    try:
        wb.close()
    except Exception:
        pass
    return out


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to XLSX export containing UNIT sheet")
    ap.add_argument("--spreadsheet-id", required=True, help="Spreadsheet ID to store data under")
    ap.add_argument(
        "--db-path",
        default=str(Path(__file__).resolve().parent / "data" / "btlz.db"),
        help="DB path (default: backend_compat/data/btlz.db)",
    )
    ap.add_argument(
        "--dt-default",
        default="",
        help='Default dt (YYYY-MM-DD) when UNIT rows do not contain "date". Recommended: date_from of your diff window.',
    )
    ap.add_argument(
        "--min-row",
        type=int,
        default=2,
        help="Start reading UNIT rows from this row (default: 2).",
    )
    ap.add_argument(
        "--nm-ids-csv",
        default="",
        help="Optional path to checklist_cross CSV to filter UNIT rows to only referenced nm_id.",
    )
    args = ap.parse_args()

    xlsx_path = str(args.xlsx)
    db_path = str(args.db_path)
    spreadsheet_id = str(args.spreadsheet_id)
    dt_default = str(args.dt_default or "").strip()
    if not dt_default:
        dt_default = datetime.now().strftime("%Y-%m-%d")

    only_nm_ids: Optional[set] = None
    if args.nm_ids_csv:
        import csv

        nm_ids: set = set()
        with open(str(args.nm_ids_csv), "r", encoding="utf-8", newline="") as f:
            r = csv.reader(f)
            next(r, None)
            for row in r:
                if not row or not str(row[0]).strip():
                    continue
                nm_ids.add(_to_int(row[0]))
        only_nm_ids = nm_ids

    rows = _load_unit_rows(
        xlsx_path,
        dt_default=dt_default,
        min_row=int(args.min_row),
        only_nm_ids=only_nm_ids,
    )
    if not rows:
        raise RuntimeError("No UNIT rows found to import")

    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parent))
    from storage import Storage  # type: ignore

    storage = Storage(db_path)
    inserted = storage.upsert_daily_unit_settings(spreadsheet_id, rows, source="xlsx_unit")
    print(inserted)


if __name__ == "__main__":
    main()

