import argparse
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

from storage import Storage


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _ymd(value: Any) -> str:
    if value is None:
        return ""
    try:
        if hasattr(value, "strftime"):
            return value.strftime("%Y-%m-%d")
    except Exception:
        pass
    return str(value)[:10]


def _in_range(day: str, date_from: str, date_to: str) -> bool:
    if not day:
        return False
    if day < date_from:
        return False
    if day > date_to:
        return False
    return True


def _split_nm_ids(text: str) -> Set[int]:
    out: Set[int] = set()
    for part in str(text or "").split(","):
        part = part.strip()
        if not part:
            continue
        val = _to_int(part)
        if val:
            out.add(val)
    return out


def import_daily_stocks_from_xlsx_checklist(
    storage: Storage,
    spreadsheet_id: str,
    xlsx_path: str,
    date_from: str,
    date_to: str,
    nm_filter: Optional[Set[int]] = None,
) -> int:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl is required to import xlsx snapshots") from exc

    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise FileNotFoundError(str(xlsx))

    wb = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    if "checklist" not in wb.sheetnames:
        raise RuntimeError("xlsx does not contain sheet 'checklist'")

    sh = wb["checklist"]
    header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        raise RuntimeError("invalid xlsx: checklist sheet header is missing")
    headers = [str(x).strip() if x is not None else "" for x in header]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = ["date", "nm_id", "stocks", "in_way_to_client", "in_way_from_client"]
    missing = [k for k in required if k not in col]
    if missing:
        raise RuntimeError(f"xlsx checklist is missing required columns: {', '.join(missing)}")

    rows: List[Dict[str, Any]] = []
    for row in sh.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        nm_id = _to_int(row[col["nm_id"]])
        if not nm_id:
            continue
        if nm_filter and nm_id not in nm_filter:
            continue
        day = _ymd(row[col["date"]])
        if not _in_range(day, date_from, date_to):
            continue
        stocks = max(_to_int(row[col["stocks"]]), 0)
        to_client = max(_to_int(row[col["in_way_to_client"]]), 0)
        from_client = max(_to_int(row[col["in_way_from_client"]]), 0)
        rows.append(
            {
                "date": day,
                "nm_id": nm_id,
                "stocks_wb": stocks,
                "in_way_to_client": to_client,
                "in_way_from_client": from_client,
                "stocks_mp": to_client + from_client,
                "source": "xlsx_checklist",
            }
        )

    return storage.upsert_daily_stocks(spreadsheet_id, rows, source="xlsx_checklist")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to XLSX export containing sheet 'checklist'")
    ap.add_argument("--spreadsheet-id", required=True, help="Spreadsheet ID")
    ap.add_argument("--db-path", default="", help="DB path (default backend_compat/data/btlz.db)")
    ap.add_argument("--date-from", default="", help="YYYY-MM-DD (default: min date in xlsx)")
    ap.add_argument("--date-to", default="", help="YYYY-MM-DD (default: today)")
    ap.add_argument("--nm-ids", default="", help="Optional comma-separated nm_ids filter")
    args = ap.parse_args()

    db_path = str(args.db_path or "").strip()
    if not db_path:
        db_path = str(Path(__file__).resolve().parent / "data" / "btlz.db")
    storage = Storage(db_path)

    date_to = str(args.date_to or datetime.utcnow().strftime("%Y-%m-%d"))[:10]
    date_from = str(args.date_from or "0000-01-01")[:10]
    nm_filter = _split_nm_ids(args.nm_ids)
    affected = import_daily_stocks_from_xlsx_checklist(
        storage=storage,
        spreadsheet_id=str(args.spreadsheet_id),
        xlsx_path=str(args.xlsx),
        date_from=date_from,
        date_to=date_to,
        nm_filter=nm_filter if nm_filter else None,
    )
    print(affected)


if __name__ == "__main__":
    main()
