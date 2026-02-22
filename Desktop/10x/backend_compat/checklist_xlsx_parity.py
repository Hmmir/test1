import argparse
import json
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        try:
            v = float(value)
        except Exception:
            return None
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    if isinstance(value, str):
        text = value.strip().replace(" ", "").replace("\u00a0", "").replace(",", ".")
        if not text:
            return None
        try:
            v = float(text)
        except ValueError:
            return None
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    return None


def _norm_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _load_checklist_xlsx_rows(
    xlsx_path: str, sheet_name: str = "checklist"
) -> Tuple[List[str], Dict[Tuple[int, str], List[Any]]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"xlsx does not contain sheet {sheet_name!r}")
    ws = wb[sheet_name]

    headers: List[str] = []
    for col in range(1, int(ws.max_column or 0) + 1):
        raw = ws.cell(row=1, column=col).value
        if raw is None or str(raw).strip() == "":
            break
        headers.append(str(raw).strip())

    if not headers:
        raise RuntimeError("invalid xlsx: checklist sheet header is missing")

    rows: Dict[Tuple[int, str], List[Any]] = {}
    for r in range(3, int(ws.max_row or 0) + 1):
        dt = ws.cell(row=r, column=1).value
        nm_raw = ws.cell(row=r, column=2).value
        if dt is None or nm_raw is None:
            continue
        nm_id = _to_int(nm_raw)
        if not nm_id:
            continue
        day = str(dt)[:10]
        values = [ws.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
        rows[(nm_id, day)] = values

    return headers, rows


def _load_autonomous_rows(path: str) -> Dict[Tuple[int, str], Dict[str, Any]]:
    payload = json.loads(Path(path).read_text(encoding="utf-8"))
    if not isinstance(payload, list):
        raise RuntimeError("autonomous payload must be a JSON list of rows")
    out: Dict[Tuple[int, str], Dict[str, Any]] = {}
    for row in payload:
        if not isinstance(row, dict):
            continue
        nm_id = _to_int(row.get("nm_id"))
        day = str(row.get("date") or "")[:10]
        if nm_id and day:
            out[(nm_id, day)] = row
    return out


def compute_parity(
    xlsx_headers: List[str],
    xlsx_rows: Dict[Tuple[int, str], List[Any]],
    auto_rows: Dict[Tuple[int, str], Dict[str, Any]],
) -> Dict[str, Any]:
    per_field: Dict[str, Dict[str, Any]] = {
        h: {
            "missing": 0,
            "numeric_compared": 0,
            "numeric_exact": 0,
            "numeric_abs_sum": 0.0,
            "string_compared": 0,
            "string_exact": 0,
        }
        for h in xlsx_headers
    }

    numeric_compared = 0
    numeric_exact = 0
    numeric_abs_sum = 0.0
    string_compared = 0
    string_exact = 0
    missing_cells = 0
    nonzero_missing_cells = 0

    for key, base_vals in xlsx_rows.items():
        arow = auto_rows.get(key)
        for idx, h in enumerate(xlsx_headers):
            bv = base_vals[idx] if idx < len(base_vals) else None
            av = arow.get(h) if arow else None

            if isinstance(bv, str) and not bv.strip():
                bv = None
            if isinstance(av, str) and not str(av).strip():
                av = None

            if bv is None and av is None:
                continue

            if av is None:
                missing_cells += 1
                per_field[h]["missing"] += 1
                bnum = _to_float(bv)
                if bnum is not None and abs(bnum) > 1e-9:
                    nonzero_missing_cells += 1
                continue

            bnum = _to_float(bv)
            anum = _to_float(av)
            if bnum is not None and anum is not None:
                numeric_compared += 1
                per_field[h]["numeric_compared"] += 1

                delta = anum - bnum
                abs_delta = abs(delta)
                numeric_abs_sum += abs_delta
                per_field[h]["numeric_abs_sum"] += abs_delta

                tol = max(abs(bnum) * 0.01, 0.01)
                if abs_delta <= tol:
                    numeric_exact += 1
                    per_field[h]["numeric_exact"] += 1
            else:
                string_compared += 1
                per_field[h]["string_compared"] += 1
                if _norm_str(bv) == _norm_str(av):
                    string_exact += 1
                    per_field[h]["string_exact"] += 1

    per_field_rows: List[Dict[str, Any]] = []
    for h, st in per_field.items():
        ncmp = int(st["numeric_compared"])
        sexact = int(st["string_exact"])
        scmp = int(st["string_compared"])
        row = {
            "field": h,
            "missing": int(st["missing"]),
            "numeric_compared": ncmp,
            "numeric_exact_ratio": round(int(st["numeric_exact"]) / ncmp, 6) if ncmp else None,
            "numeric_mae": round(float(st["numeric_abs_sum"]) / ncmp, 6) if ncmp else None,
            "string_compared": scmp,
            "string_exact_ratio": round(sexact / scmp, 6) if scmp else None,
        }
        per_field_rows.append(row)

    per_field_rows.sort(
        key=lambda r: (
            r["numeric_exact_ratio"] if r["numeric_exact_ratio"] is not None else 2.0,
            -(r["numeric_compared"] or 0),
        )
    )

    summary = {
        "base_rows": len(xlsx_rows),
        "auto_rows": len(auto_rows),
        "fields": len(xlsx_headers),
        "numeric_compared": numeric_compared,
        "numeric_exact_ratio": round(numeric_exact / numeric_compared, 6) if numeric_compared else 0.0,
        "numeric_mae": round(numeric_abs_sum / numeric_compared, 6) if numeric_compared else 0.0,
        "string_compared": string_compared,
        "string_exact_ratio": round(string_exact / string_compared, 6) if string_compared else None,
        "missing_cells": missing_cells,
        "nonzero_missing_cells": nonzero_missing_cells,
    }

    return {
        "summary": summary,
        "worst_numeric_fields": per_field_rows[:30],
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to XLSX export containing sheet 'checklist'")
    ap.add_argument(
        "--sheet",
        default="checklist",
        help="Sheet name inside xlsx (default: checklist)",
    )
    ap.add_argument(
        "--auto-json",
        required=True,
        help="JSON file with autonomous checklist rows (list[dict])",
    )
    ap.add_argument("--out", default="", help="Write JSON report to this file")
    args = ap.parse_args()

    headers, base_rows = _load_checklist_xlsx_rows(args.xlsx, sheet_name=str(args.sheet))
    auto_rows = _load_autonomous_rows(args.auto_json)
    report = compute_parity(headers, base_rows, auto_rows)

    if args.out:
        Path(args.out).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    # Always print summary for quick feedback.
    print(json.dumps(report["summary"], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()

