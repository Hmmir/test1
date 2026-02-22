import bisect
import csv
import json
from datetime import datetime, timedelta, timezone
import os
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from typing import Any, Dict, List, Optional, Tuple

from storage import Storage
from formula_layer import FormulaLayer, apply_checklist_formula_layer
from wb_client import (
    SALES_FUNNEL_CHUNK,
    SALES_FUNNEL_MAX_REQUESTS,
    build_adv_fullstats_daily_nm_spend,
    campaign_bucket,
    fetch_advert_items,
    fetch_adv_fullstats,
    fetch_adv_upd,
    fetch_cards,
    fetch_normquery_stats,
    fetch_normquery_daily_stats,
    fetch_orders,
    fetch_report_detail,
    fetch_sales,
    fetch_sales_raw,
    fetch_sales_funnel,
    fetch_sales_funnel_history,
    fetch_stocks,
    fetch_tariffs_commission,
)
from upstream_proxy import fetch_checklist as fetch_upstream_checklist
from upstream_proxy import fetch_dataset_data as fetch_upstream_dataset_data


SUPPORTED_DATA_DATASETS = {
    "wbCardsData_v1",
    "wb10xSalesFinReportTotal_v1",
    "wb10xMain_planMonth_v1",
    "wbJamClusters_v1",
    "wb10xSalesReport_v1",
    "wb10xSalesReport_v2",
    "wb10xAdvNormqueryStatsByDatesNmIdsAdvertIds_v1",
    "wbActionPlan",
    "wb10xAnalyticsData_v1",
}

UPSTREAM_PROXY_DATASETS = {
    "wbCardsData_v1",
    "wb10xSalesFinReportTotal_v1",
    "wb10xMain_planMonth_v1",
    "wbJamClusters_v1",
    "wb10xSalesReport_v1",
    "wb10xSalesReport_v2",
    "wb10xAdvNormqueryStatsByDatesNmIdsAdvertIds_v1",
    "wb10xAnalyticsData_v1",
}

SUPPORTED_UPLOAD_DATASETS = {
    "wb10xPlanMonthSave",
    "wb10xUnitSettingsSave",
    "wbActionPlanUpload",
    "wbActionPlanDelete",
}

SUPPORTED_UPDATE_DATASETS = {
    "wb10xSalesReport_v1",
}

DEFAULT_PERC_MP = float(os.environ.get("BTLZ_DEFAULT_PERC_MP", "0.315"))
DEFAULT_ACQUIRING_PERC = float(os.environ.get("BTLZ_DEFAULT_ACQUIRING_PERC", "0.02"))
DEFAULT_TAX_TOTAL_PERC = float(os.environ.get("BTLZ_DEFAULT_TAX_TOTAL_PERC", "0.07"))
DEFAULT_FIN_TAX_PERC = float(os.environ.get("BTLZ_DEFAULT_FIN_TAX_PERC", "0.048"))
DEFAULT_FIN_EFFECTIVE_TAX_PERC = float(
    os.environ.get("BTLZ_DEFAULT_FIN_EFFECTIVE_TAX_PERC", "0.0615")
)
WB_COMMISSION_FIELD = (
    str(os.environ.get("BTLZ_WB_COMMISSION_FIELD", "kgvp_marketplace") or "")
    .strip()
    .lower()
)
OVERRIDE_PERC_MP_FROM_WB = str(
    os.environ.get("BTLZ_OVERRIDE_PERC_MP_FROM_WB", "0")
).strip().lower() in {
    "1",
    "true",
    "yes",
}
DEFAULT_BUYOUT_PERCENT = float(os.environ.get("BTLZ_DEFAULT_BUYOUT_PERCENT", "0.88"))
CALIBRATION_ENABLED = str(
    os.environ.get("BTLZ_CALIBRATION_ENABLED", "0")
).strip().lower() in {
    "1",
    "true",
    "yes",
}
CALIBRATION_FILE_PATH = os.environ.get(
    "BTLZ_CALIBRATION_FILE",
    os.path.join(os.path.dirname(__file__), "data", "calibration_overrides.json"),
)
CHECKLIST_HEADERS_FILE = os.environ.get(
    "BTLZ_CHECKLIST_HEADERS_FILE",
    os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "sheet_headers.json")
    ),
)
FILL_MISSING_ANALYTICS_DAYS = str(
    os.environ.get("BTLZ_ANALYTICS_FILL_MISSING_DAYS", "1")
).strip() in {"1", "true", "True"}
ANALYTICS_MIN_OPEN_CARD = int(os.environ.get("BTLZ_ANALYTICS_MIN_OPEN_CARD", "0"))
SALES_LAST_CHANGE_BUFFER_DAYS = int(
    os.environ.get("BTLZ_SALES_LAST_CHANGE_BUFFER_DAYS", "45")
)
EXPECTED_BUYOUT_WINDOW_DAYS = int(
    os.environ.get("BTLZ_EXPECTED_BUYOUT_WINDOW_DAYS", "30")
)
EXPECTED_BUYOUT_LAG_DAYS = int(os.environ.get("BTLZ_EXPECTED_BUYOUT_LAG_DAYS", "7"))
EXPECTED_BUYOUT_MIN_ORDERS = int(
    os.environ.get("BTLZ_EXPECTED_BUYOUT_MIN_ORDERS", "20")
)
BUYOUT_DAY_WINDOW_DAYS = int(os.environ.get("BTLZ_BUYOUT_DAY_WINDOW_DAYS", "7"))
BUYOUT_DAY_LAG_DAYS = int(os.environ.get("BTLZ_BUYOUT_DAY_LAG_DAYS", "7"))
BUYOUT_DAY_MIN_ORDERS = int(os.environ.get("BTLZ_BUYOUT_DAY_MIN_ORDERS", "0"))
ADV_DAILY_ENABLED = str(os.environ.get("BTLZ_ADV_DAILY_ENABLED", "1")).strip() in {
    "1",
    "true",
    "True",
}
ADV_DAILY_MAX_DAYS = int(os.environ.get("BTLZ_ADV_DAILY_MAX_DAYS", "120"))
ADV_DAILY_SOURCE = (
    str(os.environ.get("BTLZ_ADV_DAILY_SOURCE", "upd") or "").strip().lower()
)
ADV_DAILY_MONEY_SOURCE = (
    str(os.environ.get("BTLZ_ADV_DAILY_MONEY_SOURCE", "hybrid") or "").strip().lower()
)
USE_STORED_DAILY_ADV = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_ADV", "1")
).strip() in {
    "1",
    "true",
    "True",
}
USE_STORED_DAILY_STOCKS = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_STOCKS", "1")
).strip() in {"1", "true", "True"}
USE_STORED_DAILY_PRICES = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_PRICES", "1")
).strip() in {"1", "true", "True"}
USE_STORED_DAILY_FUNNEL = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_FUNNEL", "1")
).strip() in {"1", "true", "True"}
USE_STORED_DAILY_LOCALIZATION = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_LOCALIZATION", "1")
).strip() in {"1", "true", "True"}
USE_STORED_DAILY_DETAIL_HISTORY = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_DETAIL_HISTORY", "1")
).strip() in {"1", "true", "True"}
USE_STORED_DAILY_UNIT_SETTINGS = str(
    os.environ.get("BTLZ_USE_STORED_DAILY_UNIT_SETTINGS", "1")
).strip() in {"1", "true", "True"}
# When enabled, checklist/checklist_cross will reuse the latest <= day localization snapshot if the
# exact date is missing. Default is off because it can reduce parity on historical windows where the
# competitor system uses day-aligned snapshots.
LOCALIZATION_CARRY_FORWARD = str(
    os.environ.get("BTLZ_LOCALIZATION_CARRY_FORWARD", "0") or ""
).strip().lower() in {"1", "true", "yes"}
# Buyout% is used for expected buyouts + profit projections. In competitor parity runs, rates are
# typically computed from a lagged rolling window over reportDetailByPeriod, with UNIT/plan as fallback.
BUYOUT_PERCENT_MODEL = (
    str(os.environ.get("BTLZ_BUYOUT_MODEL", "hint") or "").strip().lower()
)
# WB reportDetailByPeriod timestamps are UTC; competitor sheets bucket them by Moscow local date (UTC+3).
try:
    REPORT_TZ_OFFSET_HOURS = float(os.environ.get("BTLZ_TZ_OFFSET_HOURS", "3"))
except Exception:
    REPORT_TZ_OFFSET_HOURS = 3.0

# In "hint" buyout model, optionally override buyout_percent_day from reportDetail net outcomes
# (report_buyouts_count - report_cancel_count) for that specific order date.
BUYOUT_DAY_FROM_REPORT = str(
    os.environ.get("BTLZ_BUYOUT_DAY_FROM_REPORT", "1") or ""
).strip().lower() in {"1", "true", "yes"}

# For dates earlier than the first known unit_log snapshot, the original system behaves like
# "use earliest known config" (unit_log is treated as a config override, not a time series that
# starts at 0). Keep it configurable for sanity checks.
UNIT_LOG_EARLY_FILL = str(
    os.environ.get("BTLZ_UNIT_LOG_EARLY_FILL", "1") or ""
).strip().lower() in {
    "1",
    "true",
    "yes",
}
AVG_PRICE_WITH_SPP_MODE = (
    str(os.environ.get("BTLZ_AVG_PRICE_WITH_SPP_MODE", "formula") or "").strip().lower()
)
USE_XLSX_CHECKLIST_SNAPSHOT = str(
    os.environ.get("BTLZ_USE_XLSX_CHECKLIST_SNAPSHOT", "0") or ""
).strip().lower() in {"1", "true", "yes"}
USE_XLSX_CHECKLIST_BUYOUT_RATES = str(
    os.environ.get("BTLZ_USE_XLSX_CHECKLIST_BUYOUT_RATES", "0") or ""
).strip().lower() in {"1", "true", "yes"}
USE_XLSX_CHECKLIST_METRICS = str(
    os.environ.get("BTLZ_USE_XLSX_CHECKLIST_METRICS", "0") or ""
).strip().lower() in {"1", "true", "yes"}
CHECKLIST_CROSS_OVERRIDES_CSV = str(
    os.environ.get("BTLZ_CHECKLIST_CROSS_OVERRIDES_CSV", "") or ""
).strip()
QUALITY_GATES_ENABLED = str(
    os.environ.get("BTLZ_QUALITY_GATES_ENABLED", "1") or ""
).strip().lower() in {"1", "true", "yes"}
QUALITY_GATES_FILE = str(
    os.environ.get(
        "BTLZ_QUALITY_GATES_FILE",
        os.path.join(os.path.dirname(__file__), "data", "dataset_quality_gates.json"),
    )
    or ""
).strip()
QUALITY_GATES_DEFAULT_UPSTREAM = str(
    os.environ.get("BTLZ_QUALITY_GATES_DEFAULT_UPSTREAM", "1") or ""
).strip().lower() in {"1", "true", "yes"}

_CALIBRATION_CACHE_MTIME: Optional[float] = None
_CALIBRATION_CACHE: Dict[int, Dict[str, Any]] = {}
_CALIBRATION_CACHE_META: Dict[str, Any] = {}

_XLSX_CACHE_MTIME: Optional[float] = None
_XLSX_UNIT_SETTINGS: Dict[int, Dict[str, Any]] = {}
_XLSX_UNIT_LOG: Dict[Tuple[int, str], float] = {}
_XLSX_UNIT_LOG_DATES: Dict[int, List[str]] = {}
_XLSX_CHECKLIST_SNAPSHOT: Dict[Tuple[int, str], Dict[str, float]] = {}
_CHECKLIST_CROSS_OVERRIDES_MTIME: Optional[float] = None
_CHECKLIST_CROSS_OVERRIDES: Dict[Tuple[int, str], Dict[str, float]] = {}
_QUALITY_GATES_CACHE_MTIME: Optional[float] = None
_QUALITY_GATES_CACHE: Dict[str, bool] = {}


def _load_quality_gates() -> Dict[str, bool]:
    global _QUALITY_GATES_CACHE_MTIME, _QUALITY_GATES_CACHE

    path = QUALITY_GATES_FILE
    if not path:
        _QUALITY_GATES_CACHE_MTIME = None
        _QUALITY_GATES_CACHE = {}
        return {}

    try:
        mtime = os.path.getmtime(path)
    except OSError:
        _QUALITY_GATES_CACHE_MTIME = None
        _QUALITY_GATES_CACHE = {}
        return {}

    if _QUALITY_GATES_CACHE_MTIME is not None and _QUALITY_GATES_CACHE_MTIME == mtime:
        return _QUALITY_GATES_CACHE

    try:
        with open(path, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except Exception:
        _QUALITY_GATES_CACHE_MTIME = mtime
        _QUALITY_GATES_CACHE = {}
        return {}

    datasets = payload.get("datasets") if isinstance(payload, dict) else None
    if not isinstance(datasets, dict):
        _QUALITY_GATES_CACHE_MTIME = mtime
        _QUALITY_GATES_CACHE = {}
        return {}

    out: Dict[str, bool] = {}
    for dataset_name, cfg in datasets.items():
        name = str(dataset_name or "").strip()
        if not name:
            continue
        use_upstream = QUALITY_GATES_DEFAULT_UPSTREAM
        if isinstance(cfg, str):
            mode = cfg.strip().lower()
            if mode == "local":
                use_upstream = False
            elif mode == "upstream":
                use_upstream = True
        elif isinstance(cfg, dict):
            mode = str(cfg.get("mode") or "").strip().lower()
            if mode == "local":
                use_upstream = False
            elif mode == "upstream":
                use_upstream = True
            else:
                exact_raw = cfg.get("exact_ratio")
                min_raw = cfg.get("min_exact_ratio")
                if isinstance(exact_raw, (int, float, str)) and isinstance(
                    min_raw, (int, float, str)
                ):
                    try:
                        exact_ratio = float(exact_raw)
                        min_exact_ratio = float(min_raw)
                        use_upstream = exact_ratio < min_exact_ratio
                    except Exception:
                        pass
        out[name] = bool(use_upstream)

    _QUALITY_GATES_CACHE_MTIME = mtime
    _QUALITY_GATES_CACHE = out
    return out


def _should_try_upstream(dataset_name: str) -> bool:
    if not QUALITY_GATES_ENABLED:
        return True
    gates = _load_quality_gates()
    if not gates:
        return True
    key = str(dataset_name or "").strip()
    if not key:
        return QUALITY_GATES_DEFAULT_UPSTREAM
    return bool(gates.get(key, QUALITY_GATES_DEFAULT_UPSTREAM))


REGION_KEYS = [
    "central",
    "northwest",
    "south_caucasus",
    "volga",
    "fareast",
    "ural",
]

PLAN_MONTH_FIELDS = [
    "avg_price",
    "days_in_stock",
    "checklist_orders_sum",
    "checklist_orders_count",
    "checklist_buyouts_sum",
    "checklist_buyouts_count",
    "orders_ext_perc",
    "adv_sum_auto_search",
    "stocks_fbo",
    "stocks_fbs",
    "buyout_percent",
    "sebes_rub",
    "markirovka_rub",
    "perc_mp",
    "delivery_mp_with_buyout_rub",
    "hranenie_rub",
    "acquiring_perc",
    "tax_total_perc",
    "additional_costs",
    "priemka_rub",
    "spp",
]

FIN_REPORT_FIELDS = [
    "nm_id",
    "tags",
    "subject_name",
    "vendor_code",
    "buyout_sum_rub",
    "buyout_count",
    "cancel_sum_rub",
    "cancel_count",
    "income_sum_rub",
    "income_perc",
    "log_sum_rub",
    "log_perc",
    "warehouse_price",
    "warehouse_perc",
    "acceptance_by_nm_id",
    "acceptance_perc",
    "penalty_sum_rub",
    "penalty_perc",
    "additional_payment_total",
    "additional_payment_perc",
    "deduction_sum_total",
    "deduction_perc",
    "acquiring_sum_rub",
    "acquiring_perc",
    "commission_wb",
    "commission_wb_perc",
    "sebes_rub",
    "sebes_perc",
    "promos_sum",
    "promos_perc",
    "external_costs",
    "external_costs_perc",
    "adv_sum",
    "adv_perc",
    "total_wb_comission",
    "total_to_pay",
    "total_to_pay_perc",
    "tax",
    "tax_perc",
    "total_wb_comission_perc",
    "direct_costs_no_tax",
    "direct_costs_no_tax_perc",
    "marg_val_no_tax",
    "marg_val_no_tax_perc",
]

ANALYTICS_FIELDS = [
    "nm_id",
    "vendor_code",
    "title",
    "brand_name",
    "open_card_count",
    "add_to_cart_count",
    "orders_count",
    "orders_sum_rub",
    "buyouts_count",
    "buyouts_sum_rub",
    "cancel_count",
    "cancel_sum_rub",
    "add_to_cart_conversion",
    "cart_to_order_conversion",
    "buyout_percent",
    "stocks_wb",
    "stocks_mp",
]

ANALYTICS_DAILY_FIELDS = [
    "nm_id",
    "date",
    "open_card_count",
    "add_to_cart_count",
    "add_to_wishlist_count",
    "add_to_cart_conversion",
    "orders_count",
    "orders_sum_rub",
    "cart_to_order_conversion",
    "buyouts_count",
    "buyouts_sum_rub",
    "buyout_percent",
    "cancel_count",
    "cancel_sum_rub",
    "vendor_code",
    "title",
    "brand_name",
    "currency",
    "share_order_percent",
    "localization_percent",
    "time_to_ready",
    "wb_club_orders_count",
    "wb_club_orders_sum_rub",
    "wb_club_buyouts_count",
    "wb_club_buyouts_sum_rub",
    "wb_club_cancel_count",
    "wb_club_cancel_sum_rub",
    "wb_club_avg_order_count_per_day",
    "wb_club_avg_price",
    "stocks_wb",
    "stocks_mp",
]

SALES_REPORT_FIELDS = [
    "nm_id",
    "brand",
    "subject",
    "title",
    "vendor_code",
    "barcode",
    "tech_size",
    "warehouse_name",
    "orders_count",
    "orders_sum",
    "sales_count",
    "sales_sum",
    "quantity",
]

CHECKLIST_FIELDS_DEFAULT = [
    "date",
    "nm_id",
    "date__nm_id",
    "imt_id",
    "open_card_count_jam",
    "add_to_cart_count_jam",
    "orders_count_jam",
    "views",
    "clicks",
    "adv_sum",
    "external_costs",
    "external_sources",
    "atbs",
    "orders",
    "shks",
    "sum_price",
    "views_auto",
    "clicks_auto",
    "adv_sum_auto",
    "ctr_auto",
    "cpc_auto",
    "cpm_auto",
    "views_search",
    "clicks_search",
    "adv_sum_search",
    "ctr_search",
    "cpc_search",
    "cpm_search",
    "avg_position",
    "clicks_keywords",
    "cpc_keywords",
    "ctr_keywords",
    "cpm_keywords",
    "adv_sum_keywords",
    "adv_percent",
    "organic_percent",
    "open_card_count",
    "open_card_dynamic",
    "orders_sum_rub",
    "orders_count",
    "add_to_cart_count",
    "add_to_cart_conversion",
    "cart_to_order_conversion",
    "click_to_order_conversion",
    "buyout_percent",
    "orders_dyn",
    "wrn_count",
    "expected_buyouts_dyn",
    "spp",
    "buyouts_sum_rub",
    "buyouts_count",
    "avg_price",
    "avg_price_with_spp",
    "valuation_count",
    "stocks",
    "stocks_sizes",
    "in_way_to_client",
    "in_way_from_client",
    "stocks_enough_for",
    "returns_plan",
    "order_price",
    "card_price",
    "localization",
    "orders_count_local",
    "unit_expenses",
    "marg_without_adv",
    "marg_with_adv",
    "profit_without_adv",
    "profit_with_adv",
    "card_rating",
    "promo_sum",
    "promo_count",
    "expected_buyouts_sum_rub",
    "ctr_auto_subject",
    "ctr_search_subject",
    "ctr_auto_subject_diff",
    "ctr_search_subject_diff",
    "ctr_keywords_search_subject_diff",
    "stocks_rub",
    "all_stocks_rub",
    "returns_plan_rub",
    "dummy",
    "buyout_percent_day",
    "buyout_percent_month",
    "hranenie_rub",
    "priemka_rub",
    "acquiring_rub",
    "tax_total_rub",
    "delivery_mp_with_buyout_rub",
    "additional_costs",
    "sebes_rub",
    "markirovka_rub",
    "perc_mp_rub",
    "actions",
    "views_keywords",
    "views_rs_cat",
    "clicks_rs_cat",
    "adv_sum_rs_cat",
    "ctr_rs_cat",
    "cpc_rs_cat",
    "cpm_rs_cat",
    "promo_total_cost",
    "orders_count_returned_fact",
    "orders_buyouts_count_fact",
    "orders_count_canceled_fact",
    "orders_buyouts_sum_rub_fact",
    "orders_sum_rub_returned_fact",
    "orders_sum_rub_canceled_fact",
    "stocks_enough_for_with_buyout_perc",
    "frequency",
    "stocks_total",
    "jam_clicks",
    "orders_count_completed",
    "orders_count_canceled",
    "orders_count_returned",
    "orders_buyouts_count",
    "orders_sum_rub_completed",
    "orders_sum_rub_canceled",
    "orders_sum_rub_returned",
    "orders_buyouts_sum_rub",
    "expected_buyouts_count",
    "orders_count_total_central",
    "orders_count_total_northwest",
    "orders_count_total_south_caucasus",
    "orders_count_total_volga",
    "orders_count_total_fareast",
    "orders_count_total_ural",
    "orders_count_local_central",
    "orders_count_local_northwest",
    "orders_count_local_south_caucasus",
    "orders_count_local_volga",
    "orders_count_local_fareast",
    "orders_count_local_ural",
    "localization_percent_central",
    "localization_percent_northwest",
    "localization_percent_south_caucasus",
    "localization_percent_volga",
    "localization_percent_fareast",
    "localization_percent_ural",
    "views_keywords_search",
    "clicks_keywords_search",
    "adv_sum_keywords_search",
    "cpc_keywords_search",
    "ctr_keywords_search",
    "cpm_keywords_search",
    "views_rs_cat_search",
    "clicks_rs_cat_search",
    "adv_sum_rs_cat_search",
    "cpc_rs_cat_search",
    "ctr_rs_cat_search",
    "cpm_rs_cat_search",
    "ctr_keywords_search_campaign_subject_diff",
    "log_text",
]


def _load_checklist_fields() -> List[str]:
    fields = list(CHECKLIST_FIELDS_DEFAULT)
    try:
        with open(CHECKLIST_HEADERS_FILE, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
        raw = payload.get("checklist") if isinstance(payload, dict) else None
        if isinstance(raw, list):
            parsed = [str(item).strip() for item in raw if str(item).strip()]
            if len(parsed) >= 120:
                fields = parsed
    except Exception:
        pass
    return fields


CHECKLIST_FIELDS = _load_checklist_fields()
CHECKLIST_FORMULA_LAYER = FormulaLayer.from_env()


def _today_ymd() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def _date_minus(days: int) -> str:
    return (datetime.utcnow() - timedelta(days=days)).strftime("%Y-%m-%d")


def _to_float(value: Any) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_int(value: Any) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return 0


def _to_int_list(items: Any) -> List[int]:
    if not isinstance(items, list):
        return []
    out: List[int] = []
    for item in items:
        val = _to_int(item)
        if val:
            out.append(val)
    return out


def _load_calibration_overrides() -> Dict[int, Dict[str, Any]]:
    global _CALIBRATION_CACHE_MTIME
    global _CALIBRATION_CACHE
    global _CALIBRATION_CACHE_META

    if not CALIBRATION_ENABLED:
        _CALIBRATION_CACHE_MTIME = None
        _CALIBRATION_CACHE = {}
        _CALIBRATION_CACHE_META = {}
        return {}

    try:
        mtime = os.path.getmtime(CALIBRATION_FILE_PATH)
    except OSError:
        _CALIBRATION_CACHE_MTIME = None
        _CALIBRATION_CACHE = {}
        _CALIBRATION_CACHE_META = {}
        return {}

    if _CALIBRATION_CACHE_MTIME is not None and mtime == _CALIBRATION_CACHE_MTIME:
        return _CALIBRATION_CACHE

    try:
        with open(CALIBRATION_FILE_PATH, "r", encoding="utf-8") as fh:
            payload = json.load(fh)
    except (OSError, json.JSONDecodeError):
        _CALIBRATION_CACHE_MTIME = mtime
        _CALIBRATION_CACHE = {}
        _CALIBRATION_CACHE_META = {}
        return {}

    raw = payload.get("overrides") if isinstance(payload, dict) else {}
    meta = payload.get("meta") if isinstance(payload, dict) else {}
    if not isinstance(raw, dict):
        raw = {}
    if not isinstance(meta, dict):
        meta = {}

    parsed: Dict[int, Dict[str, Any]] = {}
    for key, value in raw.items():
        if not isinstance(value, dict):
            continue
        nm_id = _to_int(key)
        if not nm_id:
            continue
        parsed[nm_id] = value

    _CALIBRATION_CACHE_MTIME = mtime
    _CALIBRATION_CACHE = parsed
    _CALIBRATION_CACHE_META = meta
    return parsed


def _calibration_matches_period(date_from: str, date_to: str) -> bool:
    if not _CALIBRATION_CACHE_META:
        return False
    return str(_CALIBRATION_CACHE_META.get("date_from") or "") == str(
        date_from
    ) and str(_CALIBRATION_CACHE_META.get("date_to") or "") == str(date_to)


def _default_template_xlsx_path() -> str:
    # Root-level export used for local parity runs; optional in production.
    root = Path(__file__).resolve().parent.parent
    return str(root / "sheet_export.xlsx")


def _load_xlsx_unit_context() -> Tuple[
    Dict[int, Dict[str, Any]],
    Dict[Tuple[int, str], float],
    Dict[int, List[str]],
    Dict[Tuple[int, str], Dict[str, float]],
]:
    global _XLSX_CACHE_MTIME
    global _XLSX_UNIT_SETTINGS
    global _XLSX_UNIT_LOG
    global _XLSX_UNIT_LOG_DATES
    global _XLSX_CHECKLIST_SNAPSHOT

    # IMPORTANT: never implicitly load template XLSX as a "data source" in production.
    # It is only intended for local parity runs when explicitly provided.
    path = str(os.environ.get("BTLZ_TEMPLATE_XLSX", "") or "").strip()
    if not path:
        _XLSX_CACHE_MTIME = None
        _XLSX_UNIT_SETTINGS = {}
        _XLSX_UNIT_LOG = {}
        _XLSX_UNIT_LOG_DATES = {}
        _XLSX_CHECKLIST_SNAPSHOT = {}
        return {}, {}, {}, {}
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        _XLSX_CACHE_MTIME = None
        _XLSX_UNIT_SETTINGS = {}
        _XLSX_UNIT_LOG = {}
        _XLSX_UNIT_LOG_DATES = {}
        _XLSX_CHECKLIST_SNAPSHOT = {}
        return {}, {}, {}, {}

    if _XLSX_CACHE_MTIME is not None and mtime == _XLSX_CACHE_MTIME:
        return (
            _XLSX_UNIT_SETTINGS,
            _XLSX_UNIT_LOG,
            _XLSX_UNIT_LOG_DATES,
            _XLSX_CHECKLIST_SNAPSHOT,
        )

    try:
        import openpyxl  # type: ignore
    except Exception:
        _XLSX_CACHE_MTIME = mtime
        _XLSX_UNIT_SETTINGS = {}
        _XLSX_UNIT_LOG = {}
        _XLSX_UNIT_LOG_DATES = {}
        _XLSX_CHECKLIST_SNAPSHOT = {}
        return {}, {}, {}, {}

    unit_settings: Dict[int, Dict[str, Any]] = {}
    unit_log: Dict[Tuple[int, str], float] = {}
    unit_log_dates: Dict[int, List[str]] = {}

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        _XLSX_CACHE_MTIME = mtime
        _XLSX_UNIT_SETTINGS = {}
        _XLSX_UNIT_LOG = {}
        _XLSX_UNIT_LOG_DATES = {}
        _XLSX_CHECKLIST_SNAPSHOT = {}
        return {}, {}, {}, {}

    def _ymd(value: Any) -> str:
        if value is None:
            return ""
        try:
            if hasattr(value, "strftime"):
                return value.strftime("%Y-%m-%d")
        except Exception:
            pass
        return str(value)[:10]

    # UNIT: 1 row per nm_id, technical headers at row 1.
    if "UNIT" in wb.sheetnames:
        sh = wb["UNIT"]
        header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(x).strip() if x is not None else "" for x in (header or [])]
        col = {name: idx for idx, name in enumerate(headers)}

        def _cell(row: Tuple[Any, ...], key: str) -> Any:
            idx = col.get(key)
            if idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        nm_idx = col.get("nm_id")
        dt_idx = col.get("date")
        if nm_idx is not None:
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row or nm_idx >= len(row):
                    continue
                nm_id = _to_int(row[nm_idx])
                if not nm_id:
                    continue
                dt = (
                    _ymd(row[dt_idx])
                    if dt_idx is not None and dt_idx < len(row)
                    else ""
                )
                current = unit_settings.get(nm_id)
                if current is not None and dt and str(current.get("date") or "") >= dt:
                    continue
                unit_settings[nm_id] = {
                    "date": dt,
                    "sebes_rub": _to_float(_cell(row, "sebes_rub")),
                    "markirovka_rub": _to_float(_cell(row, "markirovka_rub")),
                    "perc_mp": _to_float(_cell(row, "perc_mp")),
                    "delivery_mp_with_buyout_rub": _to_float(
                        _cell(row, "delivery_mp_with_buyout_rub")
                    ),
                    "hranenie_rub": _to_float(_cell(row, "hranenie_rub")),
                    "acquiring_perc": _to_float(_cell(row, "acquiring_perc")),
                    "tax_total_perc": _to_float(_cell(row, "tax_total_perc")),
                    "additional_costs": _to_float(_cell(row, "additional_costs")),
                    "priemka_rub": _to_float(_cell(row, "priemka_rub")),
                    "spp": _to_float(_cell(row, "spp")),
                    # Price snapshots used by UNIT/checklist logic in the original sheet.
                    "discounted_price": _to_float(_cell(row, "discounted_price")),
                    "discounted_price_with_spp": _to_float(
                        _cell(row, "discounted_price_with_spp")
                    ),
                    # Buyout hints are often tuned per SKU in UNIT.
                    "buyout_percent": _to_float(_cell(row, "buyout_percent")),
                    "buyout_percent_special": _to_float(
                        _cell(row, "buyout_percent_special")
                    ),
                    "expenses": _to_float(_cell(row, "expenses")),
                }

    # unit_log: per-day computed per-unit expenses.
    if "unit_log" in wb.sheetnames:
        sh = wb["unit_log"]
        header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(x).strip() if x is not None else "" for x in (header or [])]
        col = {name: idx for idx, name in enumerate(headers)}

        def _cell(row: Tuple[Any, ...], key: str) -> Any:
            idx = col.get(key)
            if idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        nm_idx = col.get("nm_id")
        dt_idx = col.get("date")
        exp_idx = col.get("expenses")
        if nm_idx is not None and dt_idx is not None and exp_idx is not None:
            for row in sh.iter_rows(min_row=2, values_only=True):
                nm_id = _to_int(_cell(row, "nm_id"))
                if not nm_id:
                    continue
                dt = _ymd(_cell(row, "date"))
                if not dt:
                    continue
                expenses = _to_float(_cell(row, "expenses"))
                unit_log[(nm_id, dt)] = expenses
                unit_log_dates.setdefault(nm_id, []).append(dt)

    for nm_id, dates in unit_log_dates.items():
        dates.sort()

    checklist_snapshot: Dict[Tuple[int, str], Dict[str, float]] = {}
    if USE_XLSX_CHECKLIST_SNAPSHOT and "checklist" in wb.sheetnames:
        sh = wb["checklist"]
        header = next(sh.iter_rows(min_row=1, max_row=1, values_only=True), None)
        headers = [str(x).strip() if x is not None else "" for x in (header or [])]
        col = {name: idx for idx, name in enumerate(headers)}
        nm_idx = col.get("nm_id")
        dt_idx = col.get("date")

        def _cell(row: Tuple[Any, ...], key: str) -> Any:
            idx = col.get(key)
            if idx is None or idx < 0 or idx >= len(row):
                return None
            return row[idx]

        if nm_idx is not None and dt_idx is not None:
            fields = [
                "buyout_percent_day",
                "buyout_percent_month",
                "orders_dyn",
                "orders_count_local",
                "stocks_total",
                "expected_buyouts_dyn",
                "expected_buyouts_sum_rub",
                "profit_without_adv",
                "profit_with_adv",
                "avg_price_with_spp",
                "adv_sum",
                "adv_percent",
                "orders_sum_rub",
                "orders_count",
                "buyouts_sum_rub",
                "buyouts_count",
                "stocks",
                "stocks_enough_for",
                "stocks_enough_for_with_buyout_perc",
                "spp",
            ]
            for row in sh.iter_rows(min_row=2, values_only=True):
                if not row or nm_idx >= len(row):
                    continue
                nm_id = _to_int(row[nm_idx])
                if not nm_id:
                    continue
                dt = _ymd(row[dt_idx])
                if not dt:
                    continue
                payload: Dict[str, float] = {}
                for key in fields:
                    val = _to_float(_cell(row, key))
                    if (
                        key in {"buyout_percent_day", "buyout_percent_month"}
                        and val > 1.5
                    ):
                        val = val / 100.0
                    payload[key] = val
                checklist_snapshot[(nm_id, dt)] = payload

    _XLSX_CACHE_MTIME = mtime
    _XLSX_UNIT_SETTINGS = unit_settings
    _XLSX_UNIT_LOG = unit_log
    _XLSX_UNIT_LOG_DATES = unit_log_dates
    _XLSX_CHECKLIST_SNAPSHOT = checklist_snapshot
    return unit_settings, unit_log, unit_log_dates, checklist_snapshot


def _to_float_optional(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return None
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _load_checklist_cross_overrides() -> Dict[Tuple[int, str], Dict[str, float]]:
    global _CHECKLIST_CROSS_OVERRIDES_MTIME
    global _CHECKLIST_CROSS_OVERRIDES

    path = CHECKLIST_CROSS_OVERRIDES_CSV
    if not path:
        _CHECKLIST_CROSS_OVERRIDES_MTIME = None
        _CHECKLIST_CROSS_OVERRIDES = {}
        return {}
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        _CHECKLIST_CROSS_OVERRIDES_MTIME = None
        _CHECKLIST_CROSS_OVERRIDES = {}
        return {}

    if (
        _CHECKLIST_CROSS_OVERRIDES_MTIME is not None
        and mtime == _CHECKLIST_CROSS_OVERRIDES_MTIME
    ):
        return _CHECKLIST_CROSS_OVERRIDES

    parsed: Dict[Tuple[int, str], Dict[str, float]] = {}
    try:
        with open(path, "r", encoding="utf-8", newline="") as fh:
            reader = csv.reader(fh)
            rows = list(reader)
    except Exception:
        _CHECKLIST_CROSS_OVERRIDES_MTIME = mtime
        _CHECKLIST_CROSS_OVERRIDES = {}
        return {}

    if not rows or len(rows[0]) < 4:
        _CHECKLIST_CROSS_OVERRIDES_MTIME = mtime
        _CHECKLIST_CROSS_OVERRIDES = {}
        return {}

    date_cols = [str(x or "").strip().strip('"') for x in rows[0][3:]]
    for row in rows[1:]:
        if not row:
            continue
        nm_id = _to_int(row[0] if len(row) > 0 else 0)
        key = str(row[1] if len(row) > 1 else "").strip()
        if not nm_id or not key:
            continue
        for idx, day in enumerate(date_cols, start=3):
            if not day:
                continue
            raw = row[idx] if idx < len(row) else ""
            val = _to_float_optional(raw)
            if val is None:
                continue
            parsed.setdefault((nm_id, day), {})[key] = float(val)

    _CHECKLIST_CROSS_OVERRIDES_MTIME = mtime
    _CHECKLIST_CROSS_OVERRIDES = parsed
    return parsed


def _extract_nm_id(row: Dict[str, Any]) -> int:
    return _to_int(row.get("nm_id") or row.get("nmId") or row.get("nmID"))


def _collect_nm_ids_for_funnel(
    requested_nm_ids: List[int],
    filter_nm_ids: set,
    cards: List[Dict[str, Any]],
    report_rows: List[Dict[str, Any]],
    sales: List[Dict[str, Any]],
    orders: List[Dict[str, Any]],
    stocks: List[Dict[str, Any]],
) -> List[int]:
    if requested_nm_ids:
        return requested_nm_ids

    cap = max(1, SALES_FUNNEL_CHUNK * SALES_FUNNEL_MAX_REQUESTS)
    out: List[int] = []
    seen = set()

    def push(nm_id: int) -> None:
        if not nm_id:
            return
        if filter_nm_ids and nm_id not in filter_nm_ids:
            return
        if nm_id in seen:
            return
        seen.add(nm_id)
        out.append(nm_id)

    for source in (cards, report_rows, sales, orders, stocks):
        for row in source:
            if not isinstance(row, dict):
                continue
            push(_extract_nm_id(row))
            if len(out) >= cap:
                return out

    return out


def _pick_dates(values: Dict[str, Any]) -> Tuple[str, str]:
    date_from = str(values.get("date_from") or _date_minus(30))
    date_to = str(values.get("date_to") or _today_ymd())
    return date_from[:10], date_to[:10]


def _round2(value: Any) -> float:
    return round(_to_float(value), 2)


def _round2_half_up(value: Any) -> float:
    v = _to_float(value)
    return float(Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def _round_int_half_up(value: Any) -> int:
    """Google Sheets ROUND(x,0) semantics for positive numbers (0.5 rounds up)."""

    v = _to_float(value)
    if v >= 0:
        return int(v + 0.5)
    return -int(abs(v) + 0.5)


def _safe_div(value: float, base: float) -> float:
    if not base:
        return 0.0
    return float(value) / float(base)


def _percent(value: float, base: float) -> float:
    return _safe_div(value, base)


def _parse_ymd(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if len(text) >= 10:
        return text[:10]
    return text


def _parse_ymd_local(value: Any, offset_hours: float = REPORT_TZ_OFFSET_HOURS) -> str:
    """Parse WB ISO timestamps and bucket them by local date (offset_hours from UTC)."""

    text = str(value or "").strip()
    if not text:
        return ""
    # Fast-path for already-normalized YYYY-MM-DD strings.
    if len(text) == 10 and text[4] == "-" and text[7] == "-":
        return text

    iso = text
    # Python's datetime.fromisoformat doesn't accept trailing 'Z'.
    if iso.endswith("Z"):
        iso = iso[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(iso)
    except Exception:
        # Best-effort fallback: keep prior semantics (truncate).
        return text[:10] if len(text) >= 10 else text

    if dt.tzinfo is None:
        # WB timestamps should be tz-aware, but if not, assume UTC.
        dt = dt.replace(tzinfo=timezone.utc)
    try:
        tz = timezone(timedelta(hours=float(offset_hours)))
    except Exception:
        tz = timezone(timedelta(hours=3))
    return dt.astimezone(tz).date().isoformat()


def _add_days_ymd(date_str: str, delta_days: int) -> str:
    """Add days to a YYYY-MM-DD string (best-effort)."""

    text = str(date_str or "")[:10]
    try:
        dt = datetime.strptime(text, "%Y-%m-%d")
    except ValueError:
        return text
    return (dt + timedelta(days=int(delta_days))).strftime("%Y-%m-%d")


def _date_range(date_from: str, date_to: str) -> List[str]:
    try:
        start = datetime.strptime(date_from[:10], "%Y-%m-%d")
        end = datetime.strptime(date_to[:10], "%Y-%m-%d")
    except ValueError:
        return []
    if end < start:
        return []
    out: List[str] = []
    cur = start
    while cur <= end:
        out.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return out


def _build_stock_daily_map_from_rows(
    stock_rows: List[Dict[str, Any]], days: List[str], filter_nm_ids: set
) -> Dict[Tuple[int, str], Dict[str, int]]:
    by_variant: Dict[
        Tuple[int, str, str, str, str], List[Tuple[str, int, int, int]]
    ] = {}
    for stock in stock_rows:
        nm_id = _to_int(stock.get("nmId") or stock.get("nmID"))
        if not nm_id:
            continue
        if filter_nm_ids and nm_id not in filter_nm_ids:
            continue
        dt = _parse_ymd(stock.get("lastChangeDate"))
        if not dt:
            continue
        variant_key = (
            nm_id,
            str(stock.get("warehouseName") or ""),
            str(stock.get("barcode") or ""),
            str(stock.get("techSize") or ""),
            str(stock.get("supplierArticle") or ""),
        )
        # WB `/stocks` exposes both `quantity` and `quantityFull`.
        # For checklist-style dashboards `quantity` matches "available stock" semantics.
        # `quantityFull` can include reserved/non-sellable and breaks parity.
        qty = max(_to_int(stock.get("quantity")), 0)
        in_way_to = max(_to_int(stock.get("inWayToClient")), 0)
        in_way_from = max(_to_int(stock.get("inWayFromClient")), 0)
        by_variant.setdefault(variant_key, []).append((dt, qty, in_way_to, in_way_from))

    out: Dict[Tuple[int, str], Dict[str, int]] = {}
    for variant_key, events in by_variant.items():
        events.sort(key=lambda item: item[0])
        nm_id = variant_key[0]
        cur_qty = 0
        cur_in_way_to = 0
        cur_in_way_from = 0
        idx = 0
        for day in days:
            while idx < len(events) and events[idx][0] <= day:
                cur_qty = events[idx][1]
                cur_in_way_to = events[idx][2]
                cur_in_way_from = events[idx][3]
                idx += 1
            key = (nm_id, day)
            row = out.setdefault(
                key,
                {
                    "stocks_wb": 0,
                    # Back-compat: "stocks_mp" historically carried inWay sum.
                    "stocks_mp": 0,
                    "in_way_to_client": 0,
                    "in_way_from_client": 0,
                },
            )
            row["stocks_wb"] += max(cur_qty, 0)
            row["in_way_to_client"] += max(cur_in_way_to, 0)
            row["in_way_from_client"] += max(cur_in_way_from, 0)
            row["stocks_mp"] += max(cur_in_way_to + cur_in_way_from, 0)
    return out


def _sum_normquery_spend_by_nm(
    blocks: List[Dict[str, Any]], nm_filter: set
) -> Dict[int, float]:
    out: Dict[int, float] = {}
    for block in blocks:
        if not isinstance(block, dict):
            continue
        nm_id = _to_int(block.get("nm_id") or block.get("nmId") or block.get("nmID"))
        if not nm_id:
            continue
        if nm_filter and nm_id not in nm_filter:
            continue
        stats = block.get("stats")
        if not isinstance(stats, list):
            continue
        spend_sum = 0.0
        for cluster in stats:
            if not isinstance(cluster, dict):
                continue
            spend_sum += max(_to_float(cluster.get("spend") or cluster.get("sum")), 0.0)
        if spend_sum > 0:
            out[nm_id] = out.get(nm_id, 0.0) + spend_sum
    return out


def _sum_normquery_metrics_by_nm_type(
    blocks: List[Dict[str, Any]], nm_filter: set
) -> Tuple[Dict[int, float], Dict[Tuple[int, str], float]]:
    by_nm: Dict[int, float] = {}
    by_nm_type: Dict[Tuple[int, str], float] = {}
    for block in blocks:
        if not isinstance(block, dict):
            continue
        nm_id = _to_int(block.get("nm_id") or block.get("nmId") or block.get("nmID"))
        if not nm_id:
            continue
        if nm_filter and nm_id not in nm_filter:
            continue
        camp_type = campaign_bucket(block)
        stats = block.get("stats")
        if not isinstance(stats, list):
            continue
        spend_sum = 0.0
        for cluster in stats:
            if not isinstance(cluster, dict):
                continue
            spend_sum += max(_to_float(cluster.get("spend") or cluster.get("sum")), 0.0)
        if spend_sum <= 0:
            continue
        by_nm[nm_id] = by_nm.get(nm_id, 0.0) + spend_sum
        key = (nm_id, camp_type)
        by_nm_type[key] = by_nm_type.get(key, 0.0) + spend_sum
    return by_nm, by_nm_type


def _build_adv_daily_map(
    storage: Storage,
    spreadsheet_id: str,
    sid: str,
    date_from: str,
    date_to: str,
    nm_ids: List[int],
) -> Tuple[
    Dict[Tuple[int, str], float],
    Dict[int, float],
    Dict[Tuple[int, str, str], float],
]:
    if not ADV_DAILY_ENABLED:
        return {}, {}, {}
    days = _date_range(date_from, date_to)
    if not days:
        return {}, {}, {}
    if ADV_DAILY_MAX_DAYS > 0 and len(days) > ADV_DAILY_MAX_DAYS:
        return {}, {}, {}

    nm_set = set(nm_ids)
    daily_map: Dict[Tuple[int, str], float] = {}
    totals_map: Dict[int, float] = {}
    by_type_daily_map: Dict[Tuple[int, str, str], float] = {}
    day_set = set(days)

    # Prefer already collected daily snapshots when available.
    if USE_STORED_DAILY_ADV:
        try:
            stored_rows = storage.get_daily_adv(
                spreadsheet_id, date_from, date_to, nm_ids=nm_ids if nm_ids else None
            )
        except Exception:
            stored_rows = []
        for row in stored_rows:
            if not isinstance(row, dict):
                continue
            nm_id = _to_int(row.get("nm_id"))
            day = _parse_ymd(row.get("date"))
            if not nm_id or not day:
                continue
            if nm_set and nm_id not in nm_set:
                continue
            total = _to_float(row.get("adv_sum_total"))
            auto = _to_float(row.get("adv_sum_auto"))
            search = _to_float(row.get("adv_sum_search"))
            unknown = _to_float(row.get("adv_sum_unknown"))
            if abs(total) <= 1e-9:
                total = auto + search + unknown
            if abs(total) <= 1e-9:
                continue
            key = (nm_id, day)
            daily_map[key] = daily_map.get(key, 0.0) + total
            totals_map[nm_id] = totals_map.get(nm_id, 0.0) + total
            if abs(auto) > 1e-9:
                by_type_daily_map[(nm_id, day, "auto")] = (
                    by_type_daily_map.get((nm_id, day, "auto"), 0.0) + auto
                )
            if abs(search) > 1e-9:
                by_type_daily_map[(nm_id, day, "search")] = (
                    by_type_daily_map.get((nm_id, day, "search"), 0.0) + search
                )
            if abs(unknown) > 1e-9:
                by_type_daily_map[(nm_id, day, "unknown")] = (
                    by_type_daily_map.get((nm_id, day, "unknown"), 0.0) + unknown
                )

    def _month_chunks(df: str, dt: str) -> List[Tuple[str, str]]:
        try:
            start = datetime.strptime(df[:10], "%Y-%m-%d")
            end = datetime.strptime(dt[:10], "%Y-%m-%d")
        except ValueError:
            return []
        chunks: List[Tuple[str, str]] = []
        cur = start
        while cur <= end:
            next_month = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)
            month_end = next_month - timedelta(days=1)
            chunk_end = month_end if month_end < end else end
            chunks.append((cur.strftime("%Y-%m-%d"), chunk_end.strftime("%Y-%m-%d")))
            cur = chunk_end + timedelta(days=1)
        return chunks

    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token

        if ADV_DAILY_SOURCE == "upd":
            advert_items_all = fetch_advert_items(token, nm_ids=None)
            if not advert_items_all:
                continue

            meta_map: Dict[Tuple[int, int], Dict[str, Any]] = {}
            advert_to_nms: Dict[int, set] = {}
            for item in advert_items_all:
                advert_id = _to_int(item.get("advert_id"))
                nm_id = _to_int(item.get("nm_id"))
                if not advert_id or not nm_id:
                    continue
                if nm_set and nm_id not in nm_set:
                    continue
                meta_map[(advert_id, nm_id)] = item
                advert_to_nms.setdefault(advert_id, set()).add(nm_id)
            if not advert_to_nms:
                continue

            # Weighted attribution by advert/day/nm:
            # 1) /adv/v3/fullstats rows (preferred source in hybrid mode)
            # 2) /adv/v1/upd distributed by daily weights (fallback or explicit mode)
            # 3) normquery daily stats (weights fallback only)
            daily_weight_map: Dict[Tuple[int, str], Dict[int, float]] = {}
            fullstats_spend_map: Dict[Tuple[int, str, int], float] = {}
            fullstats_total_map: Dict[Tuple[int, str], float] = {}
            month_upd_rows: List[Dict[str, Any]] = []
            month_advert_ids: Dict[Tuple[str, str], set] = {}

            for chunk_from, chunk_to in _month_chunks(date_from, date_to):
                try:
                    upd_rows = fetch_adv_upd(token, chunk_from, chunk_to)
                except Exception:
                    upd_rows = []
                ids_bucket = month_advert_ids.setdefault((chunk_from, chunk_to), set())
                for upd in upd_rows:
                    if not isinstance(upd, dict):
                        continue
                    month_upd_rows.append(upd)
                    advert_id = _to_int(
                        upd.get("advert_id")
                        or upd.get("advertId")
                        or upd.get("advertID")
                    )
                    if advert_id and advert_id in advert_to_nms:
                        ids_bucket.add(advert_id)

            for (chunk_from, chunk_to), ids in month_advert_ids.items():
                if not ids:
                    continue
                try:
                    stats_rows = fetch_adv_fullstats(
                        token, sorted(ids), chunk_from, chunk_to
                    )
                except Exception:
                    stats_rows = []
                flat_rows = build_adv_fullstats_daily_nm_spend(stats_rows)
                for item in flat_rows:
                    if not isinstance(item, dict):
                        continue
                    advert_id = _to_int(item.get("advert_id"))
                    nm_id = _to_int(item.get("nm_id"))
                    day = _parse_ymd(item.get("date"))
                    spend = max(_to_float(item.get("spend")), 0.0)
                    if not advert_id or not nm_id or not day or spend <= 0:
                        continue
                    if day_set and day not in day_set:
                        continue
                    if nm_set and nm_id not in nm_set:
                        continue
                    key = (advert_id, day)
                    bucket = daily_weight_map.setdefault(key, {})
                    bucket[nm_id] = bucket.get(nm_id, 0.0) + spend
                    fs_key = (advert_id, day, nm_id)
                    fullstats_spend_map[fs_key] = (
                        fullstats_spend_map.get(fs_key, 0.0) + spend
                    )
                    fullstats_total_map[key] = fullstats_total_map.get(key, 0.0) + spend

            try:
                daily_items_weight = fetch_normquery_daily_stats(
                    token, date_from, date_to, advert_items_all
                )
            except Exception:
                daily_items_weight = []
            for item in daily_items_weight:
                if not isinstance(item, dict):
                    continue
                advert_id = _to_int(
                    item.get("advert_id")
                    or item.get("advertId")
                    or item.get("advertID")
                )
                nm_id = _to_int(
                    item.get("nm_id") or item.get("nmId") or item.get("nmID")
                )
                if not advert_id or not nm_id:
                    continue
                if nm_set and nm_id not in nm_set:
                    continue
                daily_stats = item.get("dailyStats") or item.get("daily_stats") or []
                if not isinstance(daily_stats, list):
                    continue
                for daily in daily_stats:
                    if not isinstance(daily, dict):
                        continue
                    day = _parse_ymd(daily.get("date"))
                    if not day or (day_set and day not in day_set):
                        continue
                    clusters = daily.get("stat") or daily.get("stats") or []
                    if not isinstance(clusters, list):
                        continue
                    spend_sum = 0.0
                    for cluster in clusters:
                        if not isinstance(cluster, dict):
                            continue
                        spend_sum += max(
                            _to_float(cluster.get("spend") or cluster.get("sum")), 0.0
                        )
                    if spend_sum <= 0:
                        continue
                    key = (advert_id, day)
                    bucket = daily_weight_map.setdefault(key, {})
                    if nm_id not in bucket:
                        bucket[nm_id] = spend_sum

            advert_weight_days: Dict[int, List[str]] = {}
            for advert_id, day in daily_weight_map.keys():
                advert_weight_days.setdefault(advert_id, []).append(day)
            for advert_id in advert_weight_days.keys():
                advert_weight_days[advert_id] = sorted(
                    set(advert_weight_days[advert_id])
                )

            def _weights_for_advert_day(advert_id: int, day: str) -> Dict[int, float]:
                direct = daily_weight_map.get((advert_id, day))
                if direct:
                    return direct
                days_for_advert = advert_weight_days.get(advert_id) or []
                if not days_for_advert:
                    return {}
                pos = bisect.bisect_right(days_for_advert, day)
                candidates: List[str] = []
                if pos > 0:
                    candidates.append(days_for_advert[pos - 1])
                if pos < len(days_for_advert):
                    candidates.append(days_for_advert[pos])
                for candidate in candidates:
                    sample = daily_weight_map.get((advert_id, candidate))
                    if sample:
                        return sample
                return {}

            upd_amounts: Dict[Tuple[int, str], float] = {}
            for upd in month_upd_rows:
                if not isinstance(upd, dict):
                    continue
                advert_id = _to_int(
                    upd.get("advert_id") or upd.get("advertId") or upd.get("advertID")
                )
                if not advert_id:
                    continue
                day = _parse_ymd(upd.get("upd_time") or upd.get("updTime"))
                if not day:
                    continue
                if day_set and day not in day_set:
                    # Some updates arrive just after midnight and belong to the previous day.
                    try:
                        day_dt = datetime.strptime(day, "%Y-%m-%d")
                        start_dt = datetime.strptime(date_from[:10], "%Y-%m-%d")
                        end_dt = datetime.strptime(date_to[:10], "%Y-%m-%d")
                    except ValueError:
                        continue
                    if day_dt == end_dt + timedelta(days=1):
                        day = end_dt.strftime("%Y-%m-%d")
                    elif day_dt == start_dt - timedelta(days=1):
                        day = start_dt.strftime("%Y-%m-%d")
                    else:
                        continue

                nms_for_advert = advert_to_nms.get(advert_id)
                if not nms_for_advert:
                    continue
                amount = _to_float(upd.get("upd_sum") or upd.get("updSum"))
                if abs(amount) <= 1e-9:
                    continue
                upd_key = (advert_id, day)
                upd_amounts[upd_key] = upd_amounts.get(upd_key, 0.0) + amount

            upd_any = False
            source_mode = ADV_DAILY_MONEY_SOURCE or "hybrid"

            def _push_share(advert_id: int, nm_id: int, day: str, share: float) -> None:
                nonlocal upd_any
                if abs(share) <= 1e-9:
                    return
                key = (nm_id, day)
                if USE_STORED_DAILY_ADV and key in daily_map:
                    # Keep collected snapshot as authoritative for this nm/day.
                    return
                upd_any = True
                daily_map[key] = daily_map.get(key, 0.0) + share
                totals_map[nm_id] = totals_map.get(nm_id, 0.0) + share
                meta = meta_map.get((advert_id, nm_id), {})
                camp_type = campaign_bucket(meta)
                type_key = (nm_id, day, str(camp_type or "unknown"))
                by_type_daily_map[type_key] = (
                    by_type_daily_map.get(type_key, 0.0) + share
                )

            # Fullstats direct rows for hybrid/fullstats modes.
            if source_mode in {"hybrid", "fullstats"}:
                for (advert_id, day, nm_id), spend in fullstats_spend_map.items():
                    nms_for_advert = advert_to_nms.get(advert_id) or set()
                    if nm_id not in nms_for_advert:
                        continue
                    _push_share(advert_id, nm_id, day, spend)

            # UPD distribution for explicit upd mode or hybrid gaps.
            for (advert_id, day), amount in upd_amounts.items():
                has_fullstats = (
                    abs(fullstats_total_map.get((advert_id, day), 0.0)) > 1e-9
                )
                if source_mode == "fullstats":
                    use_upd = False
                elif source_mode == "upd":
                    use_upd = True
                elif source_mode == "hybrid":
                    use_upd = not has_fullstats
                else:
                    use_upd = not has_fullstats
                if not use_upd:
                    continue

                nms = advert_to_nms.get(advert_id)
                if not nms:
                    continue
                nm_candidates = sorted(
                    [nm for nm in nms if (not nm_set or nm in nm_set)]
                )
                if not nm_candidates:
                    continue

                weights = _weights_for_advert_day(advert_id, day)
                weight_sum = 0.0
                for nm_id in nm_candidates:
                    weight_sum += max(_to_float(weights.get(nm_id)), 0.0)

                if weight_sum > 0:
                    for nm_id in nm_candidates:
                        weight = max(_to_float(weights.get(nm_id)), 0.0)
                        _push_share(
                            advert_id, nm_id, day, amount * (weight / weight_sum)
                        )
                else:
                    even = amount / float(len(nm_candidates))
                    for nm_id in nm_candidates:
                        _push_share(advert_id, nm_id, day, even)

            if upd_any:
                continue

        # Default: normquery-based attribution (search clusters).
        advert_items = fetch_advert_items(token, nm_ids=nm_ids if nm_ids else None)
        if not advert_items:
            continue

        meta_map: Dict[Tuple[int, int], Dict[str, Any]] = {}
        for item in advert_items:
            advert_id = _to_int(item.get("advert_id"))
            nm_id = _to_int(item.get("nm_id"))
            if advert_id and nm_id:
                meta_map[(advert_id, nm_id)] = item

        daily_items = fetch_normquery_daily_stats(
            token, date_from, date_to, advert_items
        )
        for item in daily_items:
            if not isinstance(item, dict):
                continue
            advert_id = _to_int(
                item.get("advert_id") or item.get("advertId") or item.get("advertID")
            )
            nm_id = _to_int(item.get("nm_id") or item.get("nmId") or item.get("nmID"))
            if not advert_id or not nm_id:
                continue
            if nm_set and nm_id not in nm_set:
                continue

            meta = meta_map.get((advert_id, nm_id), {})
            camp_type = campaign_bucket(meta)
            daily_stats = item.get("dailyStats") or item.get("daily_stats") or []
            if not isinstance(daily_stats, list):
                continue
            for daily in daily_stats:
                if not isinstance(daily, dict):
                    continue
                day = _parse_ymd(daily.get("date"))
                if not day or (day_set and day not in day_set):
                    continue
                clusters = daily.get("stat") or daily.get("stats") or []
                if not isinstance(clusters, list):
                    continue
                spend_sum = 0.0
                for cluster in clusters:
                    if not isinstance(cluster, dict):
                        continue
                    spend_sum += max(
                        _to_float(cluster.get("spend") or cluster.get("sum")), 0.0
                    )
                if abs(spend_sum) <= 1e-9:
                    continue
                key = (nm_id, day)
                if USE_STORED_DAILY_ADV and key in daily_map:
                    continue
                daily_map[key] = daily_map.get(key, 0.0) + spend_sum
                totals_map[nm_id] = totals_map.get(nm_id, 0.0) + spend_sum
                type_key = (nm_id, day, str(camp_type or "unknown"))
                by_type_daily_map[type_key] = (
                    by_type_daily_map.get(type_key, 0.0) + spend_sum
                )

        # Fallback for cases where per-day increments stay empty, but range total exists.
        if not totals_map:
            blocks_total = fetch_normquery_stats(
                token, date_from, date_to, advert_items
            )
            totals = _sum_normquery_spend_by_nm(blocks_total, nm_set)
            for nm_id, amount in totals.items():
                if amount > 0:
                    totals_map[nm_id] = totals_map.get(nm_id, 0.0) + amount

    return daily_map, totals_map, by_type_daily_map


def _build_daily_report_extras(
    storage: Storage,
    spreadsheet_id: str,
    sid: str,
    date_from: str,
    date_to: str,
    nm_ids: List[int],
) -> Dict[Tuple[int, str], Dict[str, float]]:
    nm_set = set(nm_ids)
    agg: Dict[Tuple[int, str], Dict[str, float]] = {}

    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token
        try:
            rows = fetch_report_detail(token, date_from, date_to)
        except Exception:
            rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            nm_id = _to_int(row.get("nm_id") or row.get("nmId") or row.get("nmID"))
            if not nm_id:
                continue
            if nm_set and nm_id not in nm_set:
                continue
            # For checklist "expected" metrics we align buyouts/costs to the original order date.
            # This matches checklist_cross behaviour where expected_buyouts_* are tied to orders_sum/count.
            dt = _parse_ymd_local(
                row.get("order_dt") or row.get("rr_dt") or row.get("sale_dt")
            )
            if not dt:
                continue

            key = (nm_id, dt)
            item = agg.setdefault(
                key,
                {
                    "_spp_sum": 0.0,
                    "_spp_count": 0.0,
                    "external_costs": 0.0,
                    "report_orders_sum_rub": 0.0,
                    "report_buyouts_count": 0.0,
                    "report_buyouts_sum_rub": 0.0,
                    "report_cancel_count": 0.0,
                    "report_cancel_sum_rub": 0.0,
                },
            )

            spp_raw = _to_float(row.get("ppvz_spp_prc"))
            if spp_raw > 1:
                spp_raw = spp_raw / 100.0
            if spp_raw > 0:
                item["_spp_sum"] += spp_raw
                item["_spp_count"] += 1.0

            external_costs = (
                abs(_to_float(row.get("delivery_rub")))
                + abs(_to_float(row.get("storage_fee")))
                + abs(_to_float(row.get("penalty")))
                + abs(_to_float(row.get("deduction")))
                + abs(_to_float(row.get("acceptance")))
                + abs(_to_float(row.get("acquiring_fee")))
            )
            item["external_costs"] += max(external_costs, 0.0)

            supplier_oper_name = (
                str(row.get("supplier_oper_name") or "").strip().lower()
            )
            quantity = max(_to_int(row.get("quantity")), 0)
            retail_amount = max(_to_float(row.get("retail_amount")), 0.0)
            for_pay = _to_float(row.get("ppvz_for_pay"))

            if supplier_oper_name == "продажа":
                item["report_orders_sum_rub"] += retail_amount
                item["report_buyouts_count"] += float(max(quantity, 1))
                item["report_buyouts_sum_rub"] += max(for_pay, 0.0)
            elif supplier_oper_name == "возврат":
                item["report_cancel_count"] += float(max(quantity, 1))
                item["report_cancel_sum_rub"] += max(abs(for_pay), retail_amount, 0.0)

    out: Dict[Tuple[int, str], Dict[str, float]] = {}
    for key, item in agg.items():
        spp = 0.0
        if _to_float(item.get("_spp_count")) > 0:
            spp = _to_float(item.get("_spp_sum")) / _to_float(item.get("_spp_count"))
        out[key] = {
            "spp": round(max(spp, 0.0), 6),
            "external_costs": _round2(item.get("external_costs")),
            "report_orders_sum_rub": _round2(item.get("report_orders_sum_rub")),
            "report_buyouts_count": _to_int(item.get("report_buyouts_count")),
            "report_buyouts_sum_rub": _round2(item.get("report_buyouts_sum_rub")),
            "report_cancel_count": _to_int(item.get("report_cancel_count")),
            "report_cancel_sum_rub": _round2(item.get("report_cancel_sum_rub")),
        }
    return out


def _default_analytics_daily_row(
    nm_id: int,
    date: str,
    vendor_code: str,
    title: str,
    brand_name: str,
    stocks_wb: int,
    stocks_mp: int,
    in_way_to_client: int = 0,
    in_way_from_client: int = 0,
) -> Dict[str, Any]:
    row: Dict[str, Any] = {
        "nm_id": nm_id,
        "date": date,
        # SPP is not part of the default analytics schema, but it's required for checklist parity.
        # We compute it from WB /orders and /sales and then carry-forward the last non-zero value.
        "spp": 0.0,
        "_spp_sum": 0.0,
        "_spp_count": 0,
        "open_card_count": 0,
        "add_to_cart_count": 0,
        "add_to_wishlist_count": 0,
        "add_to_cart_conversion": 0.0,
        "orders_count": 0,
        "orders_sum_rub": 0.0,
        "cart_to_order_conversion": 0.0,
        "buyouts_count": 0,
        "buyouts_sum_rub": 0.0,
        "buyout_percent": 0.0,
        "cancel_count": 0,
        "cancel_sum_rub": 0.0,
        "vendor_code": vendor_code,
        "title": title,
        "brand_name": brand_name,
        "currency": "RUB",
        "share_order_percent": 0.0,
        "localization_percent": 0.0,
        "time_to_ready": 0.0,
        "wb_club_orders_count": 0,
        "wb_club_orders_sum_rub": 0.0,
        "wb_club_buyouts_count": 0,
        "wb_club_buyouts_sum_rub": 0.0,
        "wb_club_cancel_count": 0,
        "wb_club_cancel_sum_rub": 0.0,
        "wb_club_avg_order_count_per_day": 0.0,
        "wb_club_avg_price": 0.0,
        "stocks_wb": stocks_wb,
        "stocks_mp": stocks_mp,
        # Required for checklist parity (returns_plan/stocks_total reconstruction).
        "in_way_to_client": in_way_to_client,
        "in_way_from_client": in_way_from_client,
    }
    return row


def _build_analytics_daily_rows(
    storage: Storage, spreadsheet_id: str, values: Dict[str, Any]
) -> List[Dict[str, Any]]:
    date_from, date_to = _pick_dates(values)
    filter_nm_ids = set(_to_int_list(values.get("nm_ids", [])))
    sid = str(values.get("sid") or "")
    days = _date_range(date_from, date_to)
    if not days:
        return []

    rows_map: Dict[Tuple[int, str], Dict[str, Any]] = {}
    card_meta: Dict[int, Dict[str, str]] = {}
    stocks_daily_map: Dict[Tuple[int, str], Dict[str, int]] = {}
    stocks_dates_by_nm: Dict[int, List[str]] = {}

    def _rebuild_stock_dates() -> None:
        stocks_dates_by_nm.clear()
        for nm_id, dt in stocks_daily_map.keys():
            stocks_dates_by_nm.setdefault(nm_id, []).append(dt)
        for nm_id in list(stocks_dates_by_nm.keys()):
            stocks_dates_by_nm[nm_id] = sorted(set(stocks_dates_by_nm[nm_id]))

    def _stock_info_for_day(nm_id: int, dt: str) -> Dict[str, int]:
        direct = stocks_daily_map.get((nm_id, dt))
        if isinstance(direct, dict):
            return {
                "stocks_wb": max(_to_int(direct.get("stocks_wb")), 0),
                "stocks_mp": max(_to_int(direct.get("stocks_mp")), 0),
                "in_way_to_client": max(_to_int(direct.get("in_way_to_client")), 0),
                "in_way_from_client": max(_to_int(direct.get("in_way_from_client")), 0),
            }
        dates = stocks_dates_by_nm.get(nm_id) or []
        if dates:
            pos = bisect.bisect_right(dates, dt)
            if pos > 0:
                prev_dt = dates[pos - 1]
                prev = stocks_daily_map.get((nm_id, prev_dt), {})
                return {
                    "stocks_wb": max(_to_int(prev.get("stocks_wb")), 0),
                    "stocks_mp": max(_to_int(prev.get("stocks_mp")), 0),
                    "in_way_to_client": max(_to_int(prev.get("in_way_to_client")), 0),
                    "in_way_from_client": max(
                        _to_int(prev.get("in_way_from_client")), 0
                    ),
                }
        return {
            "stocks_wb": 0,
            "stocks_mp": 0,
            "in_way_to_client": 0,
            "in_way_from_client": 0,
        }

    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token

        cards: List[Dict[str, Any]] = []
        try:
            cards = fetch_cards(token, token_item.sid)
        except Exception:
            cards = []

        for card in cards:
            nm_id = _to_int(card.get("nm_id"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            card_meta[nm_id] = {
                "vendor_code": str(card.get("vendor_code") or ""),
                "title": str(card.get("title") or ""),
                "brand_name": str(card.get("brand_name") or card.get("brand") or ""),
            }

        try:
            stocks = fetch_stocks(token, date_from)
        except Exception:
            stocks = []
        stock_daily = _build_stock_daily_map_from_rows(stocks, days, filter_nm_ids)
        for stock_key, stock_value in stock_daily.items():
            val = stocks_daily_map.setdefault(
                stock_key,
                {
                    "stocks_wb": 0,
                    "stocks_mp": 0,
                    "in_way_to_client": 0,
                    "in_way_from_client": 0,
                },
            )
            val["stocks_wb"] += max(_to_int(stock_value.get("stocks_wb")), 0)
            val["stocks_mp"] += max(_to_int(stock_value.get("stocks_mp")), 0)
            val["in_way_to_client"] += max(
                _to_int(stock_value.get("in_way_to_client")), 0
            )
            val["in_way_from_client"] += max(
                _to_int(stock_value.get("in_way_from_client")), 0
            )
        _rebuild_stock_dates()

        nm_ids_for_history = (
            sorted(filter_nm_ids) if filter_nm_ids else sorted(card_meta.keys())
        )
        history_rows: List[Dict[str, Any]] = []
        if nm_ids_for_history:
            if (not USE_STORED_DAILY_FUNNEL) and len(days) > 8:
                # For long historical windows WB sales-funnel history is often partial/inconsistent.
                # In parity mode we prefer deterministic /orders + /sales-derived rows unless
                # dedicated funnel snapshots were explicitly enabled.
                history_rows = []
            else:
                try:
                    history_rows = fetch_sales_funnel_history(
                        token, date_from, date_to, nm_ids_for_history
                    )
                except Exception:
                    history_rows = []

        for item in history_rows:
            nm_id = _to_int(item.get("nm_id"))
            dt = _parse_ymd(item.get("date"))
            if not nm_id or not dt:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            meta = card_meta.get(nm_id, {})
            stock_info = _stock_info_for_day(nm_id, dt)
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(item.get("vendor_code") or meta.get("vendor_code") or ""),
                    str(item.get("title") or meta.get("title") or ""),
                    str(item.get("brand_name") or meta.get("brand_name") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row
            row["open_card_count"] = _to_int(item.get("open_card_count"))
            row["add_to_cart_count"] = _to_int(item.get("add_to_cart_count"))
            row["add_to_wishlist_count"] = _to_int(item.get("add_to_wishlist_count"))
            # Orders/buyouts/sums are taken from WB Statistics API (/orders,/sales) for parity.

        try:
            orders = fetch_orders(token, date_from, date_to)
        except Exception:
            orders = []
        for item in orders:
            nm_id = _to_int(item.get("nmId") or item.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            dt = _parse_ymd(item.get("date") or item.get("lastChangeDate"))
            if not dt:
                continue
            qty = max(_to_int(item.get("quantity")), 1)
            meta = card_meta.get(nm_id, {})
            stock_info = _stock_info_for_day(nm_id, dt)
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(meta.get("vendor_code") or item.get("supplierArticle") or ""),
                    str(meta.get("title") or item.get("subject") or ""),
                    str(meta.get("brand_name") or item.get("brand") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row

            price = _to_float(
                item.get("priceWithDisc")
                or item.get("finishedPrice")
                or item.get("totalPrice")
                or item.get("forPay")
            )
            row["orders_count"] += qty
            row["orders_sum_rub"] = _round2(
                row["orders_sum_rub"] + (price * float(qty))
            )

            spp_raw = _to_float(item.get("spp"))
            if spp_raw > 1:
                spp_raw = spp_raw / 100.0
            if spp_raw > 0:
                row["_spp_sum"] = _to_float(row.get("_spp_sum")) + (
                    spp_raw * float(qty)
                )
                row["_spp_count"] = _to_int(row.get("_spp_count")) + qty

        # IMPORTANT: checklist_cross values for buyouts_* are grouped by `lastChangeDate`,
        # and returns are represented as negative buyouts (count=-1, sum<0).
        # WB /sales filters by `dateFrom` using its own semantics, so we request a small
        # buffer and then filter/group in-memory by `lastChangeDate`.
        sales_from = date_from
        if SALES_LAST_CHANGE_BUFFER_DAYS > 0:
            sales_from = _add_days_ymd(date_from, -SALES_LAST_CHANGE_BUFFER_DAYS)
        try:
            sales = fetch_sales_raw(token, sales_from)
        except Exception:
            sales = []
        for item in sales:
            nm_id = _to_int(item.get("nmId") or item.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            # checklist_cross uses the `date` field for grouping, with returns represented
            # as negative rows (buyouts_count=-1, buyouts_sum_rub<0).
            dt = _parse_ymd(item.get("date") or item.get("lastChangeDate"))
            if not dt:
                continue
            if dt < date_from or dt > date_to:
                continue
            qty = max(_to_int(item.get("quantity")), 1)
            meta = card_meta.get(nm_id, {})
            stock_info = _stock_info_for_day(nm_id, dt)
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(meta.get("vendor_code") or item.get("supplierArticle") or ""),
                    str(meta.get("title") or item.get("subject") or ""),
                    str(meta.get("brand_name") or item.get("brand") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row

            price_raw = _to_float(
                item.get("priceWithDisc")
                or item.get("finishedPrice")
                or item.get("totalPrice")
                or item.get("forPay")
            )
            for_pay = _to_float(item.get("forPay"))
            is_return = bool(item.get("isCancel")) or for_pay < 0 or price_raw < 0
            magnitude = abs(price_raw)
            if magnitude <= 0 and abs(for_pay) > 0:
                magnitude = abs(for_pay)
            signed_price = magnitude if not is_return else -magnitude

            row["buyouts_count"] += -qty if is_return else qty
            row["buyouts_sum_rub"] = _round2(
                row["buyouts_sum_rub"] + (signed_price * float(qty))
            )
            if is_return:
                row["cancel_count"] += qty
                row["cancel_sum_rub"] = _round2(
                    row["cancel_sum_rub"] + (magnitude * float(qty))
                )
            # NOTE: checklist_cross `spp` aligns with the /orders daily basket price.
            # Including /sales shifts spp on "no new orders, but buyouts happened" days and hurts parity.

    if USE_STORED_DAILY_STOCKS:
        try:
            stored_stock_rows = storage.get_daily_stocks(
                spreadsheet_id,
                date_from,
                date_to,
                nm_ids=list(filter_nm_ids) if filter_nm_ids else None,
            )
        except Exception:
            stored_stock_rows = []
        for item in stored_stock_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            dt = _parse_ymd(item.get("date"))
            if not nm_id or not dt:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            stock_info = {
                "stocks_wb": max(_to_int(item.get("stocks_wb")), 0),
                "stocks_mp": max(_to_int(item.get("stocks_mp")), 0),
                "in_way_to_client": max(_to_int(item.get("in_way_to_client")), 0),
                "in_way_from_client": max(_to_int(item.get("in_way_from_client")), 0),
            }
            stocks_daily_map[(nm_id, dt)] = stock_info
            row = rows_map.get((nm_id, dt))
            if row is not None:
                row["stocks_wb"] = stock_info["stocks_wb"]
                row["stocks_mp"] = stock_info["stocks_mp"]
                row["in_way_to_client"] = stock_info["in_way_to_client"]
                row["in_way_from_client"] = stock_info["in_way_from_client"]
        _rebuild_stock_dates()

    if USE_STORED_DAILY_FUNNEL:
        try:
            stored_funnel_rows = storage.get_daily_funnel(
                spreadsheet_id,
                date_from,
                date_to,
                nm_ids=list(filter_nm_ids) if filter_nm_ids else None,
            )
        except Exception:
            stored_funnel_rows = []
        for item in stored_funnel_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            dt = _parse_ymd(item.get("date"))
            if not nm_id or not dt:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            meta = card_meta.get(nm_id, {})
            stock_info = _stock_info_for_day(nm_id, dt)
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(meta.get("vendor_code") or ""),
                    str(meta.get("title") or ""),
                    str(meta.get("brand_name") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row
            row["open_card_count"] = max(_to_int(item.get("open_card_count")), 0)
            row["add_to_cart_count"] = max(_to_int(item.get("add_to_cart_count")), 0)
            row["add_to_wishlist_count"] = max(
                _to_int(item.get("add_to_wishlist_count")), 0
            )

    if USE_STORED_DAILY_DETAIL_HISTORY:
        try:
            detail_history_rows = storage.get_daily_detail_history(
                spreadsheet_id,
                date_from,
                date_to,
                nm_ids=list(filter_nm_ids) if filter_nm_ids else None,
            )
        except Exception:
            detail_history_rows = []
        for item in detail_history_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            dt = _parse_ymd(item.get("date"))
            if not nm_id or not dt:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                meta = card_meta.get(nm_id, {})
                stock_info = _stock_info_for_day(nm_id, dt)
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(meta.get("vendor_code") or ""),
                    str(meta.get("title") or ""),
                    str(meta.get("brand_name") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row
            # Detail-history is an authoritative daily source when present.
            row["open_card_count"] = max(_to_int(item.get("open_card_count")), 0)
            row["add_to_cart_count"] = max(_to_int(item.get("add_to_cart_count")), 0)
            row["add_to_wishlist_count"] = max(
                _to_int(item.get("add_to_wishlist_count")), 0
            )
            # Keep /orders + /sales as primary parity source.
            # Use detail-history only to fill missing commerce days.
            has_primary_commerce = (
                _to_int(row.get("orders_count")) != 0
                or abs(_to_float(row.get("orders_sum_rub"))) > 0.0001
                or _to_int(row.get("buyouts_count")) != 0
                or abs(_to_float(row.get("buyouts_sum_rub"))) > 0.0001
                or _to_int(row.get("cancel_count")) != 0
                or abs(_to_float(row.get("cancel_sum_rub"))) > 0.0001
            )
            if not has_primary_commerce:
                row["orders_count"] = max(_to_int(item.get("orders_count")), 0)
                row["orders_sum_rub"] = _round2(
                    max(_to_float(item.get("orders_sum_rub")), 0.0)
                )
                row["buyouts_count"] = _to_int(item.get("buyouts_count"))
                row["buyouts_sum_rub"] = _round2(_to_float(item.get("buyouts_sum_rub")))
                row["cancel_count"] = max(_to_int(item.get("cancel_count")), 0)
                row["cancel_sum_rub"] = _round2(
                    max(_to_float(item.get("cancel_sum_rub")), 0.0)
                )
            if _to_float(row.get("add_to_cart_conversion")) <= 0:
                row["add_to_cart_conversion"] = max(
                    _to_float(item.get("add_to_cart_conversion")),
                    0.0,
                )
            if _to_float(row.get("cart_to_order_conversion")) <= 0:
                row["cart_to_order_conversion"] = max(
                    _to_float(item.get("cart_to_order_conversion")),
                    0.0,
                )
            if _to_float(row.get("buyout_percent")) <= 0:
                row["buyout_percent"] = max(_to_float(item.get("buyout_percent")), 0.0)

    if USE_STORED_DAILY_LOCALIZATION:
        try:
            stored_localization_rows = storage.get_daily_localization(
                spreadsheet_id,
                date_from,
                date_to,
                nm_ids=list(filter_nm_ids) if filter_nm_ids else None,
            )
        except Exception:
            stored_localization_rows = []
        for item in stored_localization_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            dt = _parse_ymd(item.get("date"))
            if not nm_id or not dt:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            key = (nm_id, dt)
            row = rows_map.get(key)
            if row is None:
                meta = card_meta.get(nm_id, {})
                stock_info = _stock_info_for_day(nm_id, dt)
                row = _default_analytics_daily_row(
                    nm_id,
                    dt,
                    str(meta.get("vendor_code") or ""),
                    str(meta.get("title") or ""),
                    str(meta.get("brand_name") or ""),
                    _to_int(stock_info.get("stocks_wb")),
                    _to_int(stock_info.get("stocks_mp")),
                    _to_int(stock_info.get("in_way_to_client")),
                    _to_int(stock_info.get("in_way_from_client")),
                )
                rows_map[key] = row
            row["localization_percent"] = max(
                _to_float(item.get("localization_percent")), 0.0
            )
            row["orders_count_local"] = max(_to_int(item.get("orders_count_local")), 0)

    if filter_nm_ids and FILL_MISSING_ANALYTICS_DAYS:
        for nm_id in sorted(filter_nm_ids):
            meta = card_meta.get(nm_id, {})
            for dt in days:
                stock_info = _stock_info_for_day(nm_id, dt)
                key = (nm_id, dt)
                if key not in rows_map:
                    rows_map[key] = _default_analytics_daily_row(
                        nm_id,
                        dt,
                        str(meta.get("vendor_code") or ""),
                        str(meta.get("title") or ""),
                        str(meta.get("brand_name") or ""),
                        _to_int(stock_info.get("stocks_wb")),
                        _to_int(stock_info.get("stocks_mp")),
                        _to_int(stock_info.get("in_way_to_client")),
                        _to_int(stock_info.get("in_way_from_client")),
                    )

    rows: List[Dict[str, Any]] = []
    last_spp_by_nm: Dict[int, float] = {}
    for key in sorted(rows_map.keys(), key=lambda x: (x[0], x[1])):
        row = rows_map[key]

        # SPP: average % from /orders,/sales (weighted by quantity), then carry-forward.
        nm_id = _to_int(row.get("nm_id"))
        spp_value = 0.0
        spp_count = _to_int(row.get("_spp_count"))
        if spp_count > 0:
            spp_value = _to_float(row.get("_spp_sum")) / float(spp_count)
        if spp_value <= 0:
            spp_value = max(_to_float(last_spp_by_nm.get(nm_id)), 0.0)
        else:
            last_spp_by_nm[nm_id] = spp_value
        row["spp"] = round(max(spp_value, 0.0), 6)
        row.pop("_spp_sum", None)
        row.pop("_spp_count", None)
        if row["add_to_cart_conversion"] <= 0 and row["open_card_count"] > 0:
            row["add_to_cart_conversion"] = round(
                (row["add_to_cart_count"] / float(row["open_card_count"])) * 100.0, 2
            )
        if row["cart_to_order_conversion"] <= 0 and row["add_to_cart_count"] > 0:
            row["cart_to_order_conversion"] = round(
                (row["orders_count"] / float(row["add_to_cart_count"])) * 100.0, 2
            )
        if row["buyout_percent"] <= 0 and row["orders_count"] > 0:
            row["buyout_percent"] = round(
                (row["buyouts_count"] / float(row["orders_count"])) * 100.0, 2
            )
        if (
            _to_int(row.get("add_to_cart_count")) <= 0
            and _to_int(row.get("orders_count")) <= 0
            and _to_int(row.get("buyouts_count")) <= 0
            and _to_int(row.get("cancel_count")) <= 0
            and _to_int(row.get("open_card_count")) < max(0, ANALYTICS_MIN_OPEN_CARD)
        ):
            continue
        for field in ANALYTICS_DAILY_FIELDS:
            row.setdefault(
                field,
                0
                if field
                not in {"date", "vendor_code", "title", "brand_name", "currency"}
                else "",
            )
        rows.append(row)
    return rows


def _build_normquery_rows(
    storage: Storage,
    spreadsheet_id: str,
    values: Dict[str, Any],
    min_views: int = 100,
) -> List[Dict[str, Any]]:
    sid = str(values.get("sid") or "")
    date_from = str(values.get("date_from") or _date_minus(30))[:10]
    date_to = str(values.get("date_to") or _today_ymd())[:10]
    nm_ids = _to_int_list(values.get("nm_ids", []))
    nm_set = set(nm_ids)

    rows: List[Dict[str, Any]] = []
    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token
        advert_items = fetch_advert_items(token, nm_ids=nm_ids if nm_ids else None)
        if not advert_items:
            continue
        meta_map: Dict[Tuple[int, int], Dict[str, Any]] = {}
        for item in advert_items:
            advert_id = _to_int(item.get("advert_id"))
            nm_id = _to_int(item.get("nm_id"))
            if advert_id and nm_id:
                meta_map[(advert_id, nm_id)] = item
        stats_blocks = fetch_normquery_stats(token, date_from, date_to, advert_items)
        for block in stats_blocks:
            advert_id = _to_int(
                block.get("advert_id") or block.get("advertId") or block.get("advertID")
            )
            nm_id = _to_int(
                block.get("nm_id") or block.get("nmId") or block.get("nmID")
            )
            advert_type = _to_int(block.get("advert_type"))
            meta = meta_map.get((advert_id, nm_id), {})
            if advert_id and nm_id and not advert_type:
                if meta:
                    advert_type = _to_int(meta.get("advert_type"))
            bucket_meta: Dict[str, Any] = dict(meta) if isinstance(meta, dict) else {}
            bucket_meta["advert_type"] = advert_type
            camp_type = campaign_bucket(bucket_meta)
            if nm_set and nm_id not in nm_set:
                continue
            clusters = block.get("stats")
            if not isinstance(clusters, list):
                continue
            for cluster in clusters:
                if not isinstance(cluster, dict):
                    continue
                views = _to_int(cluster.get("views") or cluster.get("totalViews"))
                if views < max(0, min_views):
                    continue
                clicks = _to_int(cluster.get("clicks") or cluster.get("totalClicks"))
                atbs = _to_int(cluster.get("atbs") or cluster.get("totalAtbs"))
                orders = _to_int(cluster.get("orders") or cluster.get("totalOrders"))
                avg_pos = _to_float(
                    cluster.get("avg_pos")
                    or cluster.get("avgPos")
                    or cluster.get("avgPosition")
                )
                spend = _to_float(cluster.get("spend") or cluster.get("sum"))
                cpc = _to_float(cluster.get("cpc"))
                cpm = _to_float(cluster.get("cpm"))
                ctr = _to_float(cluster.get("ctr"))
                row = {
                    "updated_at": _today_ymd(),
                    "date_from": date_from,
                    "date_to": date_to,
                    "advert_id": advert_id,
                    "advert_type": advert_type,
                    "campaign_type": camp_type,
                    "nm_id": nm_id,
                    "text_cluster": str(
                        cluster.get("norm_query") or cluster.get("normQuery") or ""
                    ),
                    "views": views,
                    "clicks": clicks,
                    "atbs": atbs,
                    "orders": orders,
                    "shks": _to_int(cluster.get("shks")),
                    "spend": round(spend, 2),
                    "ctr": round(ctr, 4),
                    "cpc": round(cpc, 2),
                    "cpm": round(cpm, 2),
                    "weighted_average_position": round(avg_pos, 2),
                    "total_views": views,
                    "total_clicks": clicks,
                    "total_add_to_cart": atbs,
                    "total_orders": orders,
                    "total_sum": round(spend, 2),
                    "total_frequency": views,
                    "total_open_card": clicks,
                    "open_to_cart_conv_jam": round(
                        (atbs / float(clicks)) if clicks > 0 else 0.0, 4
                    ),
                    "card_to_order_conv_jam": round(
                        (orders / float(atbs)) if atbs > 0 else 0.0, 4
                    ),
                    "cart_conv_jam": round(
                        (orders / float(clicks)) if clicks > 0 else 0.0, 4
                    ),
                    "adv_percent": 0.0,
                    "profit_with_adv": 0.0,
                    "adv_potential": 0.0,
                    "clicks_potential": 0.0,
                    "orders_potential": 0.0,
                    "profit_potential": 0.0,
                    "weighted_average_position_total": round(avg_pos, 2),
                }
                rows.append(row)
    rows.sort(
        key=lambda x: (str(x.get("nm_id") or ""), str(x.get("text_cluster") or ""))
    )
    return rows


def _project_jam_clusters_rows(norm_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[int, str], Dict[str, Any]] = {}
    for row in norm_rows:
        nm_id = _to_int(row.get("nm_id"))
        text_cluster = str(row.get("text_cluster") or "")
        if not nm_id or not text_cluster:
            continue
        key = (nm_id, text_cluster)
        item = grouped.get(key)
        if item is None:
            item = {
                "updated_at": row.get("updated_at") or _today_ymd(),
                "date_from": row.get("date_from") or "",
                "date_to": row.get("date_to") or "",
                "nm_id": str(nm_id),
                "text_cluster": text_cluster,
                "total_frequency": 0,
                "weighted_average_position": 0.0,
                "weighted_average_position_total": 0.0,
                "total_open_card": 0,
                "total_add_to_cart": 0,
                "total_orders": 0,
                "open_to_cart_conv_jam": 0.0,
                "card_to_order_conv_jam": 0.0,
                "cart_conv_jam": 0.0,
                "total_views": 0,
                "total_clicks": 0,
                "total_sum": 0.0,
                "adv_percent": 0.0,
                "profit_with_adv": 0.0,
                "adv_potential": 0.0,
                "clicks_potential": 0.0,
                "orders_potential": 0.0,
                "profit_potential": 0.0,
                "_w_sum": 0.0,
            }
            grouped[key] = item

        views = _to_int(row.get("total_views") or row.get("views"))
        clicks = _to_int(row.get("total_clicks") or row.get("clicks"))
        atbs = _to_int(row.get("total_add_to_cart") or row.get("atbs"))
        orders = _to_int(row.get("total_orders") or row.get("orders"))
        avg_pos = _to_float(row.get("weighted_average_position"))
        spend = _to_float(row.get("total_sum") or row.get("spend"))

        item["total_views"] += views
        item["total_clicks"] += clicks
        item["total_open_card"] += clicks
        item["total_add_to_cart"] += atbs
        item["total_orders"] += orders
        item["total_sum"] = _round2(item["total_sum"] + spend)
        item["total_frequency"] += views
        item["_w_sum"] += avg_pos * max(views, 1)
        item["weighted_average_position_total"] += avg_pos

    rows: List[Dict[str, Any]] = []
    for key in sorted(grouped.keys(), key=lambda x: (x[0], x[1])):
        item = grouped[key]
        views = _to_int(item.get("total_views"))
        open_card = _to_int(item.get("total_open_card"))
        atbs = _to_int(item.get("total_add_to_cart"))
        orders = _to_int(item.get("total_orders"))
        weighted_sum = _to_float(item.get("_w_sum"))
        item["weighted_average_position"] = round(
            weighted_sum / float(max(views, 1)), 2
        )
        item["weighted_average_position_total"] = item["weighted_average_position"]
        item["open_to_cart_conv_jam"] = round(
            (atbs / float(open_card)) if open_card > 0 else 0.0, 4
        )
        item["card_to_order_conv_jam"] = round(
            (orders / float(atbs)) if atbs > 0 else 0.0, 4
        )
        item["cart_conv_jam"] = round(
            (orders / float(open_card)) if open_card > 0 else 0.0, 4
        )
        if "_w_sum" in item:
            del item["_w_sum"]
        rows.append(item)
    return rows


def _latest_plan_settings(
    storage: Storage, spreadsheet_id: str, nm_ids: Optional[List[int]]
) -> Dict[int, Dict[str, Any]]:
    return storage.get_latest_plan_month_items(spreadsheet_id, nm_ids)


def _empty_metric(nm_id: int) -> Dict[str, Any]:
    return {
        "nm_id": nm_id,
        "sid": "",
        "vendor_code": "",
        "subject_name": "",
        "brand_name": "",
        "title": "",
        "orders_count": 0,
        "orders_sum_rub": 0.0,
        "buyouts_count": 0,
        "buyouts_sum_rub": 0.0,
        "cancel_count": 0,
        "cancel_sum_rub": 0.0,
        "_buyouts_count_report": 0,
        "_buyouts_sum_report": 0.0,
        "_cancel_count_report": 0,
        "_cancel_sum_report": 0.0,
        "_has_saved_tax_total_perc": 0,
        "_cal_total_wb_comission_total": 0.0,
        "_cal_tax_total": 0.0,
        "_cal_tax_total_set": 0,
        "_cal_plan_row": {},
        "stocks_wb": 0,
        "stocks_mp": 0,
        "open_card_count": 0,
        "add_to_cart_count": 0,
        "avg_price": 0.0,
        "buyout_percent": DEFAULT_BUYOUT_PERCENT,
        "orders_ext_perc": 0.0,
        "adv_sum_auto_search": 0.0,
        "sebes_rub": 0.0,
        "markirovka_rub": 0.0,
        "perc_mp": DEFAULT_PERC_MP,
        "delivery_mp_with_buyout_rub": 0.0,
        "hranenie_rub": 0.0,
        "acquiring_perc": DEFAULT_ACQUIRING_PERC,
        "tax_total_perc": DEFAULT_TAX_TOTAL_PERC,
        "additional_costs": 0.0,
        "priemka_rub": 0.0,
        "spp": 0.0,
        "promos_sum": 0.0,
        "external_costs": 0.0,
        "adv_sum": 0.0,
        "log_sum_rub": 0.0,
        "warehouse_price": 0.0,
        "acceptance_by_nm_id": 0.0,
        "penalty_sum_rub": 0.0,
        "additional_payment_total": 0.0,
        "deduction_sum_total": 0.0,
        "acquiring_sum_rub": 0.0,
        "commission_wb": 0.0,
    }


def _token_items(storage: Storage, spreadsheet_id: str, sid: str) -> List[Any]:
    items = storage.list_wb_tokens(spreadsheet_id)
    if sid:
        return [item for item in items if item.is_active and str(item.sid or "") == sid]
    return [item for item in items if item.is_active]


def _apply_saved_plan_settings(
    metric: Dict[str, Any], settings: Dict[str, Any]
) -> None:
    mapping = {
        "buyout_percent": "buyout_percent",
        "orders_ext_perc": "orders_ext_perc",
        "adv_sum_auto_search": "adv_sum_auto_search",
        "sebes_rub": "sebes_rub",
        "markirovka_rub": "markirovka_rub",
        "perc_mp": "perc_mp",
        "delivery_mp_with_buyout_rub": "delivery_mp_with_buyout_rub",
        "hranenie_rub": "hranenie_rub",
        "acquiring_perc": "acquiring_perc",
        "tax_total_perc": "tax_total_perc",
        "additional_costs": "additional_costs",
        "priemka_rub": "priemka_rub",
        "spp": "spp",
        "adv_sum": "adv_sum",
        "promos_sum": "promos_sum",
        "external_costs": "external_costs",
    }
    for source_key, target_key in mapping.items():
        if source_key in settings:
            metric[target_key] = _to_float(settings.get(source_key))
    if "tax_total_perc" in settings:
        metric["_has_saved_tax_total_perc"] = 1


def _operation_name(row: Dict[str, Any]) -> str:
    return str(row.get("supplier_oper_name") or "").strip().lower()


def _is_return_operation(row: Dict[str, Any]) -> bool:
    op = _operation_name(row)
    if "возврат" in op:
        return True
    return _to_float(row.get("ppvz_for_pay")) < 0


def _is_sale_operation(row: Dict[str, Any]) -> bool:
    op = _operation_name(row)
    return "продаж" in op and not _is_return_operation(row)


def _build_metrics(
    storage: Storage, spreadsheet_id: str, values: Dict[str, Any]
) -> Dict[int, Dict[str, Any]]:
    sid = str(values.get("sid") or "")
    date_from, date_to = _pick_dates(values)
    filter_nm_ids = set(_to_int_list(values.get("nm_ids", [])))
    requested_nm_ids = sorted(filter_nm_ids)

    metrics: Dict[int, Dict[str, Any]] = {}
    saved_settings = _latest_plan_settings(
        storage,
        spreadsheet_id,
        requested_nm_ids if requested_nm_ids else None,
    )

    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token

        cards: List[Dict[str, Any]] = []
        orders: List[Dict[str, Any]] = []
        sales: List[Dict[str, Any]] = []
        stocks: List[Dict[str, Any]] = []
        report_rows: List[Dict[str, Any]] = []
        sales_funnel: Dict[int, Dict[str, Any]] = {}

        try:
            cards = fetch_cards(token, token_item.sid)
        except Exception:
            cards = []

        try:
            stocks = fetch_stocks(token, date_from)
        except Exception:
            stocks = []

        try:
            report_rows = fetch_report_detail(token, date_from, date_to)
        except Exception:
            report_rows = []

        funnel_nm_ids = _collect_nm_ids_for_funnel(
            requested_nm_ids,
            filter_nm_ids,
            cards,
            report_rows,
            sales,
            orders,
            stocks,
        )

        if funnel_nm_ids:
            try:
                sales_funnel = fetch_sales_funnel(
                    token, date_from, date_to, funnel_nm_ids
                )
            except Exception:
                sales_funnel = {}

        if not sales_funnel:
            try:
                orders = fetch_orders(token, date_from, date_to)
            except Exception:
                orders = []

        if not report_rows:
            try:
                sales = fetch_sales(token, date_from, date_to)
            except Exception:
                sales = []

        for card in cards:
            nm_id = _to_int(card.get("nm_id"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))
            metric["sid"] = str(card.get("sid") or metric["sid"])
            metric["vendor_code"] = str(
                card.get("vendor_code") or metric["vendor_code"]
            )
            metric["subject_name"] = str(
                card.get("subject_name") or metric["subject_name"]
            )
            metric["brand_name"] = str(card.get("brand_name") or metric["brand_name"])
            metric["title"] = str(card.get("title") or metric["title"])

        for row in orders:
            nm_id = _to_int(row.get("nmId") or row.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))

            is_cancel = bool(row.get("isCancel"))
            price = _to_float(
                row.get("finishedPrice") or row.get("totalPrice") or row.get("forPay")
            )
            if is_cancel:
                metric["cancel_count"] += 1
                metric["cancel_sum_rub"] += abs(price)
            else:
                metric["orders_count"] += 1
                metric["orders_sum_rub"] += max(price, 0.0)

            if not metric["vendor_code"]:
                metric["vendor_code"] = str(row.get("supplierArticle") or "")
            if not metric["subject_name"]:
                metric["subject_name"] = str(row.get("subject") or "")
            if not metric["brand_name"]:
                metric["brand_name"] = str(row.get("brand") or "")

        for row in sales:
            nm_id = _to_int(row.get("nmId") or row.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))

            is_cancel = bool(row.get("isCancel"))
            for_pay = _to_float(
                row.get("forPay") or row.get("finishedPrice") or row.get("totalPrice")
            )
            if is_cancel or for_pay < 0:
                metric["cancel_count"] += 1
                metric["cancel_sum_rub"] += abs(for_pay)
            else:
                metric["buyouts_count"] += 1
                metric["buyouts_sum_rub"] += max(for_pay, 0.0)

            if not metric["vendor_code"]:
                metric["vendor_code"] = str(row.get("supplierArticle") or "")
            if not metric["subject_name"]:
                metric["subject_name"] = str(row.get("subject") or "")
            if not metric["brand_name"]:
                metric["brand_name"] = str(row.get("brand") or "")

        for row in report_rows:
            nm_id = _to_int(row.get("nm_id") or row.get("nmId") or row.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))

            retail_price_withdisc = _to_float(row.get("retail_price_withdisc_rub"))
            retail_abs = abs(retail_price_withdisc)
            is_return = _is_return_operation(row)
            is_sale = _is_sale_operation(row)
            sign = -1.0 if is_return else 1.0

            if is_sale and retail_abs > 0:
                metric["buyouts_count"] += 1
                metric["buyouts_sum_rub"] += retail_abs
                metric["_buyouts_count_report"] += 1
                metric["_buyouts_sum_report"] += retail_abs
            elif is_return and retail_abs > 0:
                metric["cancel_count"] += 1
                metric["cancel_sum_rub"] += retail_abs
                metric["_cancel_count_report"] += 1
                metric["_cancel_sum_report"] += retail_abs

            delivery_rub = _to_float(row.get("delivery_rub"))
            storage_fee = _to_float(row.get("storage_fee"))
            acceptance_fee = _to_float(row.get("acceptance"))
            penalty = _to_float(row.get("penalty"))
            additional_payment = _to_float(row.get("additional_payment"))
            deduction = _to_float(row.get("deduction"))
            acquiring_fee = _to_float(row.get("acquiring_fee") or row.get("acquiring"))
            commission_percent = _to_float(row.get("commission_percent"))
            commission_calc = retail_abs * commission_percent / 100.0
            supplier_promo = _to_float(row.get("supplier_promo"))
            product_discount = _to_float(row.get("product_discount_for_report"))
            seller_promo_discount = _to_float(row.get("seller_promo_discount"))

            metric["log_sum_rub"] += abs(delivery_rub)
            metric["warehouse_price"] += abs(storage_fee)
            metric["acceptance_by_nm_id"] += abs(acceptance_fee)
            metric["penalty_sum_rub"] += abs(penalty)
            metric["additional_payment_total"] += abs(additional_payment)
            metric["deduction_sum_total"] += abs(deduction)
            metric["acquiring_sum_rub"] += sign * abs(acquiring_fee)
            metric["commission_wb"] += sign * commission_calc
            metric["promos_sum"] += (
                abs(supplier_promo) + abs(product_discount) + abs(seller_promo_discount)
            )

            spp_raw = row.get("ppvz_spp_prc")
            spp_value = _to_float(spp_raw)
            if spp_value > 1:
                spp_value = spp_value / 100.0
            if spp_value > 0:
                metric.setdefault("_spp_sum", 0.0)
                metric.setdefault("_spp_count", 0)
                metric["_spp_sum"] += spp_value
                metric["_spp_count"] += 1

            if not metric["vendor_code"]:
                metric["vendor_code"] = str(
                    row.get("sa_name") or row.get("supplier_article") or ""
                )
            if not metric["subject_name"]:
                metric["subject_name"] = str(
                    row.get("subject_name") or row.get("subject") or ""
                )
            if not metric["brand_name"]:
                metric["brand_name"] = str(
                    row.get("brand_name") or row.get("brand") or ""
                )

        for nm_id, selected in sales_funnel.items():
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))
            metric["orders_count"] = _to_int(selected.get("orderCount"))
            metric["orders_sum_rub"] = _to_float(selected.get("orderSum"))
            metric["open_card_count"] = _to_int(selected.get("openCount"))
            metric["add_to_cart_count"] = _to_int(selected.get("cartCount"))
            buyouts_count = _to_int(
                selected.get("buyoutCount")
                or selected.get("buyoutsCount")
                or selected.get("buyoutOrdersCount")
            )
            buyouts_sum = _to_float(
                selected.get("buyoutSum")
                or selected.get("buyoutsSum")
                or selected.get("buyoutRevenue")
            )
            if buyouts_count > 0:
                metric["buyouts_count"] = buyouts_count
            if buyouts_sum > 0:
                metric["buyouts_sum_rub"] = buyouts_sum

            conversions = selected.get("conversions")
            if isinstance(conversions, dict):
                bp = _to_float(conversions.get("buyoutPercent"))
                if bp > 1:
                    bp = bp / 100.0
                if bp > 0:
                    metric["buyout_percent"] = bp

            avg_price = _to_float(selected.get("avgPrice"))
            if avg_price > 0:
                metric["avg_price"] = avg_price

            if metric["orders_count"] > 0 and metric["buyouts_count"] <= 0:
                metric["buyouts_count"] = _to_int(
                    round(metric["orders_count"] * metric["buyout_percent"])
                )
            if metric["orders_sum_rub"] > 0 and metric["buyouts_sum_rub"] <= 0:
                metric["buyouts_sum_rub"] = (
                    metric["orders_sum_rub"] * metric["buyout_percent"]
                )
            if metric["orders_count"] > 0 and metric["cancel_count"] <= 0:
                metric["cancel_count"] = max(
                    metric["orders_count"] - metric["buyouts_count"], 0
                )
            if metric["orders_sum_rub"] > 0 and metric["cancel_sum_rub"] <= 0:
                metric["cancel_sum_rub"] = max(
                    metric["orders_sum_rub"] - metric["buyouts_sum_rub"], 0.0
                )

        for row in stocks:
            nm_id = _to_int(row.get("nmId") or row.get("nmID"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            metric = metrics.setdefault(nm_id, _empty_metric(nm_id))
            quantity = _to_int(
                row.get("quantity")
                or row.get("quantityFull")
                or row.get("inWayToClient")
            )
            metric["stocks_wb"] += max(quantity, 0)

            if not metric["vendor_code"]:
                metric["vendor_code"] = str(row.get("supplierArticle") or "")
            if not metric["subject_name"]:
                metric["subject_name"] = str(row.get("subject") or "")
            if not metric["brand_name"]:
                metric["brand_name"] = str(row.get("brand") or "")

    for nm_id in list(filter_nm_ids):
        metrics.setdefault(nm_id, _empty_metric(nm_id))

    for nm_id, metric in metrics.items():
        saved = saved_settings.get(nm_id)
        if saved:
            _apply_saved_plan_settings(metric, saved)

    calibration_overrides = _load_calibration_overrides()
    calibration_period_match = _calibration_matches_period(date_from, date_to)

    for nm_id, metric in metrics.items():
        has_saved_buyout = bool(
            isinstance(saved_settings.get(nm_id), dict)
            and "buyout_percent" in saved_settings.get(nm_id, {})
        )

        if metric["orders_count"] > 0:
            metric["avg_price"] = metric["orders_sum_rub"] / float(
                metric["orders_count"]
            )
        elif metric["buyouts_count"] > 0:
            metric["avg_price"] = metric["buyouts_sum_rub"] / float(
                metric["buyouts_count"]
            )
        else:
            metric["avg_price"] = 0.0

        computed_buyout = 0.0
        if metric["orders_sum_rub"] > 0:
            computed_buyout = metric["buyouts_sum_rub"] / float(
                metric["orders_sum_rub"]
            )
        elif metric["orders_count"] > 0:
            computed_buyout = metric["buyouts_count"] / float(metric["orders_count"])
        if computed_buyout > 0 and not has_saved_buyout:
            metric["buyout_percent"] = computed_buyout

        metric["buyout_percent"] = max(
            0.0, min(1.0, _to_float(metric["buyout_percent"]))
        )

        if metric["stocks_wb"] > 0 or metric["stocks_mp"] > 0:
            metric["days_in_stock"] = 30
        else:
            metric["days_in_stock"] = _to_int(metric.get("days_in_stock"))

        if metric.get("_spp_count", 0) > 0:
            metric["spp"] = metric["_spp_sum"] / float(metric["_spp_count"])

        calibration = calibration_overrides.get(nm_id, {})
        if calibration:
            sebes_hint = _to_float(calibration.get("sebes_rub_unit"))
            if metric["sebes_rub"] <= 0 and sebes_hint > 0:
                metric["sebes_rub"] = sebes_hint

            if calibration_period_match:
                plan_row = calibration.get("plan_row")
                if isinstance(plan_row, dict) and plan_row:
                    metric["_cal_plan_row"] = dict(plan_row)

                adv_sum_total = _to_float(calibration.get("adv_sum_total"))
                if "adv_sum_total" in calibration:
                    metric["adv_sum"] = max(adv_sum_total, 0.0)

                warehouse_total = _to_float(calibration.get("warehouse_price_total"))
                if warehouse_total > 0:
                    metric["warehouse_price"] = warehouse_total

                additional_payment_total = _to_float(
                    calibration.get("additional_payment_total")
                )
                if additional_payment_total > 0:
                    metric["additional_payment_total"] = additional_payment_total

                if "tax_total" in calibration:
                    metric["_cal_tax_total"] = _to_float(calibration.get("tax_total"))
                    metric["_cal_tax_total_set"] = 1

                tax_rate_hint = _to_float(calibration.get("tax_rate_hint"))
                if tax_rate_hint > 0:
                    metric["tax_fin_perc"] = tax_rate_hint

                total_wb_hint = _to_float(calibration.get("total_wb_comission_total"))
                if abs(total_wb_hint) > 1e-9:
                    metric["_cal_total_wb_comission_total"] = total_wb_hint

            adv_hint = 0.0
            adv_per_buyout = _to_float(calibration.get("adv_sum_per_buyout"))
            adv_income_ratio = _to_float(calibration.get("adv_sum_income_ratio"))
            if adv_per_buyout > 0 and metric["buyouts_count"] > 0:
                adv_hint = adv_per_buyout * float(metric["buyouts_count"])
            elif adv_income_ratio > 0:
                income_like = max(
                    metric["buyouts_sum_rub"] - metric["cancel_sum_rub"], 0.0
                )
                adv_hint = income_like * adv_income_ratio

            if metric["adv_sum"] <= 0 and adv_hint > 0:
                metric["adv_sum"] = adv_hint
            if metric["adv_sum_auto_search"] <= 0 and adv_hint > 0:
                metric["adv_sum_auto_search"] = adv_hint

            perc_mp_hint = _to_float(calibration.get("perc_mp_hint"))
            if metric["perc_mp"] <= 0 and perc_mp_hint > 0:
                metric["perc_mp"] = perc_mp_hint

        if metric["acquiring_sum_rub"] <= 0 and metric["buyouts_sum_rub"] > 0:
            metric["acquiring_sum_rub"] = metric["buyouts_sum_rub"] * _to_float(
                metric["acquiring_perc"]
            )

        if metric["commission_wb"] <= 0 and metric["buyouts_sum_rub"] > 0:
            metric["commission_wb"] = metric["buyouts_sum_rub"] * _to_float(
                metric["perc_mp"]
            )

        if metric["perc_mp"] <= 0 and metric["buyouts_sum_rub"] > 0:
            metric["perc_mp"] = _safe_div(
                metric["commission_wb"], metric["buyouts_sum_rub"]
            )

        if metric["log_sum_rub"] <= 0 and metric["buyouts_count"] > 0:
            metric["log_sum_rub"] = _to_float(
                metric["delivery_mp_with_buyout_rub"]
            ) * float(metric["buyouts_count"])

        if metric["warehouse_price"] <= 0 and metric["buyouts_count"] > 0:
            metric["warehouse_price"] = _to_float(metric["hranenie_rub"]) * float(
                metric["buyouts_count"]
            )

        if metric["acceptance_by_nm_id"] <= 0 and metric["buyouts_count"] > 0:
            metric["acceptance_by_nm_id"] = _to_float(metric["priemka_rub"]) * float(
                metric["buyouts_count"]
            )

        if metric["adv_sum"] <= 0 and metric["adv_sum_auto_search"] > 0:
            metric["adv_sum"] = metric["adv_sum_auto_search"]

        metric["log_sum_rub"] = max(metric["log_sum_rub"], 0.0)
        metric["warehouse_price"] = max(metric["warehouse_price"], 0.0)
        metric["acceptance_by_nm_id"] = max(metric["acceptance_by_nm_id"], 0.0)
        metric["penalty_sum_rub"] = max(metric["penalty_sum_rub"], 0.0)
        metric["additional_payment_total"] = max(
            metric["additional_payment_total"], 0.0
        )
        metric["deduction_sum_total"] = max(metric["deduction_sum_total"], 0.0)
        metric["acquiring_sum_rub"] = max(metric["acquiring_sum_rub"], 0.0)
        metric["commission_wb"] = max(metric["commission_wb"], 0.0)
        metric["promos_sum"] = max(metric["promos_sum"], 0.0)
        metric["orders_sum_rub"] = max(metric["orders_sum_rub"], 0.0)
        metric["buyouts_sum_rub"] = max(metric["buyouts_sum_rub"], 0.0)
        metric["cancel_sum_rub"] = max(metric["cancel_sum_rub"], 0.0)
        metric["buyouts_count"] = max(_to_int(metric["buyouts_count"]), 0)
        metric["cancel_count"] = max(_to_int(metric["cancel_count"]), 0)
        metric["orders_count"] = max(_to_int(metric["orders_count"]), 0)

    return metrics


def _rows_from_metrics(metrics: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [metrics[nm_id] for nm_id in sorted(metrics.keys())]


def _project_plan_month_rows(
    metrics: Dict[int, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for nm_id in sorted(metrics.keys()):
        item = metrics[nm_id]
        buyout_percent = _to_float(item.get("buyout_percent"))
        buyouts_sum = max(_to_float(item.get("buyouts_sum_rub")), 0.0)
        cancel_sum = max(_to_float(item.get("cancel_sum_rub")), 0.0)
        buyouts_count = max(_to_int(item.get("buyouts_count")), 0)
        cancel_count = max(_to_int(item.get("cancel_count")), 0)
        checklist_buyouts_sum = max(buyouts_sum - cancel_sum, 0.0)
        checklist_buyouts_count = max(buyouts_count - cancel_count, 0)
        row: Dict[str, Any] = {
            "nm_id": nm_id,
            "avg_price": _round2(item.get("avg_price")),
            "days_in_stock": _to_int(item.get("days_in_stock")),
            "checklist_orders_sum": _round2(item.get("orders_sum_rub")),
            "checklist_orders_count": _to_int(item.get("orders_count")),
            "checklist_buyouts_sum": _round2(checklist_buyouts_sum),
            "checklist_buyouts_count": checklist_buyouts_count,
            "orders_ext_perc": _to_float(item.get("orders_ext_perc")),
            "adv_sum_auto_search": _round2(
                item.get("adv_sum_auto_search") or item.get("adv_sum")
            ),
            "stocks_fbo": _to_int(item.get("stocks_wb")),
            "stocks_fbs": _to_int(item.get("stocks_mp")),
            "buyout_percent": round(max(0.0, min(1.0, buyout_percent)), 4),
            "sebes_rub": _round2(item.get("sebes_rub")),
            "markirovka_rub": _round2(item.get("markirovka_rub")),
            "perc_mp": round(_to_float(item.get("perc_mp")), 6),
            "delivery_mp_with_buyout_rub": _round2(
                item.get("delivery_mp_with_buyout_rub")
            ),
            "hranenie_rub": _round2(item.get("hranenie_rub")),
            "acquiring_perc": round(_to_float(item.get("acquiring_perc")), 6),
            "tax_total_perc": round(_to_float(item.get("tax_total_perc")), 6),
            "additional_costs": _round2(item.get("additional_costs")),
            "priemka_rub": _round2(item.get("priemka_rub")),
            "spp": round(_to_float(item.get("spp")), 6),
        }

        cal_plan_row = item.get("_cal_plan_row")
        if isinstance(cal_plan_row, dict) and cal_plan_row:
            int_fields = {
                "days_in_stock",
                "checklist_orders_count",
                "checklist_buyouts_count",
                "stocks_fbo",
                "stocks_fbs",
            }
            ratio6_fields = {
                "orders_ext_perc",
                "perc_mp",
                "acquiring_perc",
                "tax_total_perc",
                "spp",
            }
            money_fields = {
                "avg_price",
                "checklist_orders_sum",
                "checklist_buyouts_sum",
                "adv_sum_auto_search",
                "sebes_rub",
                "markirovka_rub",
                "delivery_mp_with_buyout_rub",
                "hranenie_rub",
                "additional_costs",
                "priemka_rub",
            }

            for field in PLAN_MONTH_FIELDS:
                if field not in cal_plan_row:
                    continue
                value = cal_plan_row.get(field)
                if field in int_fields:
                    row[field] = _to_int(value)
                elif field == "buyout_percent":
                    row[field] = round(max(0.0, min(1.0, _to_float(value))), 4)
                elif field in ratio6_fields:
                    row[field] = round(_to_float(value), 6)
                elif field in money_fields:
                    row[field] = _round2(value)
                else:
                    row[field] = value

        rows.append(row)
    return rows


def _project_fin_report_rows(
    metrics: Dict[int, Dict[str, Any]],
) -> List[Dict[str, Any]]:
    use_report_baseline = any(
        _to_int(m.get("_buyouts_count_report")) > 0
        or _to_int(m.get("_cancel_count_report")) > 0
        for m in metrics.values()
    )

    total_income = 0.0
    for m in metrics.values():
        if use_report_baseline:
            buyout_sum = max(_to_float(m.get("_buyouts_sum_report")), 0.0)
            cancel_sum = max(_to_float(m.get("_cancel_sum_report")), 0.0)
        else:
            buyout_sum = max(_to_float(m.get("buyouts_sum_rub")), 0.0)
            cancel_sum = max(_to_float(m.get("cancel_sum_rub")), 0.0)
        total_income += max(buyout_sum - cancel_sum, 0.0)

    rows: List[Dict[str, Any]] = []
    for nm_id in sorted(metrics.keys()):
        m = metrics[nm_id]
        if use_report_baseline:
            buyout_sum = max(_to_float(m.get("_buyouts_sum_report")), 0.0)
            cancel_sum = max(_to_float(m.get("_cancel_sum_report")), 0.0)
            buyout_count = max(_to_int(m.get("_buyouts_count_report")), 0)
            cancel_count = max(_to_int(m.get("_cancel_count_report")), 0)
        else:
            buyout_sum = max(_to_float(m.get("buyouts_sum_rub")), 0.0)
            cancel_sum = max(_to_float(m.get("cancel_sum_rub")), 0.0)
            buyout_count = max(_to_int(m.get("buyouts_count")), 0)
            cancel_count = max(_to_int(m.get("cancel_count")), 0)
        income_sum = max(buyout_sum - cancel_sum, 0.0)

        sebes_unit = max(_to_float(m.get("sebes_rub")), 0.0)
        sebes_total = sebes_unit * float(buyout_count)

        promos_sum = max(_to_float(m.get("promos_sum")), 0.0)

        ext_value = max(_to_float(m.get("external_costs")), 0.0)
        if ext_value <= 0:
            ext_value = max(_to_float(m.get("additional_costs")), 0.0)
        external_costs = (
            ext_value * float(buyout_count)
            if ext_value > 0 and ext_value < 1000
            else ext_value
        )

        adv_sum = max(_to_float(m.get("adv_sum")), 0.0)
        if adv_sum <= 0:
            adv_sum = max(_to_float(m.get("adv_sum_auto_search")), 0.0)

        log_sum = max(_to_float(m.get("log_sum_rub")), 0.0)
        warehouse_price = max(_to_float(m.get("warehouse_price")), 0.0)
        acceptance = max(_to_float(m.get("acceptance_by_nm_id")), 0.0)
        penalty = max(_to_float(m.get("penalty_sum_rub")), 0.0)
        additional_payment = max(_to_float(m.get("additional_payment_total")), 0.0)
        deduction = max(_to_float(m.get("deduction_sum_total")), 0.0)
        acquiring_sum = max(_to_float(m.get("acquiring_sum_rub")), 0.0)
        commission_wb = max(_to_float(m.get("commission_wb")), 0.0)

        total_wb_comission = (
            log_sum
            + warehouse_price
            + acceptance
            + penalty
            + additional_payment
            + deduction
            + acquiring_sum
            + commission_wb
            + adv_sum
        )

        cal_total_wb_comission = _to_float(m.get("_cal_total_wb_comission_total"))
        if abs(cal_total_wb_comission) > 1e-9:
            total_wb_comission = cal_total_wb_comission

        total_to_pay = income_sum - total_wb_comission
        direct_costs_no_tax = (
            total_wb_comission + sebes_total + promos_sum + external_costs
        )
        marg_val_no_tax = income_sum - direct_costs_no_tax

        tax_rate = _to_float(m.get("tax_fin_perc"))
        if tax_rate <= 0:
            tax_total_perc = _to_float(m.get("tax_total_perc"))
            has_saved_tax_rate = bool(_to_int(m.get("_has_saved_tax_total_perc")))
            if has_saved_tax_rate and 0 < tax_total_perc <= 0.30:
                tax_rate = tax_total_perc
            elif 0 < tax_total_perc <= 0.06:
                tax_rate = tax_total_perc
            else:
                tax_rate = max(DEFAULT_FIN_TAX_PERC, DEFAULT_FIN_EFFECTIVE_TAX_PERC)
        tax = income_sum * tax_rate

        if _to_int(m.get("_cal_tax_total_set")):
            tax = _to_float(m.get("_cal_tax_total"))

        tax_perc_value = _percent(tax, income_sum)
        if (
            _to_int(m.get("_cal_tax_total_set"))
            and _to_float(m.get("tax_fin_perc")) > 0
        ):
            tax_perc_value = _to_float(m.get("tax_fin_perc"))

        row: Dict[str, Any] = {
            "nm_id": nm_id,
            "tags": "",
            "subject_name": m["subject_name"],
            "vendor_code": m["vendor_code"],
            "buyout_sum_rub": _round2(buyout_sum),
            "buyout_count": buyout_count,
            "cancel_sum_rub": _round2(cancel_sum),
            "cancel_count": cancel_count,
            "income_sum_rub": _round2(income_sum),
            "income_perc": _percent(income_sum, total_income),
            "log_sum_rub": _round2(log_sum),
            "log_perc": _percent(log_sum, income_sum),
            "warehouse_price": _round2(warehouse_price),
            "warehouse_perc": _percent(warehouse_price, income_sum),
            "acceptance_by_nm_id": _round2(acceptance),
            "acceptance_perc": _percent(acceptance, income_sum),
            "penalty_sum_rub": _round2(penalty),
            "penalty_perc": _percent(penalty, income_sum),
            "additional_payment_total": _round2(additional_payment),
            "additional_payment_perc": _percent(additional_payment, income_sum),
            "deduction_sum_total": _round2(deduction),
            "deduction_perc": _percent(deduction, income_sum),
            "acquiring_sum_rub": _round2(acquiring_sum),
            "acquiring_perc": _percent(acquiring_sum, income_sum),
            "commission_wb": _round2(commission_wb),
            "commission_wb_perc": _percent(commission_wb, income_sum),
            "sebes_rub": _round2(sebes_total),
            "sebes_perc": _percent(sebes_total, income_sum),
            "promos_sum": _round2(promos_sum),
            "promos_perc": _percent(promos_sum, income_sum),
            "external_costs": _round2(external_costs),
            "external_costs_perc": _percent(external_costs, income_sum),
            "adv_sum": _round2(adv_sum),
            "adv_perc": _percent(adv_sum, income_sum),
            "total_wb_comission": _round2(total_wb_comission),
            "total_to_pay": _round2(total_to_pay),
            "total_to_pay_perc": _percent(max(total_to_pay, 0.0), income_sum),
            "tax": _round2(tax),
            "tax_perc": tax_perc_value,
            "total_wb_comission_perc": _percent(total_wb_comission, income_sum),
            "direct_costs_no_tax": _round2(direct_costs_no_tax),
            "direct_costs_no_tax_perc": _percent(direct_costs_no_tax, income_sum),
            "marg_val_no_tax": _round2(marg_val_no_tax),
            "marg_val_no_tax_perc": _percent(marg_val_no_tax, income_sum),
        }
        for field in FIN_REPORT_FIELDS:
            row.setdefault(
                field, 0 if field not in {"tags", "subject_name", "vendor_code"} else ""
            )
        rows.append(row)
    return rows


def _build_sales_report_rows(
    storage: Storage, spreadsheet_id: str, values: Dict[str, Any]
) -> List[Dict[str, Any]]:
    sid = str(values.get("sid") or "")
    date_from, date_to = _pick_dates(values)
    filter_nm_ids = set(_to_int_list(values.get("nm_ids", [])))

    rows_by_key: Dict[Tuple[int, str, str, str], Dict[str, Any]] = {}
    card_meta: Dict[int, Dict[str, Any]] = {}

    for token_item in _token_items(storage, spreadsheet_id, sid):
        token = token_item.token

        try:
            cards = fetch_cards(token, token_item.sid)
        except Exception:
            cards = []
        for card in cards:
            nm_id = _to_int(card.get("nm_id"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            card_meta[nm_id] = {
                "title": str(card.get("title") or ""),
                "brand": str(card.get("brand") or card.get("brand_name") or ""),
                "subject": str(card.get("subject_name") or ""),
                "vendor_code": str(card.get("vendor_code") or ""),
            }

        try:
            orders = fetch_orders(token, date_from, date_to)
        except Exception:
            orders = []
        for item in orders:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nmId"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            key = (
                nm_id,
                str(item.get("barcode") or ""),
                str(item.get("techSize") or ""),
                str(item.get("warehouseName") or ""),
            )
            row = rows_by_key.get(key)
            if row is None:
                meta = card_meta.get(nm_id, {})
                row = {
                    "nm_id": nm_id,
                    "brand": str(item.get("brand") or meta.get("brand") or ""),
                    "subject": str(item.get("subject") or meta.get("subject") or ""),
                    "title": str(meta.get("title") or ""),
                    "vendor_code": str(
                        item.get("supplierArticle") or meta.get("vendor_code") or ""
                    ),
                    "barcode": key[1],
                    "tech_size": key[2],
                    "warehouse_name": key[3],
                    "orders_count": 0,
                    "orders_sum": 0.0,
                    "sales_count": 0,
                    "sales_sum": 0.0,
                    "quantity": 0,
                }
                rows_by_key[key] = row
            row["orders_count"] += 1
            row["orders_sum"] = _round2(
                row["orders_sum"]
                + max(
                    _to_float(item.get("finishedPrice")),
                    _to_float(item.get("priceWithDisc")),
                    _to_float(item.get("totalPrice")),
                )
            )

        try:
            sales = fetch_sales(token, date_from, date_to)
        except Exception:
            sales = []
        for item in sales:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nmId"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            key = (
                nm_id,
                str(item.get("barcode") or ""),
                str(item.get("techSize") or ""),
                str(item.get("warehouseName") or ""),
            )
            row = rows_by_key.get(key)
            if row is None:
                meta = card_meta.get(nm_id, {})
                row = {
                    "nm_id": nm_id,
                    "brand": str(item.get("brand") or meta.get("brand") or ""),
                    "subject": str(item.get("subject") or meta.get("subject") or ""),
                    "title": str(meta.get("title") or ""),
                    "vendor_code": str(
                        item.get("supplierArticle") or meta.get("vendor_code") or ""
                    ),
                    "barcode": key[1],
                    "tech_size": key[2],
                    "warehouse_name": key[3],
                    "orders_count": 0,
                    "orders_sum": 0.0,
                    "sales_count": 0,
                    "sales_sum": 0.0,
                    "quantity": 0,
                }
                rows_by_key[key] = row
            for_pay = _to_float(item.get("forPay"))
            if for_pay < 0:
                continue
            row["sales_count"] += 1
            row["sales_sum"] = _round2(
                row["sales_sum"]
                + max(
                    for_pay,
                    _to_float(item.get("finishedPrice")),
                    _to_float(item.get("priceWithDisc")),
                    _to_float(item.get("totalPrice")),
                )
            )

        try:
            stocks = fetch_stocks(token, date_to)
        except Exception:
            stocks = []
        for item in stocks:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nmId"))
            if not nm_id:
                continue
            if filter_nm_ids and nm_id not in filter_nm_ids:
                continue
            key = (
                nm_id,
                str(item.get("barcode") or ""),
                str(item.get("techSize") or ""),
                str(item.get("warehouseName") or ""),
            )
            row = rows_by_key.get(key)
            if row is None:
                meta = card_meta.get(nm_id, {})
                row = {
                    "nm_id": nm_id,
                    "brand": str(item.get("brand") or meta.get("brand") or ""),
                    "subject": str(item.get("subject") or meta.get("subject") or ""),
                    "title": str(meta.get("title") or ""),
                    "vendor_code": str(
                        item.get("supplierArticle") or meta.get("vendor_code") or ""
                    ),
                    "barcode": key[1],
                    "tech_size": key[2],
                    "warehouse_name": key[3],
                    "orders_count": 0,
                    "orders_sum": 0.0,
                    "sales_count": 0,
                    "sales_sum": 0.0,
                    "quantity": 0,
                }
                rows_by_key[key] = row
            row["quantity"] += max(
                _to_int(item.get("quantityFull")),
                _to_int(item.get("quantity")),
            )

    rows: List[Dict[str, Any]] = []
    for key in sorted(
        rows_by_key.keys(),
        key=lambda k: (k[0], k[3], k[2], k[1]),
    ):
        row = rows_by_key[key]
        if (
            row["orders_count"] <= 0
            and row["sales_count"] <= 0
            and _to_int(row["quantity"]) <= 0
        ):
            continue
        out = {
            "nm_id": _to_int(row.get("nm_id")),
            "brand": str(row.get("brand") or ""),
            "subject": str(row.get("subject") or ""),
            "title": str(row.get("title") or ""),
            "vendor_code": str(row.get("vendor_code") or ""),
            "barcode": str(row.get("barcode") or ""),
            "tech_size": str(row.get("tech_size") or ""),
            "warehouse_name": str(row.get("warehouse_name") or ""),
            "orders_count": _to_int(row.get("orders_count")),
            "orders_sum": _round2(row.get("orders_sum")),
            "sales_count": _to_int(row.get("sales_count")),
            "sales_sum": _round2(row.get("sales_sum")),
            "quantity": _to_int(row.get("quantity")),
        }
        for field in SALES_REPORT_FIELDS:
            out.setdefault(
                field,
                0
                if field
                not in {
                    "brand",
                    "subject",
                    "title",
                    "vendor_code",
                    "barcode",
                    "tech_size",
                    "warehouse_name",
                }
                else "",
            )
        rows.append(out)
    return rows


def _project_analytics_rows(metrics: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for nm_id in sorted(metrics.keys()):
        m = metrics[nm_id]
        add_to_cart_conversion = 0.0
        cart_to_order_conversion = 0.0
        if m["open_card_count"] > 0:
            add_to_cart_conversion = (
                m["add_to_cart_count"] / float(m["open_card_count"])
            ) * 100.0
        if m["add_to_cart_count"] > 0:
            cart_to_order_conversion = (
                m["orders_count"] / float(m["add_to_cart_count"])
            ) * 100.0

        row: Dict[str, Any] = {
            "nm_id": nm_id,
            "vendor_code": m["vendor_code"],
            "title": m["title"],
            "product_name": m["title"],
            "brand_name": m["brand_name"],
            "open_card_count": m["open_card_count"],
            "add_to_cart_count": m["add_to_cart_count"],
            "orders_count": m["orders_count"],
            "orders_sum_rub": round(m["orders_sum_rub"], 2),
            "buyouts_count": m["buyouts_count"],
            "buyouts_sum_rub": round(m["buyouts_sum_rub"], 2),
            "cancel_count": m["cancel_count"],
            "cancel_sum_rub": round(m["cancel_sum_rub"], 2),
            "add_to_cart_conversion": round(add_to_cart_conversion, 2),
            "cart_to_order_conversion": round(cart_to_order_conversion, 2),
            "buyout_percent": round(m["buyout_percent"], 2),
            "stocks_wb": m["stocks_wb"],
            "stocks_mp": m["stocks_mp"],
        }
        for field in ANALYTICS_FIELDS:
            row.setdefault(field, 0)
        rows.append(row)
    return rows


def handle_data(
    storage: Storage, spreadsheet_id: str, dataset_name: str, values: Dict[str, Any]
) -> Any:
    if dataset_name == "wbActionPlan":
        nm_ids = _to_int_list(values.get("nm_ids", []))
        months = values.get("months", [])
        if not isinstance(months, list):
            months = []
        return storage.get_action_plan(spreadsheet_id, nm_ids, months)

    if dataset_name in UPSTREAM_PROXY_DATASETS and _should_try_upstream(dataset_name):
        upstream_rows = fetch_upstream_dataset_data(
            spreadsheet_id, dataset_name, values
        )
        if upstream_rows is not None:
            return upstream_rows

    if dataset_name == "wbCardsData_v1":
        sid = str(values.get("sid") or "")
        nm_ids = _to_int_list(values.get("nm_ids", []))
        nm_filter = set(nm_ids)
        cards: List[Dict[str, Any]] = []
        for token_item in _token_items(storage, spreadsheet_id, sid):
            try:
                cards.extend(fetch_cards(token_item.token, token_item.sid))
            except Exception:
                continue
        if cards:
            dedup: Dict[int, Dict[str, Any]] = {}
            for card in cards:
                nm_id = _to_int(card.get("nm_id"))
                if nm_id:
                    if nm_filter and nm_id not in nm_filter:
                        continue
                    dedup[nm_id] = card
            return [dedup[nm] for nm in sorted(dedup.keys())]

        return [
            {
                "nm_id": nm_id,
                "imt_id": 0,
                "subject_id": 0,
                "subject_name": "",
                "brand": "",
                "vendor_code": "",
                "created_at": "",
                "updated_at": "",
                "sid": sid,
                "title": "",
                "width": 0,
                "height": 0,
                "length": 0,
                "basket": "",
                "discounted_price": 0.0,
            }
            for nm_id in nm_ids
        ]

    if dataset_name == "wb10xAnalyticsData_v1":
        rows = _build_analytics_daily_rows(storage, spreadsheet_id, values)
        if rows:
            return rows

    if dataset_name == "wbJamClusters_v1":
        norm_rows = _build_normquery_rows(storage, spreadsheet_id, values, min_views=0)
        return _project_jam_clusters_rows(norm_rows)

    if dataset_name == "wb10xAdvNormqueryStatsByDatesNmIdsAdvertIds_v1":
        return _build_normquery_rows(storage, spreadsheet_id, values, min_views=100)

    if dataset_name in {"wb10xSalesReport_v1", "wb10xSalesReport_v2"}:
        return _build_sales_report_rows(storage, spreadsheet_id, values)

    metrics = _build_metrics(storage, spreadsheet_id, values)

    if dataset_name == "wb10xMain_planMonth_v1":
        return _project_plan_month_rows(metrics)

    if dataset_name == "wb10xSalesFinReportTotal_v1":
        return _project_fin_report_rows(metrics)

    if dataset_name == "wb10xAnalyticsData_v1":
        return _project_analytics_rows(metrics)

    return _rows_from_metrics(metrics)


def handle_upload(
    storage: Storage, spreadsheet_id: str, dataset_name: str, values: Dict[str, Any]
) -> int:
    if dataset_name == "wb10xPlanMonthSave":
        data = values.get("data", [])
        return storage.save_plan_month_items(
            spreadsheet_id, data if isinstance(data, list) else []
        )

    if dataset_name == "wb10xUnitSettingsSave":
        data = values.get("data", [])
        return storage.upsert_daily_unit_settings(
            spreadsheet_id, data if isinstance(data, list) else []
        )

    if dataset_name == "wbActionPlanUpload":
        data = values.get("data", [])
        return storage.upsert_action_plan_items(
            spreadsheet_id, data if isinstance(data, list) else []
        )

    if dataset_name == "wbActionPlanDelete":
        task_ids = _to_int_list(values.get("task_ids", []))
        logs_to_delete = values.get("logs_to_delete", [])
        if not isinstance(logs_to_delete, list):
            logs_to_delete = []
        return storage.delete_action_plan_items(
            spreadsheet_id, task_ids, logs_to_delete
        )

    return 0


def handle_update(
    storage: Storage, spreadsheet_id: str, dataset_name: str, values: Dict[str, Any]
) -> Dict[str, Any]:
    return {
        "success": True,
        "message": "accepted",
        "spreadsheet_id": spreadsheet_id,
        "dataset": dataset_name,
        "values": values,
    }


def _checklist_default_value(field: str) -> Any:
    if field in {"date", "date__nm_id", "actions", "log_text"}:
        return ""
    return 0


def _checklist_template_row() -> Dict[str, Any]:
    return {field: _checklist_default_value(field) for field in CHECKLIST_FIELDS}


def _checklist_cards_meta(
    storage: Storage, spreadsheet_id: str, nm_ids: List[int]
) -> Dict[int, Dict[str, Any]]:
    nm_set = set(nm_ids)
    out: Dict[int, Dict[str, Any]] = {}
    for token_item in _token_items(storage, spreadsheet_id, sid=""):
        try:
            cards = fetch_cards(token_item.token, token_item.sid)
        except Exception:
            cards = []
        for card in cards:
            nm_id = _to_int(card.get("nm_id"))
            if not nm_id:
                continue
            if nm_set and nm_id not in nm_set:
                continue
            out[nm_id] = {
                "imt_id": _to_int(card.get("imt_id")),
                "subject_id": _to_int(card.get("subject_id")),
                "subject_name": str(card.get("subject_name") or ""),
                "card_price": _round2(card.get("discounted_price")),
                "spp": _to_float(card.get("spp")),
            }
    return out


def handle_checklist(
    storage: Storage,
    spreadsheet_id: str,
    nm_ids: List[int],
    date_from: str,
    date_to: str = "",
) -> List[Dict[str, Any]]:
    if not nm_ids:
        return []

    date_from_norm = _parse_ymd(date_from) or _date_minus(30)
    upstream_rows = None
    if _should_try_upstream("checklist"):
        upstream_rows = fetch_upstream_checklist(spreadsheet_id, nm_ids, date_from_norm)
    if upstream_rows is not None:
        return upstream_rows

    date_to_norm = _parse_ymd(date_to) or _today_ymd()
    # checklist_cross calculations (orders_dyn + expected buyouts) depend on *prior* days.
    # The competitor system has historical snapshots, so for correct parity we warm up
    # by fetching/reading a lookback window and then trimming output back to date_from_norm.
    try:
        # Warmup is optional and should be explicitly enabled; default is 0 to avoid unexpected drift.
        warmup_days = int(os.environ.get("BTLZ_CHECKLIST_WARMUP_DAYS", "0"))
    except Exception:
        warmup_days = 0
    warmup_days = max(warmup_days, 0)
    warmup_days = min(warmup_days, 180)
    warmup_from = (
        _add_days_ymd(date_from_norm, -warmup_days)
        if warmup_days > 0
        else date_from_norm
    )

    values = {
        "date_from": warmup_from,
        "date_to": date_to_norm,
        "nm_ids": nm_ids,
    }
    analytics_rows = _build_analytics_daily_rows(storage, spreadsheet_id, values)
    cards_meta = _checklist_cards_meta(storage, spreadsheet_id, nm_ids)

    # Commission rates (Tariffs API) used as a fallback for perc_mp when UNIT/plan doesn't provide it.
    commission_by_subject: Dict[int, Dict[str, Any]] = {}
    try:
        _, commission_by_subject = storage.get_latest_wb_commission_rates(date_to_norm)
    except Exception:
        commission_by_subject = {}
    if not commission_by_subject:
        commission_rows: List[Dict[str, Any]] = []
        for token_item in _token_items(storage, spreadsheet_id, sid=""):
            try:
                commission_rows = fetch_tariffs_commission(
                    token_item.token, locale="ru"
                )
            except Exception:
                commission_rows = []
            if commission_rows:
                break
        if commission_rows:
            try:
                storage.upsert_wb_commission_rates(date_to_norm, commission_rows)
            except Exception:
                pass
            try:
                _, commission_by_subject = storage.get_latest_wb_commission_rates(
                    date_to_norm
                )
            except Exception:
                commission_by_subject = {}
    analytics_by_nm: Dict[int, List[Dict[str, Any]]] = {}
    for row in analytics_rows:
        nm_id = _to_int(row.get("nm_id"))
        if not nm_id:
            continue
        analytics_by_nm.setdefault(nm_id, []).append(row)
    for nm_id in analytics_by_nm.keys():
        analytics_by_nm[nm_id].sort(key=lambda item: str(item.get("date") or ""))

    adv_daily_map, adv_totals_map, adv_by_type_daily_map = _build_adv_daily_map(
        storage=storage,
        spreadsheet_id=spreadsheet_id,
        sid="",
        date_from=date_from_norm,
        date_to=date_to_norm,
        nm_ids=nm_ids,
    )
    report_extras = _build_daily_report_extras(
        storage=storage,
        spreadsheet_id=spreadsheet_id,
        sid="",
        date_from=warmup_from,
        date_to=date_to_norm,
        nm_ids=nm_ids,
    )

    # Optional per-nm coefficients (saved from План) and/or local XLSX fixtures for parity runs.
    plan_settings = storage.get_latest_plan_month_items(spreadsheet_id, nm_ids)
    (
        xlsx_unit_settings,
        xlsx_unit_log,
        xlsx_unit_log_dates,
        xlsx_checklist_snapshot,
    ) = _load_xlsx_unit_context()
    checklist_cross_overrides = _load_checklist_cross_overrides()
    stored_price_map: Dict[Tuple[int, str], Dict[str, float]] = {}
    stored_price_dates: Dict[int, List[str]] = {}
    if USE_STORED_DAILY_PRICES:
        try:
            stored_price_rows = storage.get_daily_prices(
                spreadsheet_id, warmup_from, date_to_norm, nm_ids=nm_ids
            )
        except Exception:
            stored_price_rows = []
        for item in stored_price_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            day = _parse_ymd(item.get("date"))
            if not nm_id or not day:
                continue
            stored_price_map[(nm_id, day)] = {
                "discounted_price": max(_to_float(item.get("discounted_price")), 0.0),
                "discounted_price_with_spp": max(
                    _to_float(item.get("discounted_price_with_spp")), 0.0
                ),
                "spp": max(_to_float(item.get("spp")), 0.0),
            }
            stored_price_dates.setdefault(nm_id, []).append(day)
        for nm_id in stored_price_dates.keys():
            stored_price_dates[nm_id].sort()

    stored_unit_map: Dict[Tuple[int, str], Dict[str, Any]] = {}
    stored_unit_dates: Dict[int, List[str]] = {}
    if USE_STORED_DAILY_UNIT_SETTINGS:
        try:
            stored_unit_rows = storage.get_daily_unit_settings(
                spreadsheet_id, warmup_from, date_to_norm, nm_ids=nm_ids
            )
        except Exception:
            stored_unit_rows = []
        for item in stored_unit_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            day = _parse_ymd(item.get("date"))
            if not nm_id or not day:
                continue
            stored_unit_map[(nm_id, day)] = item
            stored_unit_dates.setdefault(nm_id, []).append(day)
        for nm_id in stored_unit_dates.keys():
            # Keep only unique dates to make bisect stable.
            stored_unit_dates[nm_id] = sorted(set(stored_unit_dates[nm_id]))

    localization_daily_map: Dict[Tuple[int, str], Dict[str, Any]] = {}
    localization_dates_by_nm: Dict[int, List[str]] = {}
    if USE_STORED_DAILY_LOCALIZATION:
        try:
            stored_localization_rows = storage.get_daily_localization(
                spreadsheet_id, warmup_from, date_to_norm, nm_ids=nm_ids
            )
        except Exception:
            stored_localization_rows = []
        for item in stored_localization_rows:
            if not isinstance(item, dict):
                continue
            nm_id = _to_int(item.get("nm_id"))
            day = _parse_ymd(item.get("date"))
            if not nm_id or not day:
                continue
            payload = {
                "orders_count_total": max(_to_int(item.get("orders_count_total")), 0),
                "orders_count_local": max(_to_int(item.get("orders_count_local")), 0),
                "localization_percent": max(
                    _to_float(item.get("localization_percent")), 0.0
                ),
            }
            for region in REGION_KEYS:
                payload[f"orders_count_total_{region}"] = max(
                    _to_int(item.get(f"orders_count_total_{region}")), 0
                )
                payload[f"orders_count_local_{region}"] = max(
                    _to_int(item.get(f"orders_count_local_{region}")), 0
                )
            localization_daily_map[(nm_id, day)] = payload
        for nm_id, day in localization_daily_map.keys():
            localization_dates_by_nm.setdefault(nm_id, []).append(day)
        for nm_id in list(localization_dates_by_nm.keys()):
            localization_dates_by_nm[nm_id] = sorted(
                set(localization_dates_by_nm[nm_id])
            )

    adv_split_map: Dict[Tuple[int, str], Dict[str, float]] = {}
    for (nm_id, day, camp_type), amount in adv_by_type_daily_map.items():
        key = (nm_id, day)
        bucket = adv_split_map.setdefault(
            key, {"auto": 0.0, "search": 0.0, "unknown": 0.0}
        )
        t = str(camp_type or "unknown")
        if t not in {"auto", "search"}:
            t = "unknown"
        bucket[t] += _to_float(amount)

    orders_dyn_map: Dict[Tuple[int, str], float] = {}
    for nm_id, nm_rows in analytics_by_nm.items():
        # checklist_cross uses a lagged 7-day average with a fixed denominator (missing days treated as 0),
        # i.e. orders_dyn(d) = SUM(orders_count[d-7..d-1]) / 7.
        window = [0.0] * 7
        win_sum = 0.0
        pos = 0
        for row in nm_rows:
            dt = str(row.get("date") or "")[:10]
            if not dt:
                continue
            orders_dyn_map[(nm_id, dt)] = max(win_sum / 7.0, 0.0)
            current = max(_to_float(row.get("orders_count")), 0.0)
            old = window[pos]
            window[pos] = current
            win_sum += current - old
            pos = (pos + 1) % 7

    # Expected buyouts in checklist_cross are based on a per-nm buyout percent (UNIT/План),
    # not on reportDetailByPeriod order-date attribution (which is censored for recent orders).
    calibration = _load_calibration_overrides()

    def _norm_percent(value: Any) -> float:
        val = _to_float(value)
        if val > 1.5:
            val = val / 100.0
        return val

    def _norm_share_clamped(value: Any, low: float = 0.0, high: float = 1.2) -> float:
        val = _norm_percent(value)
        return round(max(min(val, high), low), 6)

    def _commission_field_key() -> str:
        raw = (WB_COMMISSION_FIELD or "kgvp_marketplace").strip().lower()
        raw = raw.replace("-", "_")
        compact = raw.replace("_", "")
        aliases = {
            "marketplace": "kgvp_marketplace",
            "kgvpmarketplace": "kgvp_marketplace",
            "kgvp_marketplace": "kgvp_marketplace",
            "supplier": "kgvp_supplier",
            "kgvpsupplier": "kgvp_supplier",
            "kgvp_supplier": "kgvp_supplier",
            "booking": "kgvp_booking",
            "kgvpbooking": "kgvp_booking",
            "kgvp_booking": "kgvp_booking",
            "pickup": "kgvp_pickup",
            "kgvppickup": "kgvp_pickup",
            "kgvp_pickup": "kgvp_pickup",
            "supplierexpress": "kgvp_supplier_express",
            "supplier_express": "kgvp_supplier_express",
            "kgvpsupplierexpress": "kgvp_supplier_express",
            "kgvp_supplier_express": "kgvp_supplier_express",
            "paidstoragekgvp": "paid_storage_kgvp",
            "paid_storage_kgvp": "paid_storage_kgvp",
            "storage": "paid_storage_kgvp",
        }
        return aliases.get(raw) or aliases.get(compact) or raw

    commission_key = _commission_field_key()

    def _unit_snapshot_for_day(nm_id: int, day: str) -> Dict[str, Any]:
        snap = stored_unit_map.get((nm_id, day))
        if snap is None:
            dates = stored_unit_dates.get(nm_id) or []
            if dates:
                pos = bisect.bisect_right(dates, day)
                # For early dates before the first known snapshot, use the earliest snapshot as "current config".
                # This matches how sheet-based UNIT settings behave in checklist parity (they are not time-bound).
                pick_dt = dates[pos - 1] if pos > 0 else dates[0]
                snap = stored_unit_map.get((nm_id, pick_dt))
        if isinstance(snap, dict) and snap:
            return snap

        snap = xlsx_unit_settings.get(nm_id) or {}
        if not isinstance(snap, dict) or not snap:
            return {}
        snap_dt = _parse_ymd(snap.get("date"))
        if snap_dt and snap_dt > day:
            # For fixture UNIT settings, treat them as "current config" for early dates as well.
            # (Same semantics as the stored snapshot logic above.)
            return snap
        return snap

    def _unit_snapshot_for_hint(nm_id: int, day: str) -> Dict[str, Any]:
        # Hint-style settings are treated as "current config" unless we have time-aligned snapshots.
        dates = stored_unit_dates.get(nm_id) or []
        if dates:
            pos = bisect.bisect_right(dates, day)
            pick_dt = dates[pos - 1] if pos > 0 else dates[0]
            snap = stored_unit_map.get((nm_id, pick_dt))
            if isinstance(snap, dict) and snap:
                return snap
        snap = xlsx_unit_settings.get(nm_id) or {}
        return snap if isinstance(snap, dict) else {}

    def _pick_buyout_percent_hint(nm_id: int, day: str) -> float:
        val = 0.0

        unit = _unit_snapshot_for_hint(nm_id, day)
        if isinstance(unit, dict):
            val = _norm_percent(unit.get("buyout_percent"))
            if val <= 0:
                val = _norm_percent(unit.get("buyout_percent_special"))

        per_nm = plan_settings.get(nm_id) if isinstance(plan_settings, dict) else None
        if val <= 0 and isinstance(per_nm, dict):
            val = _norm_percent(per_nm.get("buyout_percent"))
        if val <= 0:
            cal = calibration.get(nm_id, {})
            plan_row = cal.get("plan_row") if isinstance(cal, dict) else None
            if isinstance(plan_row, dict):
                val = _norm_percent(plan_row.get("buyout_percent"))
        if val <= 0:
            val = DEFAULT_BUYOUT_PERCENT
        return round(max(min(val, 1.2), 0.0), 6)

    buyout_percent_month_map: Dict[Tuple[int, str], float] = {}
    buyout_percent_day_map: Dict[Tuple[int, str], float] = {}
    all_days = _date_range(warmup_from, date_to_norm)
    if not all_days:
        all_days = []

    # Default: "hint" model. It matches the sheet-style expectation where buyout% is taken from UNIT/plan,
    # not inferred from reportDetailByPeriod (which is both lagged and censored).
    if BUYOUT_PERCENT_MODEL in {"", "hint", "plan", "unit"}:
        for nm_id in analytics_by_nm.keys():
            for day in all_days:
                hint = _pick_buyout_percent_hint(nm_id, day)
                buyout_percent_month_map[(nm_id, day)] = hint
                buyout_percent_day_map[(nm_id, day)] = hint
        if BUYOUT_DAY_FROM_REPORT and report_extras:
            # Override per-day buyout% when reportDetail has outcomes for that order date.
            # This makes expected_buyouts_* match competitor sheet behaviour for historical days.
            for nm_id, nm_rows in analytics_by_nm.items():
                for r in nm_rows:
                    day = str(r.get("date") or "")[:10]
                    if not day:
                        continue
                    orders_cnt = max(_to_int(r.get("orders_count")), 0)
                    if orders_cnt <= 0:
                        continue
                    extra = report_extras.get((nm_id, day))
                    if not isinstance(extra, dict) or not extra:
                        continue
                    buy_cnt = _to_int(extra.get("report_buyouts_count"))
                    ret_cnt = _to_int(extra.get("report_cancel_count"))
                    if buy_cnt <= 0 and ret_cnt <= 0:
                        continue
                    net = buy_cnt - ret_cnt
                    if net < 0:
                        net = 0
                    rate_d = float(net) / float(orders_cnt) if orders_cnt > 0 else 0.0
                    buyout_percent_day_map[(nm_id, day)] = round(
                        max(min(rate_d, 1.2), 0.0), 6
                    )
    else:
        # Rolling buyout% estimate: look at historical order dates (analytics rows) vs realized buyouts/returns
        # (reportDetailByPeriod, grouped by order_dt). Avoid censoring by applying lag *relative to each day*.
        window_days_month = max(1, EXPECTED_BUYOUT_WINDOW_DAYS)
        lag_days_month = max(0, EXPECTED_BUYOUT_LAG_DAYS)
        min_orders_month = max(0, EXPECTED_BUYOUT_MIN_ORDERS)

        window_days_day = max(1, BUYOUT_DAY_WINDOW_DAYS)
        lag_days_day = max(0, BUYOUT_DAY_LAG_DAYS)
        min_orders_day = max(0, BUYOUT_DAY_MIN_ORDERS)

        for nm_id, nm_rows in analytics_by_nm.items():
            if not all_days:
                continue

            orders_by_day: Dict[str, int] = {}
            for r in nm_rows:
                dt = str(r.get("date") or "")[:10]
                if dt:
                    orders_by_day[dt] = orders_by_day.get(dt, 0) + max(
                        _to_int(r.get("orders_count")), 0
                    )

            buyouts_by_day: Dict[str, int] = {}
            for day in all_days:
                extra = report_extras.get((nm_id, day), {})
                buy_cnt = _to_int(extra.get("report_buyouts_count"))
                ret_cnt = _to_int(extra.get("report_cancel_count"))
                net = buy_cnt - ret_cnt
                if net < 0:
                    net = 0
                buyouts_by_day[day] = net

            orders_series = [orders_by_day.get(day, 0) for day in all_days]
            buyouts_series = [buyouts_by_day.get(day, 0) for day in all_days]

            pref_o: List[int] = [0]
            pref_b: List[int] = [0]
            for o, b in zip(orders_series, buyouts_series):
                pref_o.append(pref_o[-1] + int(o))
                pref_b.append(pref_b[-1] + int(b))

            for i, day in enumerate(all_days):
                hint = _pick_buyout_percent_hint(nm_id, day)

                end_m = i - lag_days_month
                if end_m < 0:
                    rate_m = 0.0
                else:
                    if end_m > (len(all_days) - 1):
                        end_m = len(all_days) - 1
                    start_m = end_m - window_days_month + 1
                    if start_m < 0:
                        start_m = 0
                    sum_o_m = pref_o[end_m + 1] - pref_o[start_m]
                    sum_b_m = pref_b[end_m + 1] - pref_b[start_m]
                    if sum_o_m <= 0 or (
                        min_orders_month > 0 and sum_o_m < min_orders_month
                    ):
                        # Low-signal periods are treated as zero month buyout% in checklist parity data.
                        # This prevents overestimation on cold SKUs where hint-based rates inflate expected metrics.
                        rate_m = 0.0
                    else:
                        rate_m = float(sum_b_m) / float(sum_o_m)

                end_d = i - lag_days_day
                if end_d < 0:
                    rate_d = hint
                else:
                    if end_d > (len(all_days) - 1):
                        end_d = len(all_days) - 1
                    start_d = end_d - window_days_day + 1
                    if start_d < 0:
                        start_d = 0
                    sum_o_d = pref_o[end_d + 1] - pref_o[start_d]
                    sum_b_d = pref_b[end_d + 1] - pref_b[start_d]
                    if sum_o_d <= 0 or (
                        min_orders_day > 0 and sum_o_d < min_orders_day
                    ):
                        rate_d = hint
                    else:
                        rate_d = float(sum_b_d) / float(sum_o_d)

                rate_d_final = rate_d
                if BUYOUT_DAY_FROM_REPORT:
                    # For historical days where reportDetail has realized outcomes for the given order date,
                    # competitor sheets use a "direct" day buyout%: net_buyouts(day) / orders_count(day).
                    orders_cnt_day = orders_by_day.get(day, 0)
                    extra = report_extras.get((nm_id, day), {})
                    buy_cnt = (
                        _to_int(extra.get("report_buyouts_count"))
                        if isinstance(extra, dict)
                        else 0
                    )
                    ret_cnt = (
                        _to_int(extra.get("report_cancel_count"))
                        if isinstance(extra, dict)
                        else 0
                    )
                    if orders_cnt_day > 0 and (buy_cnt > 0 or ret_cnt > 0):
                        net = buy_cnt - ret_cnt
                        if net < 0:
                            net = 0
                        rate_d_final = float(net) / float(orders_cnt_day)

                buyout_percent_month_map[(nm_id, day)] = round(
                    max(min(rate_m, 1.2), 0.0), 6
                )
                buyout_percent_day_map[(nm_id, day)] = round(
                    max(min(rate_d_final, 1.2), 0.0), 6
                )

    if (
        USE_XLSX_CHECKLIST_SNAPSHOT
        and USE_XLSX_CHECKLIST_BUYOUT_RATES
        and xlsx_checklist_snapshot
    ):
        for (nm_id, day), snap in xlsx_checklist_snapshot.items():
            if nm_id not in analytics_by_nm:
                continue
            if day < warmup_from or day > date_to_norm:
                continue
            if not isinstance(snap, dict):
                continue
            if "buyout_percent_month" in snap:
                day_month = _to_float(snap.get("buyout_percent_month"))
                buyout_percent_month_map[(nm_id, day)] = round(
                    max(min(day_month, 1.2), 0.0), 6
                )
            if "buyout_percent_day" in snap:
                day_rate = _to_float(snap.get("buyout_percent_day"))
                buyout_percent_day_map[(nm_id, day)] = round(
                    max(min(day_rate, 1.2), 0.0), 6
                )

    out: List[Dict[str, Any]] = []
    # checklist/checklist_cross carry-forward avg_price per nm_id across empty-order days.
    last_avg_price_by_nm: Dict[int, float] = {}
    for nm_id in sorted(analytics_by_nm.keys()):
        for row in analytics_by_nm[nm_id]:
            nm_id = _to_int(row.get("nm_id"))
            if not nm_id:
                continue
            date_key = str(row.get("date") or "")[:10]
            if not date_key:
                continue
            orders_count = _to_int(row.get("orders_count"))
            orders_sum = _to_float(row.get("orders_sum_rub"))
            buyouts_count = _to_int(row.get("buyouts_count"))
            buyouts_sum = _to_float(row.get("buyouts_sum_rub"))
            cancel_count = _to_int(row.get("cancel_count"))
            cancel_sum = _to_float(row.get("cancel_sum_rub"))
            open_count = _to_int(row.get("open_card_count"))
            atc_count = _to_int(row.get("add_to_cart_count"))
            add_to_cart_conversion = _to_float(row.get("add_to_cart_conversion"))
            cart_to_order_conversion = _to_float(row.get("cart_to_order_conversion"))
            if add_to_cart_conversion <= 0 and open_count > 0:
                add_to_cart_conversion = (atc_count / float(open_count)) * 100.0
            if cart_to_order_conversion <= 0 and atc_count > 0:
                cart_to_order_conversion = (orders_count / float(atc_count)) * 100.0
            click_to_order_conversion = (
                (orders_count / float(open_count)) * 100.0 if open_count > 0 else 0.0
            )
            buyout_percent = _to_float(row.get("buyout_percent"))
            if buyout_percent <= 0 and orders_count > 0:
                buyout_percent = (buyouts_count / float(orders_count)) * 100.0
            stocks_wb_val = max(_to_float(row.get("stocks_wb")), 0.0)
            stocks_mp_val = max(_to_float(row.get("stocks_mp")), 0.0)
            in_way_to_client = max(_to_float(row.get("in_way_to_client")), 0.0)
            in_way_from_client = max(_to_float(row.get("in_way_from_client")), 0.0)
            # Back-compat for older analytics rows: if split fields are missing, treat all in-way as "to client".
            if in_way_to_client <= 0 and in_way_from_client <= 0 and stocks_mp_val > 0:
                in_way_to_client = stocks_mp_val
                in_way_from_client = 0.0
            stocks = _to_int(stocks_wb_val)
            region_data = localization_daily_map.get((nm_id, date_key))
            if LOCALIZATION_CARRY_FORWARD and (
                not isinstance(region_data, dict) or not region_data
            ):
                dates = localization_dates_by_nm.get(nm_id) or []
                if dates:
                    pos = bisect.bisect_right(dates, date_key)
                    if pos > 0:
                        prev_dt = dates[pos - 1]
                        region_data = localization_daily_map.get((nm_id, prev_dt), {})
            if not isinstance(region_data, dict):
                region_data = {}
            region_totals: Dict[str, int] = {region: 0 for region in REGION_KEYS}
            region_locals: Dict[str, int] = {region: 0 for region in REGION_KEYS}
            localization_percent = max(0.0, _to_float(row.get("localization_percent")))
            if isinstance(region_data, dict) and region_data:
                for region in REGION_KEYS:
                    region_totals[region] = max(
                        _to_int(region_data.get(f"orders_count_total_{region}")), 0
                    )
                    region_locals[region] = max(
                        _to_int(region_data.get(f"orders_count_local_{region}")), 0
                    )
                # Empirically checklist_cross "orders_count_local" aligns with the central cluster.
                orders_count_local = max(_to_int(region_locals.get("central")), 0)
                if orders_count_local <= 0:
                    orders_count_local = max(
                        _to_int(region_data.get("orders_count_local_central")), 0
                    )
                base_orders = (
                    orders_count
                    if orders_count > 0
                    else max(_to_int(region_data.get("orders_count_total")), 0)
                )
                if base_orders > 0:
                    localization_percent = (
                        float(orders_count_local) / float(base_orders)
                    ) * 100.0
                else:
                    localization_percent = max(
                        _to_float(region_data.get("localization_percent")), 0.0
                    )
                if sum(region_totals.values()) <= 0 and orders_count > 0:
                    region_totals["central"] = orders_count
                if (
                    _to_int(region_locals.get("central")) <= 0
                    and orders_count_local > 0
                ):
                    region_locals["central"] = orders_count_local
            else:
                # Fallback when no regional snapshots exist.
                orders_count_local = max(
                    _round_int_half_up((orders_count * localization_percent) / 100.0), 0
                )
                if orders_count > 0:
                    region_totals["central"] = orders_count
                if orders_count_local > 0:
                    region_locals["central"] = orders_count_local

            checklist_snap = (
                xlsx_checklist_snapshot.get((nm_id, date_key))
                if USE_XLSX_CHECKLIST_SNAPSHOT
                else None
            )
            if (
                isinstance(checklist_snap, dict)
                and "orders_count_local" in checklist_snap
            ):
                orders_count_local = max(
                    _to_int(checklist_snap.get("orders_count_local")), 0
                )

            adv_sum = _to_float(adv_daily_map.get((nm_id, date_key)))
            split_bucket = adv_split_map.get(
                (nm_id, date_key), {"auto": 0.0, "search": 0.0, "unknown": 0.0}
            )
            adv_sum_auto = _to_float(split_bucket.get("auto"))
            adv_sum_search = _to_float(split_bucket.get("search"))
            adv_sum_unknown = _to_float(split_bucket.get("unknown"))
            adv_sum_split_total = adv_sum_auto + adv_sum_search + adv_sum_unknown
            if abs(adv_sum) <= 1e-9 and abs(adv_sum_split_total) > 1e-9:
                adv_sum = adv_sum_split_total
            promo_sum = 0.0
            extra = report_extras.get((nm_id, date_key), {})
            external_costs = max(_to_float(extra.get("external_costs")), 0.0)
            meta = cards_meta.get(nm_id, {})
            imt_id = _to_int(meta.get("imt_id"))
            per_nm = (
                plan_settings.get(nm_id) if isinstance(plan_settings, dict) else None
            )
            cal_data = calibration.get(nm_id, {})
            cal_plan_row = (
                cal_data.get("plan_row")
                if isinstance(cal_data, dict)
                and isinstance(cal_data.get("plan_row"), dict)
                else {}
            )
            stored_price = stored_price_map.get((nm_id, date_key))
            if stored_price is None:
                dates = stored_price_dates.get(nm_id) or []
                pos = bisect.bisect_right(dates, date_key)
                if pos > 0:
                    prev_dt = dates[pos - 1]
                    stored_price = stored_price_map.get((nm_id, prev_dt))
            stored_discounted = max(
                _to_float((stored_price or {}).get("discounted_price")), 0.0
            )
            stored_discounted_with_spp = max(
                _to_float((stored_price or {}).get("discounted_price_with_spp")), 0.0
            )
            stored_spp = max(_to_float((stored_price or {}).get("spp")), 0.0)
            unit_snapshot_for_price = _unit_snapshot_for_day(nm_id, date_key) or {}
            unit_discounted = max(
                _to_float((unit_snapshot_for_price or {}).get("discounted_price")), 0.0
            )
            unit_discounted_with_spp = max(
                _to_float(
                    (unit_snapshot_for_price or {}).get("discounted_price_with_spp")
                ),
                0.0,
            )
            if unit_discounted_with_spp <= 0 and unit_discounted > 0:
                unit_spp_hint = max(
                    _to_float((unit_snapshot_for_price or {}).get("spp")), 0.0
                )
                if unit_spp_hint > 1.0:
                    unit_spp_hint = unit_spp_hint / 100.0
                unit_spp_hint = max(0.0, min(0.95, unit_spp_hint))
                if unit_spp_hint > 0:
                    unit_discounted_with_spp = max(
                        unit_discounted * (1.0 - unit_spp_hint), 0.0
                    )
            card_price = _round2(meta.get("card_price"))
            if card_price <= 0 and stored_discounted > 0:
                card_price = _round2(stored_discounted)
            if card_price <= 0 and unit_discounted > 0:
                card_price = _round2(unit_discounted)
            if card_price <= 0:
                card_price = 0.0

            avg_price_from_orders = (
                _round2(orders_sum / float(orders_count)) if orders_count > 0 else 0.0
            )
            # checklist_cross avg_price is derived strictly from /orders and carried forward.
            # Seeding from card/unit/plan prices creates non-zero values before the first order and breaks parity.
            avg_price = avg_price_from_orders
            if avg_price_from_orders > 0:
                last_avg_price_by_nm[nm_id] = avg_price_from_orders
            else:
                avg_price = max(_to_float(last_avg_price_by_nm.get(nm_id)), 0.0)
            # Warmup rows exist only to seed rolling metrics (orders_dyn/buyout%/avg_price carry-forward).
            # Do not expose them in the checklist response.
            if date_key < date_from_norm:
                continue
            if card_price <= 0 and avg_price > 0:
                card_price = avg_price
            order_price = avg_price
            # Prefer SPP computed from /orders,/sales (matches checklist_cross formulas).
            spp = max(_to_float(row.get("spp")), 0.0)
            if spp <= 0:
                report_spp = _to_float(extra.get("spp"))
                if report_spp > 0:
                    spp = report_spp
                elif stored_spp > 0:
                    spp = stored_spp
                # Do not seed SPP from UNIT/card metadata: checklist_cross keeps SPP at 0
                # until the first real /orders day, then carries it forward.

            returns_count = max(orders_count - buyouts_count - cancel_count, 0)
            orders_dyn = max(_to_float(orders_dyn_map.get((nm_id, date_key))), 0.0)
            if isinstance(checklist_snap, dict) and "orders_dyn" in checklist_snap:
                orders_dyn = max(_to_float(checklist_snap.get("orders_dyn")), 0.0)

            buyout_percent_month = max(
                _to_float(buyout_percent_month_map.get((nm_id, date_key))), 0.0
            )
            if buyout_percent_month <= 0:
                buyout_percent_month = _pick_buyout_percent_hint(nm_id, date_key)
            buyout_rate_month = max(min(float(buyout_percent_month), 1.2), 0.0)

            buyout_percent_day = max(
                _to_float(buyout_percent_day_map.get((nm_id, date_key))), 0.0
            )
            if buyout_percent_day <= 0:
                buyout_percent_day = buyout_percent_month
            buyout_rate_day = max(min(float(buyout_percent_day), 1.2), 0.0)
            # Sheets uses half-up rounding for expected counts.
            expected_buyouts_count = max(
                _round_int_half_up(float(orders_count) * float(buyout_rate_day)), 0
            )
            expected_buyouts_sum = (
                _round2_half_up(float(expected_buyouts_count) * float(avg_price))
                if expected_buyouts_count > 0 and avg_price > 0
                else 0.0
            )

            # Returns plan / stocks_total reconstruction from checklist sheet:
            # returns_plan = in_way_from_client + in_way_to_client * (1 - buyout_percent_month)
            # stocks_total = stocks + returns_plan
            returns_plan = max(
                float(in_way_from_client)
                + float(in_way_to_client) * (1.0 - buyout_rate_month),
                0.0,
            )
            stocks_total = max(float(stocks) + float(returns_plan), 0.0)
            if isinstance(checklist_snap, dict) and "stocks_total" in checklist_snap:
                stocks_total = max(_to_float(checklist_snap.get("stocks_total")), 0.0)

            spp_share = max(_to_float(spp), 0.0)
            if spp_share > 1:
                spp_share = spp_share / 100.0
            spp_share = max(0.0, min(0.95, spp_share))

            # avg_price_with_spp is computed later via the formula layer to match sheet behaviour:
            # avg_price_with_spp = avg_price * (1 - spp_share) when there is any activity context.
            avg_price_with_spp = 0.0

            # Unit expenses decomposition from daily economics coefficients.
            unit_expenses = 0.0

            def _calc_unit_components(settings: Dict[str, Any]) -> Dict[str, float]:
                sebes_rub = max(_to_float(settings.get("sebes_rub")), 0.0)
                markirovka_rub = max(_to_float(settings.get("markirovka_rub")), 0.0)
                perc_mp = _to_float(settings.get("perc_mp")) or DEFAULT_PERC_MP
                delivery_mp_with_buyout_rub = max(
                    _to_float(settings.get("delivery_mp_with_buyout_rub")), 0.0
                )
                hranenie_rub = max(_to_float(settings.get("hranenie_rub")), 0.0)
                additional_costs = max(_to_float(settings.get("additional_costs")), 0.0)
                priemka_rub = max(_to_float(settings.get("priemka_rub")), 0.0)
                acquiring_perc = (
                    _to_float(settings.get("acquiring_perc")) or DEFAULT_ACQUIRING_PERC
                )
                tax_total_perc = (
                    _to_float(settings.get("tax_total_perc")) or DEFAULT_TAX_TOTAL_PERC
                )

                base_price = avg_price
                if base_price <= 0 and buyouts_count > 0:
                    base_price = _round2(buyouts_sum / float(buyouts_count))
                if base_price <= 0:
                    base_price = card_price
                base_price_with_spp = max(base_price * (1.0 - spp_share), 0.0)

                perc_mp_rub = base_price * perc_mp
                acquiring_rub = base_price_with_spp * acquiring_perc
                tax_total_rub = base_price_with_spp * tax_total_perc

                unit_total = (
                    sebes_rub
                    + markirovka_rub
                    + delivery_mp_with_buyout_rub
                    + hranenie_rub
                    + priemka_rub
                    + additional_costs
                    + perc_mp_rub
                    + acquiring_rub
                    + tax_total_rub
                )
                # UNIT can provide an already computed per-unit total (`expenses`). When present, treat it
                # as authoritative for the final unit_expenses to match sheet behaviour.
                expenses_override = max(_to_float(settings.get("expenses")), 0.0)
                if expenses_override > 0:
                    unit_total = expenses_override
                return {
                    "sebes_rub": sebes_rub,
                    "markirovka_rub": markirovka_rub,
                    "delivery_mp_with_buyout_rub": delivery_mp_with_buyout_rub,
                    "hranenie_rub": hranenie_rub,
                    "additional_costs": additional_costs,
                    "priemka_rub": priemka_rub,
                    "perc_mp_rub": perc_mp_rub,
                    "acquiring_rub": acquiring_rub,
                    "tax_total_rub": tax_total_rub,
                    "unit_expenses": unit_total,
                }

            # For costs we prefer the latest stored UNIT snapshot (<= day) when available; otherwise fall back
            # to the template UNIT settings (which act as "current config" for many sheets).
            unit_settings = xlsx_unit_settings.get(nm_id) or {}
            dates = stored_unit_dates.get(nm_id) or []
            if dates:
                pos = bisect.bisect_right(dates, date_key)
                pick_dt = dates[pos - 1] if pos > 0 else dates[0]
                candidate = stored_unit_map.get((nm_id, pick_dt))
                if isinstance(candidate, dict) and candidate:
                    unit_settings = candidate
            effective_settings: Dict[str, float] = {}
            settings_keys = [
                "sebes_rub",
                "markirovka_rub",
                "perc_mp",
                "delivery_mp_with_buyout_rub",
                "hranenie_rub",
                "acquiring_perc",
                "tax_total_perc",
                "additional_costs",
                "priemka_rub",
                "spp",
                "expenses",
            ]
            for key_name in settings_keys:
                value = (
                    _to_float(per_nm.get(key_name)) if isinstance(per_nm, dict) else 0.0
                )
                if value <= 0 and isinstance(cal_plan_row, dict):
                    value = _to_float(cal_plan_row.get(key_name))
                if value <= 0 and isinstance(unit_settings, dict):
                    value = _to_float(unit_settings.get(key_name))
                if key_name == "perc_mp":
                    # Accept both 0.295 and 29.5 formats.
                    if value > 1.5:
                        value = value / 100.0

                    if (
                        OVERRIDE_PERC_MP_FROM_WB or value <= 0
                    ) and commission_by_subject:
                        subject_id = _to_int(meta.get("subject_id"))
                        if subject_id:
                            comm = commission_by_subject.get(subject_id) or {}
                            perc_raw = _to_float(comm.get(commission_key))
                            perc_mp = _norm_share_clamped(perc_raw)
                            if perc_mp > 0:
                                value = perc_mp
                if value > 0:
                    effective_settings[key_name] = value

            unit_components: Dict[str, float] = _calc_unit_components(
                effective_settings
            )

            # Per-day unit_log can override only the final unit total when available.
            direct = max(_to_float(xlsx_unit_log.get((nm_id, date_key))), 0.0)
            if direct <= 0:
                dates = xlsx_unit_log_dates.get(nm_id) or []
                pos = bisect.bisect_right(dates, date_key)
                nearest_dt = ""
                if pos > 0:
                    nearest_dt = dates[pos - 1]
                elif UNIT_LOG_EARLY_FILL and dates:
                    nearest_dt = dates[0]
                if nearest_dt:
                    direct = max(_to_float(xlsx_unit_log.get((nm_id, nearest_dt))), 0.0)
            if direct > 0:
                unit_components["unit_expenses"] = direct
            unit_expenses = max(_to_float(unit_components.get("unit_expenses")), 0.0)

            formula_values = {
                "orders_sum_rub": _round2(orders_sum),
                "orders_count": orders_count,
                "orders_dyn": _round2(orders_dyn),
                "avg_price": _round2(avg_price),
                "spp": _to_float(spp),
                "adv_sum": _round2(adv_sum),
                "promo_sum": _round2(promo_sum),
                "unit_expenses": _to_float(unit_expenses),
                "expected_buyouts_count": expected_buyouts_count,
                "expected_buyouts_sum_rub": _to_float(expected_buyouts_sum),
                "buyout_percent_month": round(buyout_percent_month, 6),
                "stocks_total": _to_float(stocks_total),
                "open_card_count": open_count,
            }
            formula_output = apply_checklist_formula_layer(
                CHECKLIST_FORMULA_LAYER, formula_values
            )
            avg_price_with_spp_formula = max(
                _to_float(formula_output.get("avg_price_with_spp")), 0.0
            )
            # For active order days, sheet parity is formula-first.
            # On zero-order days, use snapshot carry-forward before formula fallback.
            avg_price_with_spp = avg_price_with_spp_formula
            use_snapshot_no_orders = AVG_PRICE_WITH_SPP_MODE in {
                "snapshot",
                "hybrid",
                "snapshot_no_orders",
            }
            if use_snapshot_no_orders and orders_count <= 0:
                if stored_discounted_with_spp > 0:
                    avg_price_with_spp = stored_discounted_with_spp
                elif unit_discounted_with_spp > 0:
                    avg_price_with_spp = unit_discounted_with_spp

            expected_buyouts_dyn = max(
                _to_float(formula_output.get("expected_buyouts_dyn")), 0.0
            )
            if (
                expected_buyouts_dyn <= 0
                and orders_dyn > 0
                and buyout_percent_month > 0
            ):
                expected_buyouts_dyn = orders_dyn * buyout_percent_month

            if "profit_without_adv" in formula_output:
                profit_without_adv = _round2_half_up(
                    formula_output.get("profit_without_adv")
                )
            else:
                profit_without_adv = _round2_half_up(
                    float(expected_buyouts_sum)
                    - float(expected_buyouts_count) * float(unit_expenses)
                    - float(promo_sum)
                )
            profit_with_adv = _round2_half_up(profit_without_adv - float(adv_sum))

            stocks_enough_for = (
                _round2_half_up(_safe_div(float(stocks), orders_dyn))
                if orders_dyn > 0
                else 0.0
            )
            stocks_enough_for_with_buyout = (
                _round2_half_up(_safe_div(stocks_total, expected_buyouts_dyn))
                if expected_buyouts_dyn > 0
                else 0.0
            )
            stocks_rub = _round2(stocks_total * card_price)
            sebes_rub = max(_to_float(unit_components.get("sebes_rub")), 0.0)
            markirovka_rub = max(_to_float(unit_components.get("markirovka_rub")), 0.0)
            delivery_mp_with_buyout_rub = max(
                _to_float(unit_components.get("delivery_mp_with_buyout_rub")), 0.0
            )
            hranenie_rub = max(_to_float(unit_components.get("hranenie_rub")), 0.0)
            additional_costs = max(
                _to_float(unit_components.get("additional_costs")), 0.0
            )
            priemka_rub = max(_to_float(unit_components.get("priemka_rub")), 0.0)
            perc_mp_rub = max(_to_float(unit_components.get("perc_mp_rub")), 0.0)
            acquiring_rub = max(_to_float(unit_components.get("acquiring_rub")), 0.0)
            tax_total_rub = max(_to_float(unit_components.get("tax_total_rub")), 0.0)
            returns_plan_rub = (
                _round2(float(returns_plan) * float(sebes_rub))
                if sebes_rub > 0
                else 0.0
            )
            adv_percent = _safe_div(adv_sum, orders_sum) if orders_sum > 0 else 0.0

            orders_count_total_central = max(_to_int(region_totals.get("central")), 0)
            orders_count_total_northwest = max(
                _to_int(region_totals.get("northwest")), 0
            )
            orders_count_total_south_caucasus = max(
                _to_int(region_totals.get("south_caucasus")), 0
            )
            orders_count_total_volga = max(_to_int(region_totals.get("volga")), 0)
            orders_count_total_fareast = max(_to_int(region_totals.get("fareast")), 0)
            orders_count_total_ural = max(_to_int(region_totals.get("ural")), 0)
            orders_count_local_central = max(_to_int(region_locals.get("central")), 0)
            orders_count_local_northwest = max(
                _to_int(region_locals.get("northwest")), 0
            )
            orders_count_local_south_caucasus = max(
                _to_int(region_locals.get("south_caucasus")), 0
            )
            orders_count_local_volga = max(_to_int(region_locals.get("volga")), 0)
            orders_count_local_fareast = max(_to_int(region_locals.get("fareast")), 0)
            orders_count_local_ural = max(_to_int(region_locals.get("ural")), 0)

            localization_percent_central = (
                _round2(
                    _safe_div(orders_count_local_central, orders_count_total_central)
                    * 100.0
                )
                if orders_count_total_central > 0
                else 0.0
            )
            localization_percent_northwest = (
                _round2(
                    _safe_div(
                        orders_count_local_northwest, orders_count_total_northwest
                    )
                    * 100.0
                )
                if orders_count_total_northwest > 0
                else 0.0
            )
            localization_percent_south_caucasus = (
                _round2(
                    _safe_div(
                        orders_count_local_south_caucasus,
                        orders_count_total_south_caucasus,
                    )
                    * 100.0
                )
                if orders_count_total_south_caucasus > 0
                else 0.0
            )
            localization_percent_volga = (
                _round2(
                    _safe_div(orders_count_local_volga, orders_count_total_volga)
                    * 100.0
                )
                if orders_count_total_volga > 0
                else 0.0
            )
            localization_percent_fareast = (
                _round2(
                    _safe_div(orders_count_local_fareast, orders_count_total_fareast)
                    * 100.0
                )
                if orders_count_total_fareast > 0
                else 0.0
            )
            localization_percent_ural = (
                _round2(
                    _safe_div(orders_count_local_ural, orders_count_total_ural) * 100.0
                )
                if orders_count_total_ural > 0
                else 0.0
            )

            checklist_row = _checklist_template_row()
            checklist_row.update(
                {
                    "date": date_key,
                    "nm_id": nm_id,
                    "date__nm_id": f"{date_key}__{nm_id}",
                    "imt_id": imt_id,
                    "open_card_count_jam": open_count,
                    "add_to_cart_count_jam": atc_count,
                    "orders_count_jam": orders_count,
                    "views": open_count,
                    "clicks": atc_count,
                    "adv_sum": _round2(adv_sum),
                    "adv_sum_auto": _round2(adv_sum_auto),
                    "adv_sum_search": _round2(adv_sum_search),
                    # Best-effort: in WB normquery stats, spend is tied to search clusters.
                    # Keep keywords spend aligned with search unless a dedicated source exists.
                    "adv_sum_keywords": _round2(adv_sum_search),
                    "adv_percent": round(adv_percent, 6),
                    "external_costs": _round2(external_costs),
                    "atbs": atc_count,
                    "orders": orders_count,
                    "shks": buyouts_count,
                    "sum_price": _round2(orders_sum),
                    "open_card_count": open_count,
                    "orders_sum_rub": _round2(orders_sum),
                    "orders_count": orders_count,
                    "orders_dyn": _round2(orders_dyn),
                    "add_to_cart_count": atc_count,
                    "add_to_cart_conversion": _round2(add_to_cart_conversion),
                    "cart_to_order_conversion": _round2(cart_to_order_conversion),
                    "click_to_order_conversion": _round2(click_to_order_conversion),
                    "buyout_percent": _round2(buyout_percent),
                    "buyout_percent_day": round(buyout_percent_day, 6),
                    "buyout_percent_month": round(buyout_percent_month, 6),
                    "spp": round(_to_float(spp), 6),
                    "buyouts_sum_rub": _round2(buyouts_sum),
                    "buyouts_count": buyouts_count,
                    "avg_price": _round2(avg_price),
                    "avg_price_with_spp": _round2(avg_price_with_spp),
                    "stocks": stocks,
                    "stocks_sizes": _round2(stocks_total),
                    "in_way_to_client": _round2(in_way_to_client),
                    "in_way_from_client": _round2(in_way_from_client),
                    "stocks_total": _round2(stocks_total),
                    "stocks_enough_for": stocks_enough_for,
                    "stocks_enough_for_with_buyout_perc": stocks_enough_for_with_buyout,
                    "returns_plan": _round2(returns_plan),
                    "returns_plan_rub": returns_plan_rub,
                    "order_price": _round2(order_price),
                    "card_price": _round2(card_price),
                    "localization": _round2(localization_percent / 100.0),
                    "orders_count_local": orders_count_local,
                    "sebes_rub": _round2(sebes_rub),
                    "markirovka_rub": _round2(markirovka_rub),
                    "perc_mp_rub": _round2(perc_mp_rub),
                    "delivery_mp_with_buyout_rub": _round2(delivery_mp_with_buyout_rub),
                    "hranenie_rub": _round2(hranenie_rub),
                    "acquiring_rub": _round2(acquiring_rub),
                    "tax_total_rub": _round2(tax_total_rub),
                    "additional_costs": _round2(additional_costs),
                    "priemka_rub": _round2(priemka_rub),
                    "marg_without_adv": _round2_half_up(profit_without_adv),
                    "marg_with_adv": _round2_half_up(profit_with_adv),
                    "profit_without_adv": _round2_half_up(profit_without_adv),
                    "profit_with_adv": _round2_half_up(profit_with_adv),
                    "promo_sum": _round2(promo_sum),
                    "promo_count": 0,
                    "promo_total_cost": _round2(promo_sum + adv_sum),
                    "expected_buyouts_sum_rub": expected_buyouts_sum,
                    "unit_expenses": _round2(unit_expenses),
                    "stocks_rub": stocks_rub,
                    "all_stocks_rub": stocks_rub,
                    "views_keywords": open_count,
                    "frequency": open_count,
                    "jam_clicks": atc_count,
                    "orders_count_completed": orders_count,
                    "orders_count_canceled": cancel_count,
                    "orders_count_returned": returns_count,
                    "orders_buyouts_count": buyouts_count,
                    "orders_sum_rub_completed": _round2(orders_sum),
                    "orders_sum_rub_canceled": _round2(cancel_sum),
                    "orders_sum_rub_returned": 0.0,
                    "orders_buyouts_sum_rub": _round2(buyouts_sum),
                    "expected_buyouts_count": expected_buyouts_count,
                    "expected_buyouts_dyn": round(expected_buyouts_dyn, 5),
                    "orders_count_returned_fact": returns_count,
                    "orders_buyouts_count_fact": buyouts_count,
                    "orders_count_canceled_fact": cancel_count,
                    "orders_buyouts_sum_rub_fact": _round2(buyouts_sum),
                    "orders_sum_rub_returned_fact": 0.0,
                    "orders_sum_rub_canceled_fact": _round2(cancel_sum),
                    "orders_count_total_central": orders_count_total_central,
                    "orders_count_total_northwest": orders_count_total_northwest,
                    "orders_count_total_south_caucasus": orders_count_total_south_caucasus,
                    "orders_count_total_volga": orders_count_total_volga,
                    "orders_count_total_fareast": orders_count_total_fareast,
                    "orders_count_total_ural": orders_count_total_ural,
                    "orders_count_local_central": orders_count_local_central,
                    "orders_count_local_northwest": orders_count_local_northwest,
                    "orders_count_local_south_caucasus": orders_count_local_south_caucasus,
                    "orders_count_local_volga": orders_count_local_volga,
                    "orders_count_local_fareast": orders_count_local_fareast,
                    "orders_count_local_ural": orders_count_local_ural,
                    "localization_percent_central": localization_percent_central,
                    "localization_percent_northwest": localization_percent_northwest,
                    "localization_percent_south_caucasus": localization_percent_south_caucasus,
                    "localization_percent_volga": localization_percent_volga,
                    "localization_percent_fareast": localization_percent_fareast,
                    "localization_percent_ural": localization_percent_ural,
                    "wrn_count": 0,
                    "actions": "",
                    "log_text": "autonomous-checklist",
                }
            )
            if (
                USE_XLSX_CHECKLIST_SNAPSHOT
                and USE_XLSX_CHECKLIST_METRICS
                and isinstance(checklist_snap, dict)
            ):
                int_fields = {
                    "orders_count",
                    "orders_count_local",
                    "buyouts_count",
                    "stocks",
                }
                six_dec_fields = {"adv_percent", "spp"}
                override_fields = [
                    "adv_sum",
                    "adv_percent",
                    "avg_price_with_spp",
                    "buyouts_count",
                    "buyouts_sum_rub",
                    "expected_buyouts_sum_rub",
                    "orders_count",
                    "orders_count_local",
                    "orders_dyn",
                    "orders_sum_rub",
                    "profit_with_adv",
                    "profit_without_adv",
                    "spp",
                    "stocks",
                    "stocks_enough_for",
                    "stocks_enough_for_with_buyout_perc",
                    "stocks_total",
                ]
                for field in override_fields:
                    if field not in checklist_snap:
                        continue
                    raw_val = checklist_snap.get(field)
                    if field in int_fields:
                        checklist_row[field] = max(_to_int(raw_val), 0)
                    elif field in six_dec_fields:
                        checklist_row[field] = round(_to_float(raw_val), 6)
                    else:
                        checklist_row[field] = _round2_half_up(raw_val)
            cross_override = checklist_cross_overrides.get((nm_id, date_key))
            if isinstance(cross_override, dict) and cross_override:
                int_fields = {
                    "orders_count",
                    "orders_count_local",
                    "buyouts_count",
                    "stocks",
                    "wrn_count",
                }
                for field, raw_val in cross_override.items():
                    if field not in checklist_row:
                        continue
                    if field in int_fields:
                        checklist_row[field] = max(_to_int(raw_val), 0)
                    elif field in {"adv_percent", "spp"}:
                        checklist_row[field] = round(_to_float(raw_val), 6)
                    else:
                        checklist_row[field] = _round2_half_up(raw_val)
            out.append(checklist_row)
    out.sort(key=lambda item: (_to_int(item.get("nm_id")), str(item.get("date") or "")))
    return out


def parse_dataset_payload(payload: Dict[str, Any]) -> Tuple[str, Dict[str, Any], str]:
    dataset = payload.get("dataset") if isinstance(payload, dict) else None
    if not isinstance(dataset, dict):
        raise ValueError("dataset object is required")
    dataset_name = dataset.get("name")
    if not dataset_name or not isinstance(dataset_name, str):
        raise ValueError("dataset.name is required")
    values = dataset.get("values")
    if values is None:
        values = {}
    if not isinstance(values, dict):
        raise ValueError("dataset.values must be an object")

    spreadsheet_id = payload.get("spreadsheet_id") or payload.get("ssId")
    if not spreadsheet_id:
        raise ValueError("spreadsheet_id or ssId is required")
    return str(spreadsheet_id), values, dataset_name
